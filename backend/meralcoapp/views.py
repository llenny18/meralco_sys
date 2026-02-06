from rest_framework import viewsets, filters, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from django_filters.rest_framework import DjangoFilterBackend
from django.db.models import Count, Sum, Avg, Q, F
from django.db.models.functions import TruncMonth
from django.utils import timezone
from datetime import date, timedelta
from .models import *
from .serializers import *
from rest_framework.permissions import AllowAny
from rest_framework import status, viewsets
from rest_framework.decorators import action, api_view, permission_classes
from rest_framework.response import Response
from rest_framework.permissions import AllowAny, IsAuthenticated
from rest_framework.authtoken.models import Token
from django.contrib.auth import login, logout
from django.utils import timezone
from .models import User, UserSession
from .serializers import (
    LoginSerializer, 
    UserLoginResponseSerializer,
    LogoutSerializer,
    ChangePasswordSerializer,
    RegisterUserSerializer
)


from rest_framework.response import Response
from django.shortcuts import get_object_or_404

from rest_framework.parsers import MultiPartParser, FormParser
from pathlib import Path
from django.http import FileResponse
from rest_framework.decorators import action
from rest_framework.response import Response
from django.conf import settings
from rest_framework import status, viewsets
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import AllowAny, IsAuthenticated
from rest_framework.authtoken.models import Token
from django.contrib.auth import authenticate
from django.utils import timezone
from .models import User, UserSession, UserRole
from .serializers import (
    LoginSerializer, 
    UserLoginResponseSerializer,
    LogoutSerializer,
    ChangePasswordSerializer,
    RegisterUserSerializer
)

from rest_framework.decorators import api_view
from rest_framework.response import Response
from rest_framework import status
from datetime import datetime
from .ml_service import ml_service
from .chatbot_service import chatbot_service
from .serializers import (
    DelayPredictionSerializer,
    PenaltyPredictionSerializer,
    ChatRequestSerializer
)



from rest_framework import status, viewsets
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import AllowAny
from django.db.models import Count, Sum, Avg, Q, F
from django.db.models.functions import TruncMonth
from django.utils import timezone
from datetime import date, timedelta
from .models import *
from .serializers import *


class DashboardViewSet(viewsets.ViewSet):
    """Dashboard analytics and statistics"""
    permission_classes = [AllowAny]



    @action(detail=False, methods=['get'], url_path='upcoming_deadlines')
    def upcoming_deadlines(self, request):
        """
        Get all upcoming deadlines from various sources:
        - Work Orders
        - Projects
        - SLA Tracking
        - Document Compliance
        """
        try:
            today = timezone.now().date()
            # Get deadlines from next 90 days
            end_date = today + timedelta(days=90)
            
            deadlines = []
            
            # 1. Work Order Deadlines
            work_orders = WorkOrder.objects.filter(
                date_received_by_vc__range=[today, end_date]
            ).select_related('vendor', 'supervisor')
            
            for wo in work_orders:
                days_remaining = (wo.date_received_by_vc - today).days if wo.date_received_by_vc else 0
                deadlines.append({
                    'project_code': wo.wo_no,
                    'project_name': wo.description or 'No Description',
                    'deadline_type': 'Work Order Completion',
                    'due_date': wo.date_received_by_vc,
                    'days_remaining': days_remaining,
                    
                    'status': wo.status,
                    'assigned_to': wo.supervisor_full_name if wo.supervisor_full_name else None
                })
            
            # 2. Project Completion Dates
            projects = Project.objects.filter(
                completion_date__range=[today, end_date],
                status__status_name__in=['In Progress', 'Active']
            ).select_related('vendor', 'assigned_engineer')
            
            for proj in projects:
                days_remaining = (proj.completion_date - today).days if proj.completion_date else 0
                deadlines.append({
                    'project_code': proj.project_code,
                    'project_name': proj.project_name,
                    'deadline_type': 'Project Completion',
                    'due_date': proj.completion_date,
                    'days_remaining': days_remaining,
                    'priority': proj.priority,
                    'status': proj.status.status_name if proj.status else 'Unknown',
                    'assigned_to': proj.assigned_engineer.get_full_name() if proj.assigned_engineer else None
                })
            
            # 3. SLA Tracking Deadlines
            sla_items = SLATracking.objects.filter(
                due_date__range=[today, end_date],
                status='Open'
            ).select_related('project', 'sla_rule')
            
            for sla in sla_items:
                days_remaining = (sla.due_date - today).days if sla.due_date else 0
                deadlines.append({
                    'project_code': sla.project.project_code,
                    'project_name': sla.project.project_name,
                    'deadline_type': 'SLA Deadline',
                    'due_date': sla.due_date,
                    'days_remaining': days_remaining,
                    'priority': 'High' if days_remaining <= 3 else 'Medium',
                    'status': sla.status,
                    'assigned_to': None
                })
            
            # 4. Document Compliance Deadlines
            doc_compliance = DocumentCompliance.objects.filter(
                due_date__range=[today, end_date],
                is_submitted=False
            ).select_related('project', 'doc_type')
            
            for doc in doc_compliance:
                days_remaining = (doc.due_date - today).days if doc.due_date else 0
                deadlines.append({
                    'project_code': doc.project.project_code,
                    'project_name': doc.project.project_name,
                    'deadline_type': f'Document: {doc.doc_type.doc_type_name}',
                    'due_date': doc.due_date,
                    'days_remaining': days_remaining,
                    'priority': 'Critical' if days_remaining <= 2 else 'High',
                    'status': 'Pending Submission',
                    'assigned_to': None
                })
            
            # Sort by days remaining (most urgent first)
            deadlines.sort(key=lambda x: x['days_remaining'])
            
            # Serialize and return
            serializer = UpcomingDeadlineSerializer(deadlines, many=True)
            return Response(serializer.data)
            
        except Exception as e:
            return Response(
                {'error': str(e)},
                status=500
            )
            
    @action(detail=False, methods=['get'], url_path='stats')
    def stats(self, request):
        """Get overall dashboard statistics"""
        try:
            stats = {
                'total_projects': Project.objects.count(),
                'active_projects': Project.objects.exclude(
                    status__status_name__in=['Completed', 'Cancelled', 'Billed']
                ).count(),
                'delayed_projects': Project.objects.filter(is_delayed=True).count(),
                'completed_projects': Project.objects.filter(
                    status__status_name='Completed'
                ).count(),
                'total_vendors': Vendor.objects.count(),
                'active_vendors': Vendor.objects.filter(is_active=True).count(),
                'blacklisted_vendors': Vendor.objects.filter(is_blacklisted=True).count(),
                'pending_inspections': QIInspection.objects.filter(is_completed=False).count(),
                'overdue_documents': DocumentCompliance.objects.filter(
                    is_overdue=True, is_submitted=False
                ).count(),
                'sla_breaches': SLATracking.objects.filter(is_breached=True).count(),
                'total_penalties': float(Penalty.objects.aggregate(
                    total=Sum('penalty_amount')
                )['total'] or 0),
                'pending_invoices': Invoice.objects.exclude(payment_status='Paid').count(),
            }
            return Response(stats)
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=False, methods=['get'], url_path='project_status_summary')
    def project_status_summary(self, request):
        """Get project status distribution"""
        try:
            total_projects = Project.objects.count()
            status_summary = ProjectStatus.objects.annotate(
                project_count=Count('projects')
            ).values('status_name', 'project_count', 'status_color').order_by('status_order')
            
            status_list = list(status_summary)
            for item in status_list:
                if total_projects > 0:
                    item['percentage'] = round(
                        (item['project_count'] / total_projects) * 100, 2
                    )
                else:
                    item['percentage'] = 0
            
            return Response(status_list)
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=False, methods=['get'], url_path='vendor_performance')
    def vendor_performance(self, request):
        """Get vendor performance summary"""
        try:
            limit = int(request.query_params.get('limit', 10))
            
            vendors = Vendor.objects.filter(is_active=True).annotate(
                total_projects=Count('projects'),
                delayed_projects=Count('projects', filter=Q(projects__is_delayed=True)),
                total_penalties=Sum('penalties__penalty_amount', 
                                   filter=~Q(penalties__penalty_status='Waived')),
                sla_breaches=Count('projects__sla_tracking', 
                                  filter=Q(projects__sla_tracking__is_breached=True))
            ).values(
                'vendor_id', 'vendor_code', 'vendor_name', 'compliance_score',
                'total_projects', 'delayed_projects', 'total_penalties', 'sla_breaches'
            )[:limit]
            
            vendors_list = list(vendors)
            for vendor in vendors_list:
                if vendor['total_projects'] > 0:
                    on_time = vendor['total_projects'] - vendor['delayed_projects']
                    vendor['on_time_percentage'] = round(
                        (on_time / vendor['total_projects']) * 100, 2
                    )
                else:
                    vendor['on_time_percentage'] = 0
                vendor['total_penalties'] = float(vendor['total_penalties'] or 0)
                vendor['compliance_score'] = float(vendor['compliance_score'] or 0)
            
            return Response(vendors_list)
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=False, methods=['get'], url_path='delay_analysis')
    def delay_analysis(self, request):
        """Get delay factor analysis"""
        try:
            analysis = ProjectDelay.objects.values(
                'factor__factor_name',
                'factor__factor_category'
            ).annotate(
                occurrence_count=Count('delay_id'),
                total_delay_days=Sum('delay_days'),
                avg_delay_days=Avg('delay_days')
            ).order_by('-occurrence_count')[:10]
            
            result = []
            for item in analysis:
                result.append({
                    'factor__factor_name': item['factor__factor_name'],
                    'factor__factor_category': item['factor__factor_category'],
                    'occurrence_count': item['occurrence_count'],
                    'total_delay_days': item['total_delay_days'],
                    'avg_delay_days': float(item['avg_delay_days'] or 0)
                })
            
            return Response(result)
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=False, methods=['get'], url_path='monthly_trends')
    def monthly_trends(self, request):
        """Get monthly project trends"""
        try:
            twelve_months_ago = timezone.now() - timedelta(days=365)
            
            trends = Project.objects.filter(
                created_at__gte=twelve_months_ago
            ).annotate(
                month=TruncMonth('created_at')
            ).values('month').annotate(
                total=Count('project_id'),
                completed=Count('project_id', filter=Q(status__status_name='Completed')),
                delayed=Count('project_id', filter=Q(is_delayed=True))
            ).order_by('month')
            
            return Response(list(trends))
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=False, methods=['get'], url_path='project_priority_distribution')
    def project_priority_distribution(self, request):
        """Get project distribution by priority"""
        try:
            total_projects = Project.objects.count()
            
            distribution = Project.objects.values('priority').annotate(
                count=Count('project_id')
            ).order_by('-count')
            
            result = []
            for item in distribution:
                percentage = (item['count'] / total_projects * 100) if total_projects > 0 else 0
                result.append({
                    'priority': item['priority'],
                    'count': item['count'],
                    'percentage': round(percentage, 2)
                })
            
            return Response(result)
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=False, methods=['get'], url_path='upcoming_deadlines')
    def upcoming_deadlines(self, request):
        """Get upcoming deadlines (next 7 days)"""
        try:
            today = date.today()
            next_week = today + timedelta(days=7)
            
            # SLA deadlines
            sla_deadlines = SLATracking.objects.filter(
                due_date__gte=today,
                due_date__lte=next_week,
                completion_date__isnull=True
            ).select_related('project').values(
                'project__project_code',
                'project__project_name',
                'due_date',
                'project__priority'
            )
            
            result = []
            for sla in sla_deadlines:
                days_remaining = (sla['due_date'] - today).days
                result.append({
                    'project_code': sla['project__project_code'],
                    'project_name': sla['project__project_name'],
                    'deadline_type': 'SLA Deadline',
                    'due_date': sla['due_date'].isoformat(),
                    'days_remaining': days_remaining,
                    'priority': sla['project__priority']
                })
            
            # Document deadlines
            doc_deadlines = DocumentCompliance.objects.filter(
                due_date__gte=today,
                due_date__lte=next_week,
                is_submitted=False
            ).select_related('project').values(
                'project__project_code',
                'project__project_name',
                'due_date',
                'project__priority'
            )
            
            for doc in doc_deadlines:
                days_remaining = (doc['due_date'] - today).days
                result.append({
                    'project_code': doc['project__project_code'],
                    'project_name': doc['project__project_name'],
                    'deadline_type': 'Document Deadline',
                    'due_date': doc['due_date'].isoformat(),
                    'days_remaining': days_remaining,
                    'priority': doc['project__priority']
                })
            
            # Sort by days remaining
            result.sort(key=lambda x: x['days_remaining'])
            
            return Response(result[:10])
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=False, methods=['get'], url_path='financial_overview')
    def financial_overview(self, request):
        """Get financial overview"""
        try:
            overview = {
                'total_contract_value': float(Project.objects.aggregate(
                    total=Sum('contract_value')
                )['total'] or 0),
                'completed_contract_value': float(Project.objects.filter(
                    status__status_name='Completed'
                ).aggregate(total=Sum('contract_value'))['total'] or 0),
                'total_penalties': float(Penalty.objects.exclude(
                    penalty_status='Waived'
                ).aggregate(total=Sum('penalty_amount'))['total'] or 0),
                'total_invoiced': float(Invoice.objects.aggregate(
                    total=Sum('invoice_amount')
                )['total'] or 0),
                'total_paid': float(Invoice.objects.filter(
                    payment_status='Paid'
                ).aggregate(total=Sum('net_amount'))['total'] or 0),
                'outstanding_payments': float(Invoice.objects.exclude(
                    payment_status='Paid'
                ).aggregate(total=Sum('net_amount'))['total'] or 0),
            }
            
            return Response(overview)
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=False, methods=['get'], url_path='sector_summary')
    def sector_summary(self, request):
        """Get project summary by sector"""
        try:
            sector_summary = Sector.objects.annotate(
                total_projects=Count('projects'),
                active_projects=Count('projects', 
                    filter=~Q(projects__status__status_name__in=['Completed', 'Cancelled', 'Billed'])),
                delayed_projects=Count('projects', filter=Q(projects__is_delayed=True))
            ).values(
                'sector_code', 'sector_name', 'total_projects', 
                'active_projects', 'delayed_projects'
            ).order_by('-total_projects')
            
            return Response(list(sector_summary))
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['GET'])
def health_check(request):
    return Response({
        'status': 'healthy',
        'timestamp': datetime.now().isoformat()
    })

@api_view(['POST'])
def predict_delay(request):
    serializer = DelayPredictionSerializer(data=request.data)
    if not serializer.is_valid():
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
    try:
        prediction = ml_service.predict_delay(serializer.validated_data)
        return Response(prediction)
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['POST'])
def predict_penalty(request):
    serializer = PenaltyPredictionSerializer(data=request.data)
    if not serializer.is_valid():
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
    try:
        prediction = ml_service.predict_penalty(
            serializer.validated_data['violation_type'],
            serializer.validated_data['delay_days']
        )
        return Response(prediction)
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

from rest_framework.decorators import api_view, permission_classes
from rest_framework.response import Response
from rest_framework.permissions import AllowAny
from rest_framework import status
from .chatbot_service import chatbot_service
from .serializers import ChatRequestSerializer
import logging

logger = logging.getLogger(__name__)

@api_view(['POST'])
@permission_classes([AllowAny])
def chat(request):
    """
    Chat endpoint for AI assistant
    
    Expected payload:
    {
        "question": "What is the Smart Vendor Monitoring System?"
    }
    
    Returns:
    {
        "question": "What is the Smart Vendor Monitoring System?",
        "answer": "The Smart Vendor Monitoring System is...",
        "confidence": 0.85,
        "matched_question": "What is the Smart Vendor Monitoring System?"
    }
    """
    serializer = ChatRequestSerializer(data=request.data)
    
    if not serializer.is_valid():
        return Response(
            {
                'error': 'Invalid request',
                'details': serializer.errors
            }, 
            status=status.HTTP_400_BAD_REQUEST
        )
    
    try:
        question = serializer.validated_data['question']
        
        # Log the incoming question
        logger.info(f"Chat question received: {question}")
        print(f"\n{'='*80}")
        print(f"📥 INCOMING QUESTION: {question}")
        print(f"{'='*80}")
        
        # Get answer from chatbot service
        answer = chatbot_service.answer(question)
        
        # Log the response
        logger.info(f"Chat answer generated: {answer[:100]}...")
        print(f"📤 RESPONSE: {answer[:200]}...")
        print(f"{'='*80}\n")
        
        return Response({
            'question': question,
            'answer': answer,
            'timestamp': request.META.get('HTTP_DATE', None)
        }, status=status.HTTP_200_OK)
        
    except Exception as e:
        logger.error(f"Chat error: {str(e)}", exc_info=True)
        print(f"❌ ERROR in chat endpoint: {e}")
        
        return Response(
            {
                'error': 'Internal server error',
                'message': 'Sorry, I encountered an error processing your question.',
                'details': str(e) if request.user.is_staff else None
            }, 
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


@api_view(['GET'])
@permission_classes([AllowAny])
def chat_health(request):
    """Check if chatbot is loaded and working"""
    try:
        # Test the chatbot
        test_answer = chatbot_service.answer("test")
        
        return Response({
            'status': 'healthy',
            'chatbot_loaded': True,
            'knowledge_base_size': len(chatbot_service.kb_questions),
            'model': 'all-MiniLM-L6-v2',
            'test_response': test_answer[:100] + "..."
        })
    except Exception as e:
        return Response({
            'status': 'error',
            'chatbot_loaded': False,
            'error': str(e)
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['POST'])
@permission_classes([AllowAny])
def chat_debug(request):
    """
    Debug endpoint to see similarity scores for a question
    Only use in development!
    """
    question = request.data.get('question', '')
    
    if not question:
        return Response({'error': 'question required'}, status=400)
    
    try:
        # Get similar questions
        similar = chatbot_service.get_similar_questions(question, top_k=10)
        
        return Response({
            'question': question,
            'similar_questions': similar,
            'threshold': 0.35
        })
    except Exception as e:
        return Response({'error': str(e)}, status=500)

def get_client_ip(request):
    """Get client IP address from request"""
    x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
    if x_forwarded_for:
        ip = x_forwarded_for.split(',')[0]
    else:
        ip = request.META.get('REMOTE_ADDR')
    return ip


class AuthViewSet(viewsets.ViewSet):
    """
    Authentication ViewSet for login, logout, and password management
    """
    
    @action(detail=False, methods=['post'], permission_classes=[AllowAny])
    def login(self, request):
        """
        User login endpoint
        
        Expected payload:
        {
            "username": "admin",
            "password": "admin123",
            "user_type": "admin"
        }
        """
        serializer = LoginSerializer(data=request.data)
        
        if serializer.is_valid():
            user = serializer.validated_data['user']
            
            # Update last login - FIX: Use update() to avoid datetime conversion issues
            User.objects.filter(pk=user.pk).update(last_login=timezone.now())
            
            # Refresh user instance to get updated last_login
            user.refresh_from_db()
            
            # Create or get token
            token, created = Token.objects.get_or_create(user=user)
            
            # Create user session
            user_agent = request.META.get('HTTP_USER_AGENT', '')
            ip_address = get_client_ip(request)
            
            try:
                UserSession.objects.create(
                    user=user,
                    ip_address=ip_address,
                    user_agent=user_agent,
                    is_active=True
                )
            except Exception as e:
                # Log error but don't fail login if session creation fails
                print(f"Session creation warning: {e}")
            
            # Get user data
            user_serializer = UserLoginResponseSerializer(user)
            
            # Determine redirect path based on role
            redirect_paths = {
                'System Administrator': '/admin/dashboard',
                'Team Leader': '/leader/dashboard',
                'Sector Manager': '/sector-manager/dashboard',
                'Engineer': '/engineer/dashboard',
                'Vendor Representative': '/vendor/dashboard',
                'Quality Inspector': '/qi/dashboard',
                'Clerk': '/clerk/dashboard',
                'Engineering Aide': '/aide/dashboard',
                'WO Supervisor': '/supervisor/dashboard'
            }
            
            role_name = user.role.role_name if user.role else 'user'
            redirect_path = redirect_paths.get(role_name, '/dashboard')
            
            return Response({
                'success': True,
                'message': 'Login successful',
                'token': token.key,
                'user': user_serializer.data,
                'redirect_path': redirect_path
            }, status=status.HTTP_200_OK)
        
        return Response({
            'success': False,
            'message': 'Invalid credentials or user type mismatch',
            'errors': serializer.errors
        }, status=status.HTTP_400_BAD_REQUEST)
    
    
    @action(detail=False, methods=['post'], permission_classes=[AllowAny])
    def logout(self, request):
        """
        User logout endpoint
        Requires authentication token
        """
        try:
            # Deactivate current session
            ip_address = get_client_ip(request)
            active_sessions = UserSession.objects.filter(
                user=request.user,
                is_active=True,
                ip_address=ip_address
            )
            
            for session in active_sessions:
                session.is_active = False
                session.logout_time = timezone.now()
                session.save()
            
            # Delete token
            request.user.auth_token.delete()
            
            return Response({
                'success': True,
                'message': 'Logout successful'
            }, status=status.HTTP_200_OK)
        
        except Exception as e:
            return Response({
                'success': False,
                'message': 'Logout failed',
                'error': str(e)
            }, status=status.HTTP_400_BAD_REQUEST)
    
    
    @action(detail=False, methods=['post'], permission_classes=[AllowAny])
    def change_password(self, request):
        """
        Change user password endpoint
        Requires authentication token
        
        Expected payload:
        {
            "old_password": "current_password",
            "new_password": "new_password",
            "confirm_password": "new_password"
        }
        """
        serializer = ChangePasswordSerializer(data=request.data)
        
        if serializer.is_valid():
            user = request.user
            
            # Check old password
            if not user.check_password(serializer.validated_data['old_password']):
                return Response({
                    'success': False,
                    'message': 'Old password is incorrect'
                }, status=status.HTTP_400_BAD_REQUEST)
            
            # Set new password
            user.set_password(serializer.validated_data['new_password'])
            user.save()
            
            # Update token
            Token.objects.filter(user=user).delete()
            token = Token.objects.create(user=user)
            
            return Response({
                'success': True,
                'message': 'Password changed successfully',
                'token': token.key
            }, status=status.HTTP_200_OK)
        
        return Response({
            'success': False,
            'message': 'Password change failed',
            'errors': serializer.errors
        }, status=status.HTTP_400_BAD_REQUEST)
    
    
    @action(detail=False, methods=['get'], permission_classes=[AllowAny])
    def me(self, request):
        """
        Get current authenticated user details
        """
        serializer = UserLoginResponseSerializer(request.user)
        return Response({
            'success': True,
            'user': serializer.data
        }, status=status.HTTP_200_OK)
    
    
    @action(detail=False, methods=['post'], permission_classes=[AllowAny])
    def register(self, request):
        """
        User registration endpoint (optional - can be restricted)
        
        Expected payload:
        {
            "username": "newuser",
            "email": "user@example.com",
            "first_name": "John",
            "last_name": "Doe",
            "password": "password123",
            "confirm_password": "password123",
            "role_name": "engineer",
            "phone_number": "+1234567890"
        }
        """
        serializer = RegisterUserSerializer(data=request.data)
        
        if serializer.is_valid():
            user = serializer.save()
            token = Token.objects.create(user=user)
            user_data = UserLoginResponseSerializer(user).data
            
            return Response({
                'success': True,
                'message': 'User registered successfully',
                'token': token.key,
                'user': user_data
            }, status=status.HTTP_201_CREATED)
        
        return Response({
            'success': False,
            'message': 'Registration failed',
            'errors': serializer.errors
        }, status=status.HTTP_400_BAD_REQUEST)


# Standalone view for getting user roles
from rest_framework.decorators import api_view, permission_classes

@api_view(['GET'])
@permission_classes([AllowAny])
def get_user_roles(request):
    """
    Get list of available user roles for the login dropdown
    """
    roles = UserRole.objects.values('role_id', 'role_name', 'role_description')
    
    # Map to frontend format
    user_types = [
        {
            'value': role['role_name'],
            'label': role['role_name'].replace('-', ' ').title()
        }
        for role in roles
    ]
    
    return Response({
        'success': True,
        'user_types': user_types
    }, status=status.HTTP_200_OK)


# ============================================
# USER MANAGEMENT VIEWSETS
# ============================================

class UserRoleViewSet(viewsets.ModelViewSet):
    queryset = UserRole.objects.all()
    serializer_class = UserRoleSerializer
    permission_classes = [AllowAny]
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['role_name', 'role_description']
    ordering_fields = ['role_name', 'created_at']


class PermissionViewSet(viewsets.ModelViewSet):
    queryset = Permission.objects.all()
    serializer_class = PermissionSerializer
    permission_classes = [AllowAny]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    filterset_fields = ['module_name']
    search_fields = ['permission_name', 'permission_description']


class RolePermissionViewSet(viewsets.ModelViewSet):
    queryset = RolePermission.objects.select_related('role', 'permission').all()
    serializer_class = RolePermissionSerializer
    permission_classes = [AllowAny]
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['role', 'permission']


class UserViewSet(viewsets.ModelViewSet):
    queryset = User.objects.select_related('role').all()
    serializer_class = UserSerializer
    permission_classes = [AllowAny]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['role', 'is_active', 'user_id']
    search_fields = ['username', 'email', 'first_name', 'last_name']
    ordering_fields = ['username', 'created_at', 'last_login']

    @action(detail=False, methods=['get'])
    def me(self, request):
        """Get current user details"""
        serializer = self.get_serializer(request.user)
        return Response(serializer.data)

    @action(detail=True, methods=['post'])
    def deactivate(self, request, pk=None):
        """Deactivate a user"""
        user = self.get_object()
        user.is_active = False
        user.save()
        return Response({'status': 'user deactivated'})


class UserSessionViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = UserSession.objects.select_related('user').all()
    serializer_class = UserSessionSerializer
    permission_classes = [AllowAny]
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_fields = ['user', 'is_active']
    ordering_fields = ['login_time']


# ============================================
# VENDOR MANAGEMENT VIEWSETS
# ============================================

class VendorViewSet(viewsets.ModelViewSet):
    queryset = Vendor.objects.prefetch_related('contacts').all()
    permission_classes = [AllowAny]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['is_active', 'is_blacklisted', 'city', 'region', 'user_id']
    search_fields = ['vendor_code', 'vendor_name', 'company_name', 'email']
    ordering_fields = ['vendor_name', 'compliance_score', 'created_at']

    def get_serializer_class(self):
        if self.action == 'list':
            return VendorListSerializer
        return VendorSerializer

    @action(detail=True, methods=['post'])
    def blacklist(self, request, pk=None):
        """Blacklist a vendor"""
        vendor = self.get_object()
        vendor.is_blacklisted = True
        vendor.blacklist_reason = request.data.get('reason', '')
        vendor.blacklist_date = timezone.now()
        vendor.save()
        return Response({'status': 'vendor blacklisted'})

    @action(detail=True, methods=['post'])
    def remove_blacklist(self, request, pk=None):
        """Remove vendor from blacklist"""
        vendor = self.get_object()
        vendor.is_blacklisted = False
        vendor.blacklist_reason = None
        vendor.blacklist_date = None
        vendor.save()
        return Response({'status': 'vendor removed from blacklist'})

    @action(detail=True, methods=['get'])
    def performance_history(self, request, pk=None):
        """Get vendor performance history"""
        vendor = self.get_object()
        performance = vendor.performance_records.all()
        serializer = VendorPerformanceSerializer(performance, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=['get'])
    def top_performers(self, request):
        """Get top performing vendors"""
        limit = int(request.query_params.get('limit', 10))
        vendors = Vendor.objects.filter(is_active=True, is_blacklisted=False).order_by('-compliance_score')[:limit]
        serializer = VendorListSerializer(vendors, many=True)
        return Response(serializer.data)


class VendorContactViewSet(viewsets.ModelViewSet):
    queryset = VendorContact.objects.select_related('vendor').all()
    serializer_class = VendorContactSerializer
    permission_classes = [AllowAny]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    filterset_fields = ['vendor', 'is_primary']
    search_fields = ['contact_name', 'contact_email', 'contact_phone']


class VendorPerformanceViewSet(viewsets.ModelViewSet):
    queryset = VendorPerformance.objects.select_related('vendor', 'evaluator').all()
    serializer_class = VendorPerformanceSerializer
    permission_classes = [AllowAny]
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_fields = ['vendor', 'evaluator']
    ordering_fields = ['evaluation_date', 'overall_rating']


# ============================================
# PROJECT MANAGEMENT VIEWSETS
# ============================================

class SectorViewSet(viewsets.ModelViewSet):
    queryset = Sector.objects.select_related('sector_manager').all()
    serializer_class = SectorSerializer
    permission_classes = [AllowAny]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    filterset_fields = ['is_active', 'sector_manager']
    search_fields = ['sector_code', 'sector_name', 'location']


class ProjectStatusViewSet(viewsets.ModelViewSet):
    queryset = ProjectStatus.objects.all()
    serializer_class = ProjectStatusSerializer
    permission_classes = [AllowAny]
    filter_backends = [filters.OrderingFilter]
    ordering_fields = ['status_order', 'status_name']


class ProjectViewSet(viewsets.ModelViewSet):
    queryset = Project.objects.select_related(
        'vendor', 'sector', 'status', 'assigned_engineer', 'assigned_qi', 'wo_supervisor'
    ).prefetch_related('milestones', 'team_members').all()
    permission_classes = [AllowAny]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['vendor', 'sector', 'status', 'priority', 'risk_score', 
                        'is_delayed', 'assigned_engineer', 'assigned_qi', 'project_id']
    search_fields = ['project_code', 'project_name', 'project_location']
    ordering_fields = ['project_code', 'start_date', 'completion_date', 'created_at']

    def get_serializer_class(self):
        if self.action == 'list':
            return ProjectListSerializer
        return ProjectSerializer

    @action(detail=False, methods=['get'])
    def delayed(self, request):
        """Get all delayed projects"""
        projects = self.queryset.filter(is_delayed=True)
        serializer = ProjectListSerializer(projects, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=['get'])
    def critical(self, request):
        """Get critical priority projects"""
        projects = self.queryset.filter(priority='Critical')
        serializer = ProjectListSerializer(projects, many=True)
        return Response(serializer.data)

    @action(detail=True, methods=['get'])
    def timeline(self, request, pk=None):
        """Get project workflow timeline"""
        project = self.get_object()
        workflow = project.workflow_stages.select_related('stage', 'assigned_user').all()
        serializer = ProjectWorkflowSerializer(workflow, many=True)
        return Response(serializer.data)

    @action(detail=True, methods=['get'])
    def documents_status(self, request, pk=None):
        """Get project document compliance status"""
        project = self.get_object()
        compliance = project.document_compliance.select_related('doc_type').all()
        serializer = DocumentComplianceSerializer(compliance, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=['get'])
    def by_status(self, request):
        """Get projects grouped by status"""
        status_summary = Project.objects.values(
            'status__status_name', 'status__status_color'
        ).annotate(
            count=Count('id')
        ).order_by('status__status_order')
        return Response(status_summary)


class ProjectMilestoneViewSet(viewsets.ModelViewSet):
    queryset = ProjectMilestone.objects.select_related('project').all()
    serializer_class = ProjectMilestoneSerializer
    permission_classes = [AllowAny]
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_fields = ['project', 'is_completed']
    ordering_fields = ['milestone_order', 'target_date']


class ProjectTeamViewSet(viewsets.ModelViewSet):
    queryset = ProjectTeam.objects.select_related('project', 'user').all()
    serializer_class = ProjectTeamSerializer
    permission_classes = [AllowAny]
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['project', 'user', 'is_active']


# ============================================
# WORKFLOW MANAGEMENT VIEWSETS
# ============================================

class WorkflowStageViewSet(viewsets.ModelViewSet):
    queryset = WorkflowStage.objects.all()
    serializer_class = WorkflowStageSerializer
    permission_classes = [AllowAny]
    filter_backends = [filters.OrderingFilter]
    ordering_fields = ['stage_order', 'stage_name']


class ProjectWorkflowViewSet(viewsets.ModelViewSet):
    queryset = ProjectWorkflow.objects.select_related(
        'project', 'stage', 'assigned_user'
    ).all()
    serializer_class = ProjectWorkflowSerializer
    permission_classes = [AllowAny]
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_fields = ['project', 'stage', 'status', 'is_current_stage', 'assigned_user']
    ordering_fields = ['start_date', 'due_date']

    @action(detail=False, methods=['get'])
    def overdue(self, request):
        """Get overdue workflow stages"""
        overdue = self.queryset.filter(
            completion_date__isnull=True,
            due_date__lt=date.today()
        )
        serializer = self.get_serializer(overdue, many=True)
        return Response(serializer.data)


# ============================================
# DOCUMENT MANAGEMENT VIEWSETS
# ============================================

class DocumentTypeViewSet(viewsets.ModelViewSet):
    queryset = DocumentType.objects.all()
    serializer_class = DocumentTypeSerializer
    permission_classes = [AllowAny]
    filter_backends = [filters.SearchFilter]
    search_fields = ['doc_type_name', 'doc_type_description']


class ProjectDocumentViewSet(viewsets.ModelViewSet):
    queryset = ProjectDocument.objects.select_related(
        'project', 'doc_type', 'uploaded_by', 'approved_by'
    ).all()
    serializer_class = ProjectDocumentSerializer
    permission_classes = [AllowAny]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['project', 'doc_type', 'approval_status', 'is_current_version', 'uploaded_by']
    search_fields = ['document_name']
    ordering_fields = ['upload_date', 'document_name']

    @action(detail=True, methods=['post'])
    def approve(self, request, pk=None):
        """Approve a document"""
        document = self.get_object()
        document.approval_status = 'Approved'
        document.approved_by = request.user
        document.approval_date = timezone.now()
        document.save()
        
        # Email sent automatically via signal
        return Response({'status': 'document approved'})

    @action(detail=True, methods=['post'])
    def reject(self, request, pk=None):
        """Reject a document"""
        document = self.get_object()
        document.approval_status = 'Rejected'
        document.rejection_reason = request.data.get('reason', '')
        document.save()
    
        # Email sent automatically via signal
        return Response({'status': 'document rejected'})

    @action(detail=False, methods=['get'])
    def pending_approval(self, request):
        """Get documents pending approval"""
        pending = self.queryset.filter(approval_status='Pending')
        serializer = self.get_serializer(pending, many=True)
        return Response(serializer.data)


class DocumentComplianceViewSet(viewsets.ModelViewSet):
    queryset = DocumentCompliance.objects.select_related('project', 'doc_type').all()
    serializer_class = DocumentComplianceSerializer
    permission_classes = [AllowAny]
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['project', 'doc_type', 'is_submitted', 'is_approved', 'is_overdue']

    @action(detail=False, methods=['get'])
    def overdue(self, request):
        """Get overdue documents"""
        overdue = self.queryset.filter(is_overdue=True, is_submitted=False)
        serializer = self.get_serializer(overdue, many=True)
        return Response(serializer.data)


# ============================================
# SLA MANAGEMENT VIEWSETS
# ============================================

class SLARuleViewSet(viewsets.ModelViewSet):
    queryset = SLARule.objects.select_related('stage').all()
    serializer_class = SLARuleSerializer
    permission_classes = [AllowAny]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    filterset_fields = ['stage', 'is_active']
    search_fields = ['rule_name', 'rule_description']


class SLATrackingViewSet(viewsets.ModelViewSet):
    queryset = SLATracking.objects.all()
    serializer_class = SLATrackingSerializer
    permission_classes = [AllowAny]
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_fields = ['project', 'sla_rule', 'status', 'is_breached']
    ordering_fields = ['start_date', 'due_date', 'breach_days']

    @action(detail=False, methods=['get'])
    def breached(self, request):
        """Get all SLA breaches"""
        breaches = self.queryset.filter(is_breached=True, status='Breached')
        serializer = self.get_serializer(breaches, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=['get'])
    def at_risk(self, request):
        """Get SLAs at risk of breach"""
        warning_date = date.today() + timedelta(days=2)
        at_risk = self.queryset.filter(
            completion_date__isnull=True,
            due_date__lte=warning_date,
            is_breached=False
        )
        serializer = self.get_serializer(at_risk, many=True)
        return Response(serializer.data)

    @action(detail=True, methods=['post'])
    def waive(self, request, pk=None):
        """Waive an SLA breach"""
        sla = self.get_object()
        sla.status = 'Waived'
        sla.waiver_reason = request.data.get('reason', '')
        sla.waived_by = request.user
        sla.waiver_date = timezone.now()
        sla.save()
        return Response({'status': 'SLA breach waived'})


# ============================================
# QUALITY INSPECTION VIEWSETS
# ============================================

class InspectionTypeViewSet(viewsets.ModelViewSet):
    queryset = InspectionType.objects.all()
    serializer_class = InspectionTypeSerializer
    permission_classes = [AllowAny]


class QIInspectionViewSet(viewsets.ModelViewSet):
    queryset = QIInspection.objects.select_related(
        'project', 'inspection_type', 'assigned_qi'
    ).all()
    serializer_class = QIInspectionSerializer
    permission_classes = [AllowAny]
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_fields = ['project', 'inspection_type', 'assigned_qi', 'inspection_result', 'is_completed']
    ordering_fields = ['scheduled_date', 'inspection_date']
    
        
     
    @action(detail=True, methods=['post'], url_path='submit_corrections')
    def submit_corrections(self, request, pk=None):
        """
        Handle vendor correction submissions with file uploads
        """
        try:
            inspection = self.get_object()
            
            # Check if already submitted
            if inspection.correction_status in ['SUBMITTED', 'APPROVED']:
                return Response({
                    'error': 'Corrections already submitted for this inspection',
                    'status': inspection.correction_status
                }, status=status.HTTP_400_BAD_REQUEST)
            
            # Get correction notes
            correction_notes = request.data.get('correction_notes', '')
            
            if not correction_notes:
                return Response(
                    {'error': 'Correction notes are required'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Handle uploaded photos
            uploaded_files = request.FILES.getlist('corrective_photos')
            
            if not uploaded_files:
                return Response(
                    {'error': 'At least one correction photo is required'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            photo_urls = []
            
            for idx, photo_file in enumerate(uploaded_files):
                # Validate file type
                if not photo_file.content_type.startswith('image/'):
                    return Response(
                        {'error': f'Invalid file type: {photo_file.name}'},
                        status=status.HTTP_400_BAD_REQUEST
                    )
                
                # Validate file size (5MB max)
                if photo_file.size > 5 * 1024 * 1024:
                    return Response(
                        {'error': f'File too large: {photo_file.name}. Max 5MB'},
                        status=status.HTTP_400_BAD_REQUEST
                    )
                
                # Create photo record
                correction_photo = QIInspectionCorrectionPhoto.objects.create(
                    inspection=inspection,
                    photo_file=photo_file,
                    caption=f"Correction photo {idx + 1}",
                    uploaded_by_id=request.data.get('uploaded_by')
                )
                
                # ✅ FIX: Build full URL properly
                full_url = request.build_absolute_uri(correction_photo.photo_file.url)
                photo_urls.append(full_url)
            
            # Update inspection with correction data
            inspection.correction_notes = correction_notes
            inspection.correction_completed_at = timezone.now()
            inspection.correction_status = 'SUBMITTED'
            inspection.correction_photos = photo_urls  # Store full URLs
            inspection.save()
            
            # Send notification email
            try:
                from .email_notification_service import email_service
                email_service.notify_correction_submitted(inspection)
            except Exception as e:
                print(f"Failed to send notification: {e}")
            
            return Response({
                'status': 'success',
                'message': 'Corrections submitted successfully',
                'inspection_id': inspection.inspection_id,
                'correction_status': inspection.correction_status,
                'photo_count': len(photo_urls),
                'photo_urls': photo_urls
            }, status=status.HTTP_200_OK)
            
        except Exception as e:
            import traceback
            traceback.print_exc()
            return Response({
                'status': 'error',
                'message': str(e)
            }, status=status.HTTP_400_BAD_REQUEST)
    
    @action(detail=True, methods=['post'])
    def complete_with_checklist(self, request, pk=None):
        """
        Complete inspection and auto-generate defect flags
        
        Request body:
        {
            "checklist_items": [
                {
                    "item_name": "Electrical wiring",
                    "item_category": "Safety",
                    "status": "Fail",
                    "notes": "Exposed wiring found",
                    "photos": ["photo1.jpg", "photo2.jpg"]
                }
            ],
            "inspection_result": "Fail",
            "findings": "Multiple safety violations",
            "location_coordinates": "14.5995,120.9842"
        }
        """
        inspection = self.get_object()
        
        # Update inspection
        inspection.is_completed = True
        inspection.inspection_date = timezone.now().date()
        inspection.inspection_result = request.data.get('inspection_result', 'Fail')
        inspection.findings = request.data.get('findings', '')
        inspection.location_coordinates = request.data.get('location_coordinates', '')
        inspection.save()
        
        # Create checklist items
        checklist_data = request.data.get('checklist_items', [])
        failed_items = []
        
        for item_data in checklist_data:
            item = InspectionChecklistItem.objects.create(
                inspection=inspection,
                item_name=item_data['item_name'],
                item_category=item_data.get('item_category', ''),
                status=item_data['status'],
                notes=item_data.get('notes', ''),
                photos=item_data.get('photos', []),
                checked_at=timezone.now(),
                checked_by=request.user
            )
            
            if item.status == 'Fail':
                failed_items.append(item)
        
        # Auto-generate flag if there are failures
        if failed_items:
            flag = InspectionFlag.objects.create(
                inspection=inspection,
                flag_type='FAILED_ITEMS',
                item_count=len(failed_items),
                requires_action=True,
                status='PENDING_QI_REVIEW'
            )
            
            # Generate AI suggestions
            suggestions = self._generate_ai_suggestions(failed_items)
            flag.ai_suggestions = {'suggestions': suggestions}
            flag.save()
            
            return Response({
                'status': 'completed',
                'flag_generated': True,
                'flag_id': flag.id,
                'failed_items_count': len(failed_items),
                'message': '⚠️ Inspection completed with failures. Flag created for QI review.'
            })
        
        return Response({
            'status': 'completed',
            'flag_generated': False,
            'message': '✅ Inspection completed successfully with no failures.'
        })
    
    def _generate_ai_suggestions(self, failed_items):
        """AI logic to suggest defect groupings"""
        suggestions = []
        
        # Group by category
        categories = {}
        for item in failed_items:
            cat = item.item_category or 'General'
            if cat not in categories:
                categories[cat] = []
            categories[cat].append(item)
        
        # Create suggestions
        for category, items in categories.items():
            # Determine severity
            severity = self._determine_severity(items)
            
            # Generate description
            item_names = [item.item_name for item in items]
            description = self._generate_description(category, item_names)
            
            suggestions.append({
                'suggested_defect_type': f"{category} Non-Compliance",
                'suggested_severity': severity,
                'suggested_description': description,
                'related_item_ids': [item.id for item in items],
                'confidence_score': self._calculate_confidence(items),
                'reasoning': f"Grouped {len(items)} related failures in {category} category"
            })
        
        return suggestions
    
    def _determine_severity(self, items):
        """Rule-based severity determination"""
        combined_text = ' '.join([item.item_name + ' ' + item.notes for item in items]).lower()
        
        critical_keywords = ['safety', 'structural', 'electrical', 'fire', 'collapse', 'hazard']
        major_keywords = ['code', 'violation', 'non-compliance', 'regulation', 'standard']
        
        if any(keyword in combined_text for keyword in critical_keywords):
            return 'CRITICAL'
        elif any(keyword in combined_text for keyword in major_keywords):
            return 'MAJOR'
        else:
            return 'MINOR'
    
    def _generate_description(self, category, item_names):
        """Generate human-readable description"""
        if len(item_names) == 1:
            return f"Issue found: {item_names[0]}"
        elif len(item_names) <= 3:
            return f"Multiple {category.lower()} issues: {', '.join(item_names)}"
        else:
            return f"Multiple {category.lower()} issues including: {', '.join(item_names[:3])} and {len(item_names)-3} more"
    
    def _calculate_confidence(self, items):
        """Calculate AI confidence score"""
        # More items in same category = higher confidence
        if len(items) >= 5:
            return 0.95
        elif len(items) >= 3:
            return 0.85
        elif len(items) == 2:
            return 0.75
        else:
            return 0.65
    
    @action(detail=True, methods=['post'], url_path='archive-documents') 
    def archive_documents(self, request, pk=None):
        """
        Custom action to archive all documents related to an inspection.
        
        POST /api/v1/qi-inspections/{id}/archive-documents/
        
        Body:
        {
            "archived_by": user_id,
            "archive_date": "2025-01-28T10:00:00Z",
            "retention_period": 10  # years
        }
        """
        try:
            inspection = self.get_object()
            
            # Get request data
            archived_by = request.data.get('archived_by')
            archive_date = request.data.get('archive_date', timezone.now().isoformat())
            retention_period = request.data.get('retention_period', 10)
            
            # Calculate retention expiry date
            archive_datetime = timezone.now()
            expiry_date = archive_datetime + timedelta(days=365 * retention_period)
            
            # Get all related documents for this inspection
            archived_documents = []
            
            # 1. Archive inspection photos
            if hasattr(inspection, 'inspection_photos'):
                photos = inspection.inspection_photos.all()
                for photo in photos:
                    photo.is_archived = True
                    photo.archived_at = archive_datetime
                    photo.archived_by_id = archived_by
                    photo.archive_expiry_date = expiry_date
                    photo.save()
                    archived_documents.append({
                        'type': 'photo',
                        'id': photo.id,
                        'filename': photo.photo.name if photo.photo else 'N/A'
                    })
            
            # 2. Archive inspection reports/PDFs
            if hasattr(inspection, 'inspection_reports'):
                reports = inspection.inspection_reports.all()
                for report in reports:
                    report.is_archived = True
                    report.archived_at = archive_datetime
                    report.archived_by_id = archived_by
                    report.archive_expiry_date = expiry_date
                    report.save()
                    archived_documents.append({
                        'type': 'report',
                        'id': report.id,
                        'filename': report.report_file.name if report.report_file else 'N/A'
                    })
            
            # 3. Archive inspection attachments
            if hasattr(inspection, 'attachments'):
                attachments = inspection.attachments.all()
                for attachment in attachments:
                    attachment.is_archived = True
                    attachment.archived_at = archive_datetime
                    attachment.archived_by_id = archived_by
                    attachment.archive_expiry_date = expiry_date
                    attachment.save()
                    archived_documents.append({
                        'type': 'attachment',
                        'id': attachment.id,
                        'filename': attachment.file.name if attachment.file else 'N/A'
                    })
            
            # 4. Update inspection record
            inspection.documents_archived = True
            inspection.documents_archived_at = archive_datetime
            inspection.documents_archived_by_id = archived_by
            inspection.archive_retention_years = retention_period
            inspection.archive_expiry_date = expiry_date
            inspection.save()
            
            # 5. Create archive log entry (optional)
            try:
                ArchiveLog.objects.create(
                    inspection=inspection,
                    archived_by_id=archived_by,
                    archive_date=archive_datetime,
                    retention_period_years=retention_period,
                    expiry_date=expiry_date,
                    document_count=len(archived_documents),
                    document_details=json.dumps(archived_documents)
                )
            except Exception as log_error:
                # Log creation failed but archiving succeeded
                print(f"Archive log creation failed: {log_error}")
            
            return Response({
                'success': True,
                'message': 'Documents archived successfully',
                'inspection_id': inspection.inspection_id,
                'archived_count': len(archived_documents),
                'archived_documents': archived_documents,
                'retention_period_years': retention_period,
                'archive_date': archive_datetime.isoformat(),
                'expiry_date': expiry_date.isoformat()
            }, status=status.HTTP_200_OK)
            
        except QIInspection.DoesNotExist:
            return Response({
                'success': False,
                'error': 'Inspection not found'
            }, status=status.HTTP_404_NOT_FOUND)
            
        except Exception as e:
            return Response({
                'success': False,
                'error': str(e),
                'detail': 'An error occurred while archiving documents'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)



    @action(detail=False, methods=['get'])
    def pending(self, request):
        """Get pending inspections"""
        pending = self.queryset.filter(is_completed=False)
        serializer = self.get_serializer(pending, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=['get'])
    def overdue(self, request):
        """Get overdue inspections"""
        overdue = self.queryset.filter(
            is_completed=False,
            scheduled_date__lt=date.today()
        )
        serializer = self.get_serializer(overdue, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=['get'])
    def my_inspections(self, request):
        """Get inspections assigned to current user"""
        my_inspections = self.queryset.filter(assigned_qi=request.user)
        serializer = self.get_serializer(my_inspections, many=True)
        return Response(serializer.data)


class QIDailyTargetViewSet(viewsets.ModelViewSet):
    queryset = QIDailyTarget.objects.select_related('qi_user').all()
    serializer_class = QIDailyTargetSerializer
    permission_classes = [AllowAny]
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_fields = ['qi_user', 'target_met', 'target_date']
    ordering_fields = ['target_date']

    @action(detail=False, methods=['get'])
    def my_targets(self, request):
        """Get targets for current user"""
        targets = self.queryset.filter(qi_user=request.user)
        serializer = self.get_serializer(targets, many=True)
        return Response(serializer.data)


class QIPerformanceViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = QIPerformance.objects.select_related('qi_user').all()
    serializer_class = QIPerformanceSerializer
    permission_classes = [AllowAny]
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_fields = ['qi_user']
    ordering_fields = ['evaluation_period_end', 'quality_rating']


# ============================================
# PENALTY MANAGEMENT VIEWSETS
# ============================================

class PenaltyRuleViewSet(viewsets.ModelViewSet):
    queryset = PenaltyRule.objects.all()
    serializer_class = PenaltyRuleSerializer
    permission_classes = [AllowAny]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    filterset_fields = ['violation_type', 'is_active']
    search_fields = ['rule_name', 'rule_description']


class PenaltyViewSet(viewsets.ModelViewSet):
    queryset = Penalty.objects.select_related(
        'project', 'vendor', 'penalty_rule', 'created_by', 'approved_by', 'waived_by'
    ).all()
    serializer_class = PenaltySerializer
    permission_classes = [AllowAny]
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_fields = ['project', 'vendor', 'penalty_rule', 'penalty_status']
    ordering_fields = ['violation_date', 'penalty_amount', 'created_at']

    @action(detail=True, methods=['post'])
    def approve(self, request, pk=None):
        """Approve a penalty"""
        penalty = self.get_object()
        penalty.penalty_status = 'Issued'
        penalty.approved_by = request.user
        penalty.approval_date = timezone.now()
        penalty.issue_date = date.today()
        penalty.save()
        
        # Email sent automatically via signal
        return Response({'status': 'penalty approved'})


   

    @action(detail=True, methods=['post'])
    def waive(self, request, pk=None):
        """Waive a penalty"""
        penalty = self.get_object()
        penalty.penalty_status = 'Waived'
        penalty.waiver_reason = request.data.get('reason', '')
        penalty.waived_by = request.user
        penalty.waiver_date = timezone.now()
        penalty.save()
        return Response({'status': 'penalty waived'})

    @action(detail=False, methods=['get'])
    def by_vendor(self, request):
        """Get penalties grouped by vendor"""
        vendor_id = request.query_params.get('vendor_id')
        if vendor_id:
            penalties = self.queryset.filter(vendor_id=vendor_id)
            serializer = self.get_serializer(penalties, many=True)
            return Response(serializer.data)
        return Response({'error': 'vendor_id required'}, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=['get'])
    def summary(self, request):
        """Get penalty summary statistics"""
        summary = self.queryset.aggregate(
            total_penalties=Count('id'),
            total_amount=Sum('penalty_amount'),
            issued=Count('id', filter=Q(penalty_status='Issued')),
            paid=Count('id', filter=Q(penalty_status='Paid')),
            waived=Count('id', filter=Q(penalty_status='Waived')),
            disputed=Count('id', filter=Q(penalty_status='Disputed'))
        )
        return Response(summary)


 # Add scheduled task endpoint for checking overdue documents
@api_view(['POST'])
def check_overdue_documents(request):
    """Check and notify vendors about overdue documents (run daily via cron)"""
    
    from datetime import date
    from .models import DocumentCompliance, Vendor
    from collections import defaultdict
    
    # Get all overdue, unsubmitted documents
    overdue_docs = DocumentCompliance.objects.filter(
        is_overdue=True,
        is_submitted=False
    ).select_related('project', 'project__vendor')
    
    # Group by vendor
    vendor_overdue = defaultdict(int)
    for doc in overdue_docs:
        if doc.project.vendor:
            vendor_overdue[doc.project.vendor] += 1
    
    # Send notifications
    notifications_sent = 0
    for vendor, count in vendor_overdue.items():
        email_service.notify_vendor_document_overdue(vendor, count)
        notifications_sent += 1
    
    return Response({
    'status': 'success',
    'notifications_sent': notifications_sent,
    'vendors_notified': [vendor.vendor_name for vendor in vendor_overdue.keys()]
})


# ============================================
# BILLING MANAGEMENT VIEWSETS
# ============================================
from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import AllowAny
from django.http import HttpResponse
from django.db import transaction
from django.utils import timezone
from django.core.mail import EmailMessage
from django.conf import settings
from datetime import datetime
import os
import tempfile
import traceback
import logging

# ReportLab imports for PDF generation
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.units import inch
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak, Image
from reportlab.pdfgen import canvas
from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT

logger = logging.getLogger(__name__)

from .models import Invoice, Penalty, Project, Vendor, User
from .serializers import InvoiceSerializer


class InvoiceViewSet(viewsets.ModelViewSet):
    """
    ViewSet for Invoice operations with PDF generation and email functionality
    """
    queryset = Invoice.objects.all().select_related('project', 'vendor', 'created_by').order_by('-created_at')
    serializer_class = InvoiceSerializer
    filterset_fields = {
        'payment_status': ['exact'],
        'approved_by': ['isnull'],
    }
    permission_classes = [AllowAny]
    
    def get_queryset(self):
        """Filter invoices based on query parameters"""
        queryset = super().get_queryset()
        
        # Filter by status
        status_param = self.request.query_params.get('status', None)
        if status_param:
            queryset = queryset.filter(payment_status=status_param)
        
        # Filter by vendor
        vendor = self.request.query_params.get('vendor', None)
        if vendor:
            queryset = queryset.filter(vendor_id=vendor)
        
        # Filter by project
        project = self.request.query_params.get('project', None)
        if project:
            queryset = queryset.filter(project_id=project)
        
        # Date range filter
        start_date = self.request.query_params.get('start_date', None)
        end_date = self.request.query_params.get('end_date', None)
        if start_date and end_date:
            queryset = queryset.filter(invoice_date__range=[start_date, end_date])
        
        return queryset
    
    @transaction.atomic
    def create(self, request, *args, **kwargs):
        """Create invoice with automatic calculations and project status update"""
        try:
            data = request.data.copy()
            
            # Set created_by
            if hasattr(request.user, 'user_id'):
                data['created_by'] = request.user.user_id
            
            # Calculate net amount
            invoice_amount = float(data.get('invoice_amount', 0))
            penalty_amount = float(data.get('penalty_amount', 0))
            data['net_amount'] = str(invoice_amount - penalty_amount)
            
            serializer = self.get_serializer(data=data)
            serializer.is_valid(raise_exception=True)
            self.perform_create(serializer)
            
            # Update project status to 8 (Invoiced/Billed)
            project_id = data.get('project')
            if project_id:
                try:
                    project = Project.objects.get(project_id=project_id)
                    project.status_id = 8  # Set to "Invoiced" status
                    project.save()
                    logger.info(f"Updated project {project.project_code} status to Invoiced (8)")
                except Project.DoesNotExist:
                    logger.warning(f"Project {project_id} not found for status update")
            
            headers = self.get_success_headers(serializer.data)
            return Response(serializer.data, status=status.HTTP_201_CREATED, headers=headers)
            
        except Exception as e:
            logger.error(f"Error creating invoice: {e}")
            logger.error(traceback.format_exc())
            return Response(
                {'error': str(e), 'detail': 'Failed to create invoice'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    @transaction.atomic
    def update(self, request, *args, **kwargs):
        """Update invoice with recalculation"""
        try:
            partial = kwargs.pop('partial', False)
            instance = self.get_object()
            data = request.data.copy()
            
            # Recalculate net amount if amounts changed
            if 'invoice_amount' in data or 'penalty_amount' in data:
                invoice_amount = float(data.get('invoice_amount', instance.invoice_amount))
                penalty_amount = float(data.get('penalty_amount', instance.penalty_amount))
                data['net_amount'] = str(invoice_amount - penalty_amount)
            
            serializer = self.get_serializer(instance, data=data, partial=partial)
            serializer.is_valid(raise_exception=True)
            self.perform_update(serializer)
            
            return Response(serializer.data)
            
        except Exception as e:
            logger.error(f"Error updating invoice: {e}")
            logger.error(traceback.format_exc())
            return Response(
                {'error': str(e), 'detail': 'Failed to update invoice'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    def _create_invoice_pdf(self, invoice):
        """
        Create a professional invoice PDF using ReportLab
        """
        try:
            logger.info(f"Creating PDF for invoice {invoice.invoice_number}")
            
            # Get related data
            vendor = invoice.vendor
            project = invoice.project
            
            # Get penalties
            penalties = Penalty.objects.filter(
                project=project,
                penalty_status='Issued'
            ).select_related('penalty_rule')
            
            # Create temporary file
            temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.pdf')
            temp_path = temp_file.name
            temp_file.close()
            
            # Create PDF document
            doc = SimpleDocTemplate(
                temp_path,
                pagesize=letter,
                rightMargin=0.75*inch,
                leftMargin=0.75*inch,
                topMargin=0.75*inch,
                bottomMargin=0.75*inch
            )
            
            # Container for the 'Flowable' objects
            elements = []
            
            # Define styles
            styles = getSampleStyleSheet()
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=24,
                textColor=colors.HexColor('#1976d2'),
                spaceAfter=30,
                alignment=TA_CENTER,
                fontName='Helvetica-Bold'
            )
            
            heading_style = ParagraphStyle(
                'CustomHeading',
                parent=styles['Heading2'],
                fontSize=14,
                textColor=colors.HexColor('#1976d2'),
                spaceAfter=12,
                spaceBefore=12,
                fontName='Helvetica-Bold'
            )
            
            normal_style = styles['Normal']
            
            # Title
            title = Paragraph("INVOICE", title_style)
            elements.append(title)
            elements.append(Spacer(1, 0.2*inch))
            
            # Invoice details table
            invoice_info = [
                ['Invoice Number:', invoice.invoice_number],
                ['Invoice Date:', invoice.invoice_date.strftime('%B %d, %Y')],
                ['Due Date:', invoice.due_date.strftime('%B %d, %Y')],
                ['Status:', invoice.payment_status]
            ]
            
            invoice_table = Table(invoice_info, colWidths=[2*inch, 3*inch])
            invoice_table.setStyle(TableStyle([
                ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
                ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('TEXTCOLOR', (0, 0), (0, -1), colors.HexColor('#555555')),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ]))
            elements.append(invoice_table)
            elements.append(Spacer(1, 0.3*inch))
            
            # Vendor information
            elements.append(Paragraph("Bill To:", heading_style))
            vendor_info = f"""
            <b>{vendor.vendor_name}</b><br/>
            {vendor.company_name or ''}<br/>
            {vendor.address or ''}<br/>
            Tax ID: {vendor.tax_id or 'N/A'}<br/>
            Contact: {vendor.email or 'N/A'}
            """
            elements.append(Paragraph(vendor_info, normal_style))
            elements.append(Spacer(1, 0.2*inch))
            
            # Project information
            elements.append(Paragraph("Project Details:", heading_style))
            project_info = f"""
            <b>Project Code:</b> {project.project_code}<br/>
            <b>Project Name:</b> {project.project_name}
            """
            elements.append(Paragraph(project_info, normal_style))
            elements.append(Spacer(1, 0.3*inch))
            
            # Invoice items table
            elements.append(Paragraph("Invoice Items", heading_style))
            
            items_data = [
                ['Description', 'Quantity', 'Unit Price', 'Amount'],
                ['Project Contract Value', '1', f'₱{float(invoice.invoice_amount):,.2f}', 
                 f'₱{float(invoice.invoice_amount):,.2f}']
            ]
            
            items_table = Table(items_data, colWidths=[3*inch, 1*inch, 1.5*inch, 1.5*inch])
            items_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1976d2')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 11),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 10),
                ('ALIGN', (1, 1), (-1, -1), 'RIGHT'),
                ('ALIGN', (0, 1), (0, -1), 'LEFT'),
            ]))
            elements.append(items_table)
            elements.append(Spacer(1, 0.2*inch))
            
            # Penalties if any
            if penalties.exists():
                elements.append(Paragraph("Applied Penalties", heading_style))
                
                penalty_data = [['Penalty Description', 'Amount']]
                for penalty in penalties:
                    penalty_data.append([
                        penalty.penalty_rule.rule_name,
                        f'₱{float(penalty.penalty_amount):,.2f}'
                    ])
                
                penalty_table = Table(penalty_data, colWidths=[5*inch, 2*inch])
                penalty_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#d32f2f')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 11),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#ffebee')),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                    ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                    ('FONTSIZE', (0, 1), (-1, -1), 10),
                ]))
                elements.append(penalty_table)
                elements.append(Spacer(1, 0.2*inch))
            
            # Totals
            totals_data = [
                ['Subtotal:', f'₱{float(invoice.invoice_amount):,.2f}'],
                ['Penalties:', f'- ₱{float(invoice.penalty_amount):,.2f}'],
                ['', ''],
                ['Net Amount:', f'₱{float(invoice.net_amount):,.2f}']
            ]
            
            totals_table = Table(totals_data, colWidths=[5.5*inch, 1.5*inch])
            totals_table.setStyle(TableStyle([
                ('ALIGN', (0, 0), (-1, -1), 'RIGHT'),
                ('FONTNAME', (0, 0), (0, 2), 'Helvetica-Bold'),
                ('FONTNAME', (0, 3), (0, 3), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 3), (-1, 3), 14),
                ('TEXTCOLOR', (1, 1), (1, 1), colors.HexColor('#d32f2f')),
                ('LINEABOVE', (0, 2), (-1, 2), 1, colors.black),
                ('LINEABOVE', (0, 3), (-1, 3), 2, colors.black),
                ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
                ('FONTSIZE', (1, 0), (1, 2), 10),
            ]))
            elements.append(totals_table)
            elements.append(Spacer(1, 0.3*inch))
            
            # Payment terms
            elements.append(Paragraph("Payment Terms", heading_style))
            payment_info = f"""
            <b>Payment Status:</b> {invoice.payment_status}<br/>
            <b>Due Date:</b> {invoice.due_date.strftime('%B %d, %Y')}
            """
            elements.append(Paragraph(payment_info, normal_style))
            
            # Notes
            if invoice.notes:
                elements.append(Spacer(1, 0.2*inch))
                elements.append(Paragraph("Notes", heading_style))
                elements.append(Paragraph(invoice.notes, normal_style))
            
            elements.append(Spacer(1, 0.4*inch))
            
            # Footer
            footer_style = ParagraphStyle(
                'Footer',
                parent=styles['Normal'],
                fontSize=10,
                textColor=colors.HexColor('#666666'),
                alignment=TA_CENTER,
                fontName='Helvetica-Oblique'
            )
            elements.append(Paragraph("Thank you for your business!", footer_style))
            
            # Build PDF
            doc.build(elements)
            logger.info(f"PDF created successfully at {temp_path}")
            
            return temp_path
            
        except Exception as e:
            logger.error(f"Error in _create_invoice_pdf: {e}")
            logger.error(traceback.format_exc())
            raise Exception(f'PDF creation failed: {str(e)}')
    
    def _create_receipt_pdf(self, invoice):
        """
        Create a professional receipt PDF
        """
        try:
            logger.info(f"Creating receipt PDF for invoice {invoice.invoice_number}")
            
            vendor = invoice.vendor
            
            # Create temporary file
            temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.pdf')
            temp_path = temp_file.name
            temp_file.close()
            
            # Create PDF document
            doc = SimpleDocTemplate(
                temp_path,
                pagesize=letter,
                rightMargin=inch,
                leftMargin=inch,
                topMargin=inch,
                bottomMargin=inch
            )
            
            elements = []
            styles = getSampleStyleSheet()
            
            # Title
            title_style = ParagraphStyle(
                'ReceiptTitle',
                parent=styles['Heading1'],
                fontSize=28,
                textColor=colors.HexColor('#2e7d32'),
                spaceAfter=30,
                alignment=TA_CENTER,
                fontName='Helvetica-Bold'
            )
            
            elements.append(Paragraph("PAYMENT RECEIPT", title_style))
            elements.append(Spacer(1, 0.3*inch))
            
            # Receipt box
            receipt_style = ParagraphStyle(
                'ReceiptInfo',
                parent=styles['Normal'],
                fontSize=12,
                alignment=TA_CENTER,
                spaceAfter=6
            )
            
            receipt_number = f"REC-{invoice.invoice_number}"
            elements.append(Paragraph(f"<b>Receipt Number:</b> {receipt_number}", receipt_style))
            elements.append(Spacer(1, 0.4*inch))
            
            # Receipt details
            normal_style = ParagraphStyle(
                'ReceiptNormal',
                parent=styles['Normal'],
                fontSize=12,
                spaceAfter=10,
                leading=18
            )
            
            receipt_info = f"""
            <b>Received from:</b> {vendor.vendor_name}<br/>
            <b>Company:</b> {vendor.company_name or 'N/A'}<br/>
            <br/>
            <b>Amount Received:</b> <font size="16" color="#2e7d32">₱{float(invoice.net_amount):,.2f}</font><br/>
            <br/>
            <b>Payment Date:</b> {invoice.payment_date.strftime('%B %d, %Y') if invoice.payment_date else 'N/A'}<br/>
            <b>Payment Reference:</b> {invoice.payment_reference or 'N/A'}<br/>
            <b>Invoice Number:</b> {invoice.invoice_number}<br/>
            """
            
            elements.append(Paragraph(receipt_info, normal_style))
            elements.append(Spacer(1, 0.5*inch))
            
            # Signature box
            sig_table = Table([
                ['_' * 40],
                ['Authorized Signature']
            ], colWidths=[4*inch])
            sig_table.setStyle(TableStyle([
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 1), (0, 1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (0, 1), 10),
                ('TOPPADDING', (0, 1), (0, 1), 10),
            ]))
            elements.append(sig_table)
            
            elements.append(Spacer(1, 0.5*inch))
            
            # Footer
            footer_style = ParagraphStyle(
                'ReceiptFooter',
                parent=styles['Normal'],
                fontSize=11,
                textColor=colors.HexColor('#2e7d32'),
                alignment=TA_CENTER,
                fontName='Helvetica-Bold'
            )
            elements.append(Paragraph("Thank you for your payment!", footer_style))
            
            # Build PDF
            doc.build(elements)
            logger.info(f"Receipt PDF created successfully at {temp_path}")
            
            return temp_path
            
        except Exception as e:
            logger.error(f"Error in _create_receipt_pdf: {e}")
            logger.error(traceback.format_exc())
            raise Exception(f'Receipt PDF creation failed: {str(e)}')
    
    @action(detail=True, methods=['post'])
    def generate_document(self, request, pk=None):
        """
        Generate invoice document in PDF format
        """
        logger.info(f"=== generate_document called for invoice {pk} ===")
        
        try:
            # Get invoice
            invoice = self.get_object()
            logger.info(f"Invoice found: {invoice.invoice_number}")
            
            # Create PDF
            logger.info("Creating PDF document")
            pdf_path = self._create_invoice_pdf(invoice)
            logger.info("PDF created successfully")
            
            # Read file content
            with open(pdf_path, 'rb') as f:
                file_content = f.read()
            
            file_size = len(file_content)
            logger.info(f"PDF size: {file_size} bytes")
            
            # Clean up temp file
            try:
                os.unlink(pdf_path)
                logger.info("Temp file cleaned up")
            except Exception as e:
                logger.warning(f"Could not delete temp file: {e}")
            
            # Create response
            response = HttpResponse(
                file_content,
                content_type='application/pdf'
            )
            response['Content-Disposition'] = f'attachment; filename="Invoice_{invoice.invoice_number}.pdf"'
            response['Content-Length'] = file_size
            
            logger.info("=== generate_document completed successfully ===")
            return response
            
        except Invoice.DoesNotExist:
            logger.error(f"Invoice {pk} not found")
            return Response(
                {'error': f'Invoice with ID {pk} not found'},
                status=status.HTTP_404_NOT_FOUND
            )
        except Exception as e:
            logger.error(f"Error in generate_document: {e}")
            logger.error(traceback.format_exc())
            return Response(
                {
                    'error': 'PDF generation failed',
                    'detail': str(e),
                    'type': type(e).__name__
                },
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    @action(detail=True, methods=['post'])
    def generate_receipt(self, request, pk=None):
        """Generate payment receipt in PDF format"""
        logger.info(f"=== generate_receipt called for invoice {pk} ===")
        
        try:
            invoice = self.get_object()
            
            if invoice.payment_status != 'Paid':
                return Response(
                    {'error': 'Cannot generate receipt for unpaid invoice'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Create receipt PDF
            pdf_path = self._create_receipt_pdf(invoice)
            
            # Read file content
            with open(pdf_path, 'rb') as f:
                file_content = f.read()
            
            # Clean up
            os.unlink(pdf_path)
            
            # Create response
            response = HttpResponse(
                file_content,
                content_type='application/pdf'
            )
            response['Content-Disposition'] = f'attachment; filename="Receipt_{invoice.invoice_number}.pdf"'
            
            logger.info("=== generate_receipt completed successfully ===")
            return response
            
        except Exception as e:
            logger.error(f"Error in generate_receipt: {e}")
            logger.error(traceback.format_exc())
            return Response(
                {'error': str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    @action(detail=True, methods=['post'])
    def send_email(self, request, pk=None):
        """Send invoice via email with PDF attachment"""
        logger.info(f"=== send_email called for invoice {pk} ===")
        
        try:
            invoice = self.get_object()
            vendor = invoice.vendor
            
            if not vendor.email:
                return Response(
                    {'error': 'Vendor email not found'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Generate PDF
            pdf_path = self._create_invoice_pdf(invoice)
            
            # Create email
            subject = f'Invoice {invoice.invoice_number} - {invoice.project.project_name}'
            
            message = f"""
Dear {vendor.vendor_name},

Please find attached invoice {invoice.invoice_number} for project {invoice.project.project_name}.

Invoice Details:
- Invoice Number: {invoice.invoice_number}
- Invoice Date: {invoice.invoice_date.strftime('%B %d, %Y')}
- Due Date: {invoice.due_date.strftime('%B %d, %Y')}
- Invoice Amount: ₱{float(invoice.invoice_amount):,.2f}
- Penalty Amount: ₱{float(invoice.penalty_amount):,.2f}
- Net Amount Due: ₱{float(invoice.net_amount):,.2f}

Please process payment by the due date.

Thank you for your business.

Best regards,
Billing Department
            """
            
            email = EmailMessage(
                subject=subject,
                body=message,
                from_email=settings.DEFAULT_FROM_EMAIL,
                to=[vendor.email],
            )
            
            # Attach PDF
            with open(pdf_path, 'rb') as f:
                email.attach(
                    f'Invoice_{invoice.invoice_number}.pdf',
                    f.read(),
                    'application/pdf'
                )
            
            # Send email
            email.send()
            
            # Clean up
            os.unlink(pdf_path)
            
            logger.info(f"Email sent successfully to: {vendor.email}")
            
            return Response({
                'message': f'Invoice sent successfully to {vendor.email}',
                'recipient': vendor.email
            })
            
        except Exception as e:
            logger.error(f"Error in send_email: {e}")
            logger.error(traceback.format_exc())
            return Response(
                {
                    'error': 'Failed to send email',
                    'detail': str(e)
                },
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    @action(detail=True, methods=['post'])
    def approve(self, request, pk=None):
        """Approve invoice"""
        try:
            invoice = self.get_object()
            
            if hasattr(request.user, 'user_id'):
                invoice.approved_by_id = request.user.user_id
                invoice.approval_date = timezone.now()
                invoice.save()
            
            serializer = self.get_serializer(invoice)
            return Response(serializer.data)
            
        except Exception as e:
            logger.error(f"Error in approve: {e}")
            return Response(
                {'error': str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    @action(detail=False, methods=['get'])
    def statistics(self, request):
        """Get invoice statistics"""
        try:
            from django.db.models import Sum, Count, Q
            
            stats = {
                'total_invoices': Invoice.objects.count(),
                'total_amount': float(Invoice.objects.aggregate(Sum('invoice_amount'))['invoice_amount__sum'] or 0),
                'total_penalties': float(Invoice.objects.aggregate(Sum('penalty_amount'))['penalty_amount__sum'] or 0),
                'total_net': float(Invoice.objects.aggregate(Sum('net_amount'))['net_amount__sum'] or 0),
                'by_status': {
                    'unpaid': Invoice.objects.filter(payment_status='Unpaid').count(),
                    'partially_paid': Invoice.objects.filter(payment_status='Partially Paid').count(),
                    'paid': Invoice.objects.filter(payment_status='Paid').count(),
                    'overdue': Invoice.objects.filter(payment_status='Overdue').count(),
                },
                'total_paid': float(Invoice.objects.filter(payment_status='Paid').aggregate(Sum('net_amount'))['net_amount__sum'] or 0),
                'total_pending': float(Invoice.objects.filter(
                    Q(payment_status='Unpaid') | Q(payment_status='Partially Paid')
                ).aggregate(Sum('net_amount'))['net_amount__sum'] or 0),
            }
            
            return Response(stats)
            
        except Exception as e:
            logger.error(f"Error in statistics: {e}")
            return Response(
                {'error': str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
            
            
            
class PaymentViewSet(viewsets.ModelViewSet):
    """
    ViewSet for Payment tracking
    """
    queryset = Payment.objects.all().select_related('invoice', 'processed_by').order_by('-payment_date')
    serializer_class = PaymentSerializer
    permission_classes = [IsAuthenticated]
    
    @transaction.atomic
    def create(self, request, *args, **kwargs):
        """Create payment and update invoice status"""
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        
        payment = serializer.save(processed_by=request.user)
        invoice = payment.invoice
        
        # Calculate total payments
        total_paid = Payment.objects.filter(invoice=invoice).aggregate(
            Sum('payment_amount')
        )['payment_amount__sum'] or 0
        
        # Update invoice status
        net_amount = float(invoice.net_amount)
        if total_paid >= net_amount:
            invoice.payment_status = 'Paid'
            invoice.payment_date = payment.payment_date
        elif total_paid > 0:
            invoice.payment_status = 'Partially Paid'
        
        invoice.save()
        
        return Response(serializer.data, status=status.HTTP_201_CREATED)


# ============================================
# NOTIFICATION MANAGEMENT VIEWSETS
# ============================================

class NotificationTemplateViewSet(viewsets.ModelViewSet):
    queryset = NotificationTemplate.objects.all()
    serializer_class = NotificationTemplateSerializer
    permission_classes = [AllowAny]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    filterset_fields = ['notification_type', 'is_active']
    search_fields = ['template_name', 'template_subject']


class NotificationViewSet(viewsets.ModelViewSet):
    queryset = Notification.objects.select_related(
        'recipient_user', 'related_project'
    ).all()
    serializer_class = NotificationSerializer
    permission_classes = [AllowAny]
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_fields = ['recipient_user', 'notification_type', 'status', 'related_project']
    ordering_fields = ['created_at', 'sent_at']

    @action(detail=False, methods=['get'])
    def my_notifications(self, request):
        """Get notifications for current user"""
        notifications = self.queryset.filter(recipient_user=request.user)
        serializer = self.get_serializer(notifications, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=['get'])
    def unread(self, request):
        """Get unread notifications for current user"""
        unread = self.queryset.filter(
            recipient_user=request.user,
            read_at__isnull=True
        )
        serializer = self.get_serializer(unread, many=True)
        return Response(serializer.data)

    @action(detail=True, methods=['post'])
    def mark_read(self, request, pk=None):
        """Mark notification as read"""
        notification = self.get_object()
        notification.read_at = timezone.now()
        notification.status = 'Read'
        notification.save()
        return Response({'status': 'marked as read'})

    @action(detail=False, methods=['post'])
    def mark_all_read(self, request):
        """Mark all notifications as read for current user"""
        self.queryset.filter(
            recipient_user=request.user,
            read_at__isnull=True
        ).update(read_at=timezone.now(), status='Read')
        return Response({'status': 'all notifications marked as read'})


# ============================================
# ESCALATION MANAGEMENT VIEWSETS
# ============================================

class EscalationRuleViewSet(viewsets.ModelViewSet):
    queryset = EscalationRule.objects.select_related(
        'escalate_to_role', 'notification_template'
    ).all()
    serializer_class = EscalationRuleSerializer
    permission_classes = [AllowAny]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    filterset_fields = ['is_active', 'escalate_to_role']
    search_fields = ['rule_name', 'rule_description']


class EscalationViewSet(viewsets.ModelViewSet):
    queryset = Escalation.objects.select_related(
        'project', 'escalation_rule', 'escalated_from_user', 
        'escalated_to_user', 'resolved_by'
    ).all()
    serializer_class = EscalationSerializer
    permission_classes = [AllowAny]
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_fields = ['project', 'status', 'escalated_to_user']
    ordering_fields = ['escalation_date', 'response_delay_hours']

    @action(detail=False, methods=['get'])
    def my_escalations(self, request):
        """Get escalations assigned to current user"""
        escalations = self.queryset.filter(escalated_to_user=request.user)
        serializer = self.get_serializer(escalations, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=['get'])
    def open(self, request):
        """Get open escalations"""
        open_escalations = self.queryset.filter(status='Open')
        serializer = self.get_serializer(open_escalations, many=True)
        return Response(serializer.data)

    @action(detail=True, methods=['post'])
    def resolve(self, request, pk=None):
        """Resolve an escalation"""
        escalation = self.get_object()
        escalation.status = 'Resolved'
        escalation.resolution = request.data.get('resolution', '')
        escalation.resolved_by = request.user
        escalation.resolution_date = timezone.now()
        escalation.save()
        return Response({'status': 'escalation resolved'})


# ============================================
# ANALYTICS VIEWSETS
# ============================================

class DelayFactorViewSet(viewsets.ModelViewSet):
    queryset = DelayFactor.objects.all()
    serializer_class = DelayFactorSerializer
    permission_classes = [AllowAny]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    filterset_fields = ['factor_category', 'is_active']
    search_fields = ['factor_name', 'factor_description']


class ProjectDelayViewSet(viewsets.ModelViewSet):
    queryset = ProjectDelay.objects.select_related(
        'project', 'factor', 'reported_by'
    ).all()
    serializer_class = ProjectDelaySerializer
    permission_classes = [AllowAny]
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_fields = ['project', 'factor', 'responsible_party']
    ordering_fields = ['delay_start_date', 'delay_days']

    @action(detail=False, methods=['get'])
    def analysis(self, request):
        """Get delay analysis"""
        analysis = self.queryset.values(
            'factor__factor_name',
            'factor__factor_category'
        ).annotate(
            occurrence_count=Count('id'),
            total_delay_days=Sum('delay_days'),
            avg_delay_days=Avg('delay_days')
        ).order_by('-occurrence_count')[:10]
        
        serializer = DelayAnalysisSerializer(analysis, many=True)
        return Response(serializer.data)


# ============================================
# VENDOR PORTAL VIEWSETS
# ============================================

class VendorDisputeViewSet(viewsets.ModelViewSet):
    queryset = VendorDispute.objects.select_related(
        'vendor', 'project', 'related_penalty', 'assigned_to', 'resolved_by'
    ).all()
    serializer_class = VendorDisputeSerializer
    permission_classes = [AllowAny]
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_fields = ['vendor', 'project', 'dispute_status', 'dispute_type']
    ordering_fields = ['submitted_date']

    @action(detail=False, methods=['get'])
    def my_disputes(self, request):
        """Get disputes for current user's vendor"""
        disputes = self.queryset.filter(vendor__email=request.user.email)
        serializer = self.get_serializer(disputes, many=True)
        return Response(serializer.data)

    @action(detail=True, methods=['post'])
    def assign(self, request, pk=None):
        """Assign dispute to a user"""
        dispute = self.get_object()
        user_id = request.data.get('user_id')
        try:
            user = User.objects.get(id=user_id)
            dispute.assigned_to = user
            dispute.save()
            return Response({'status': 'dispute assigned'})
        except User.DoesNotExist:
            return Response({'error': 'user not found'}, status=status.HTTP_404_NOT_FOUND)

    @action(detail=True, methods=['post'])
    def resolve(self, request, pk=None):
        """Resolve a dispute"""
        dispute = self.get_object()
        dispute.dispute_status = 'Resolved'
        dispute.resolution = request.data.get('resolution', '')
        dispute.resolved_by = request.user
        dispute.resolution_date = timezone.now()
        dispute.save()
        return Response({'status': 'dispute resolved'})


class VendorFeedbackViewSet(viewsets.ModelViewSet):
    queryset = VendorFeedback.objects.select_related(
        'vendor', 'reviewed_by'
    ).all()
    serializer_class = VendorFeedbackSerializer
    permission_classes = [AllowAny]
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_fields = ['vendor', 'feedback_type', 'status', 'rating']
    ordering_fields = ['created_at', 'rating']


# ============================================
# AUDIT & CHANGE LOG VIEWSETS
# ============================================

class ChangeLogViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = ChangeLog.objects.select_related('changed_by').all()
    serializer_class = ChangeLogSerializer
    permission_classes = [AllowAny]
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_fields = ['table_name', 'change_type', 'changed_by']
    ordering_fields = ['created_at']


class SystemAuditLogViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = SystemAuditLog.objects.select_related('user').all()
    serializer_class = SystemAuditLogSerializer
    permission_classes = [AllowAny]
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_fields = ['user', 'action_type', 'status', 'entity_type']
    ordering_fields = ['created_at']


# ============================================
# SYSTEM CONFIGURATION VIEWSETS
# ============================================

class SystemSettingViewSet(viewsets.ModelViewSet):
    queryset = SystemSetting.objects.all()
    serializer_class = SystemSettingSerializer
    permission_classes = [AllowAny]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    filterset_fields = ['setting_type', 'is_editable']
    search_fields = ['setting_key', 'setting_description']




from rest_framework import viewsets, status, filters
from rest_framework.decorators import action
from rest_framework.response import Response
from django.db.models import Count, Sum, Avg, Q, F
from django.utils import timezone
from datetime import datetime, timedelta
from django_filters.rest_framework import DjangoFilterBackend

# ============================================
# WORK ORDER VIEWSETS
# ============================================
from rest_framework import viewsets, status, filters
from rest_framework.decorators import action
from rest_framework.response import Response
from django_filters.rest_framework import DjangoFilterBackend
from django.db.models import Count, Avg, Q, Sum, F
from django.db.models.functions import TruncMonth
from datetime import datetime, timedelta
import pandas as pd
from io import BytesIO
from django.http import HttpResponse

class WorkOrderViewSet(viewsets.ModelViewSet):
    """
    Complete ViewSet for Work Order management
    
    Features:
    - Full CRUD operations
    - Excel import/export with improved column mapping
    - Dashboard statistics
    - Timeline tracking
    - Performance metrics
    - Filtering and search
    """
    queryset = WorkOrder.objects.all()
    serializer_class = WorkOrderSerializer
    permission_classes = [AllowAny]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]

    
    # Enhanced filtering options
    filterset_fields = {
        'status': ['exact', 'in'],
        'vendor_id': ['exact'],
        'supervisor_full_name': ['exact', 'in'],
        'vip': ['exact'],
        'municipality': ['exact', 'in'],
        'assigned': ['exact', 'in'],
        'for_ccti_exclusion': ['exact'],
        'for_apt_exclusion': ['exact'],
        'actual_field_status': ['exact', 'in'],
        'date_received_jacket_ps': ['gte', 'lte', 'exact'],
        'date_comp': ['gte', 'lte', 'exact'],
        'days_comp': ['gte', 'lte'],
    }
    
    # Search functionality
    search_fields = [
        'wo_no',
        'description',
        'location',
        'municipality',
        'area_of_responsibility',
        'assigned',
        'supervisor_full_name',
        'vendor_remarks',
        'c1_remarks'
    ]
    
    # Ordering options
    ordering_fields = [
        'date_received_jacket_ps',
        'date_comp',
        'days_comp',
        'created_at',
        'wo_no',
        'status'
    ]
    ordering = ['-date_received_jacket_ps']
    
    def get_serializer_class(self):
        """Return appropriate serializer based on action"""
        if self.action == 'list':
            return WorkOrderListSerializer
        elif self.action in ['create', 'update', 'partial_update']:
            return WorkOrderCreateUpdateSerializer
        elif self.action == 'timeline':
            return WorkOrderTimelineSerializer
        return WorkOrderSerializer
    
    def get_queryset(self):
        """Enhanced queryset with additional filters"""
        queryset = super().get_queryset()
        params = self.request.query_params

        # ----------------------
        # Project ID filter (multiple allowed) - CUSTOM HANDLING
        # ----------------------
        project_ids_list = params.getlist('project_id')
        print(f"🔢 Raw project_ids_list: {project_ids_list}")
        
        project_ids = []
        
        if project_ids_list:
            for pid in project_ids_list:
                if ',' in str(pid):
                    project_ids.extend([p.strip() for p in str(pid).split(',')])
                else:
                    project_ids.append(str(pid).strip())
        
        print(f"🧹 After processing: {project_ids}")
        
        if project_ids:
            try:
                project_ids = [int(pid) for pid in project_ids if str(pid).strip().isdigit()]
                print(f"🔍 Filtering by project IDs: {project_ids}")
                
                from meralcoapp.models import Project
                existing_projects = Project.objects.filter(project_id__in=project_ids).values_list('project_id', flat=True)
                print(f"✅ Existing projects in DB: {list(existing_projects)}")
                
                queryset = queryset.filter(project_id__in=project_ids)
                print(f"✅ Found {queryset.count()} work orders matching project IDs")
                
                if queryset.count() == 0:
                    all_project_ids = queryset.model.objects.values_list('project_id', flat=True).distinct()
                    print(f"⚠️ Available project_ids in work orders: {list(all_project_ids)[:20]}")
                    
            except ValueError as e:
                print(f"❌ Error converting project IDs: {e}")
        else:
            print("⚠️ No valid project IDs provided")

        # ----------------------
        # Date range filter - CUSTOM HANDLING
        # ----------------------
        start_date = params.get('start_date')
        end_date = params.get('end_date')
        if start_date:
            queryset = queryset.filter(date_received_jacket_ps__gte=start_date)
        if end_date:
            queryset = queryset.filter(date_received_jacket_ps__lte=end_date)

        # ----------------------
        # Overdue filter - CUSTOM HANDLING
        # ----------------------
        is_overdue = params.get('is_overdue')
        if is_overdue == 'true':
            queryset = queryset.filter(
                Q(days_comp__gt=60) |
                Q(date_received_jacket_ps__lt=timezone.now().date() - timedelta(days=90), date_comp__isnull=True)
            )

        # ----------------------
        # VIP filter - CUSTOM HANDLING (different param name)
        # ----------------------
        vip_only = params.get('vip_only')
        if vip_only == 'true':
            queryset = queryset.filter(vip=True)

        # ❌ REMOVE THESE - LET DRF HANDLE THEM:
        # municipality, assigned, status are in filterset_fields
        
        # ----------------------
        # Search filter - CUSTOM HANDLING
        # ----------------------
        search_query = params.get('search')
        if search_query:
            queryset = queryset.filter(
                Q(wo_no__icontains=search_query) |
                Q(description__icontains=search_query) |
                Q(location__icontains=search_query) |
                Q(area_of_responsibility__icontains=search_query)
            )

        final_queryset = queryset
        print(f"📊 FINAL QUERYSET COUNT: {final_queryset.count()}")
        print(f"📊 FIRST ITEM: {final_queryset.first()}")
        
        return final_queryset

    # ============================================================
    # DASHBOARD & STATISTICS ENDPOINTS
    # ============================================================
    
    @action(detail=False, methods=['get'])
    def dashboard_stats(self, request):
        """
        GET /api/work-orders/dashboard_stats/
        
        Comprehensive dashboard statistics
        """
        queryset = self.filter_queryset(self.get_queryset())
        total_count = queryset.count()
        
        # Status breakdown
        status_breakdown = queryset.values('status').annotate(
            count=Count('id')
        ).order_by('-count')
        
        # VIP count
        vip_count = queryset.filter(vip=True).count()
        
        # Overdue count
        overdue_count = queryset.filter(
            Q(days_comp__gt=60) | 
            Q(date_received_jacket_ps__lt=timezone.now().date() - timedelta(days=90), date_comp__isnull=True)
        ).count()
        
        # Average completion time
        avg_completion = queryset.filter(
            days_comp__isnull=False
        ).aggregate(avg_days=Avg('days_comp'))
        
        # Completion rate
        completed_count = queryset.filter(date_comp__isnull=False).count()
        completion_rate = (completed_count / total_count * 100) if total_count > 0 else 0
        
        # By municipality (top 10)
        by_municipality = queryset.values('municipality').annotate(
            count=Count('id')
        ).order_by('-count')[:10]
        
        # By assigned crew (top 10)
        by_assigned = queryset.values('assigned').annotate(
            count=Count('id')
        ).order_by('-count')[:10]
        
        # Recent work orders
        recent_wo = queryset.order_by('-created_at')[:10]
        
        return Response({
            'total_count': total_count,
            'status_breakdown': list(status_breakdown),
            'vip_count': vip_count,
            'overdue_count': overdue_count,
            'overdue_percentage': round((overdue_count / total_count * 100) if total_count > 0 else 0, 2),
            'avg_completion_days': round(avg_completion['avg_days'], 2) if avg_completion['avg_days'] else 0,
            'completion_rate': round(completion_rate, 2),
            'by_municipality': list(by_municipality),
            'by_assigned': list(by_assigned),
            'recent_work_orders': WorkOrderListSerializer(recent_wo, many=True).data
        })
    
    @action(detail=False, methods=['get'])
    def performance_metrics(self, request):
        """
        GET /api/work-orders/performance_metrics/
        
        Detailed performance metrics and KPIs
        """
        queryset = self.filter_queryset(self.get_queryset())
        completed = queryset.filter(date_comp__isnull=False)
        total = queryset.count()
        
        metrics = {
            'total_work_orders': total,
            'completed': completed.count(),
            'in_progress': queryset.filter(date_comp__isnull=True).count(),
            'completion_rate': round((completed.count() / total * 100) if total > 0 else 0, 2),
            
            # Time-based metrics
            'avg_wmtrl_to_fcomp': round(completed.aggregate(
                avg=Avg('days_wmtrl_to_fcomp_apt')
            )['avg'] or 0, 2),
            'avg_sched_to_fcomp': round(completed.aggregate(
                avg=Avg('days_sched_to_fcomp')
            )['avg'] or 0, 2),
            'avg_total_days': round(completed.aggregate(
                avg=Avg('days_comp')
            )['avg'] or 0, 2),
            
            # Index metrics
            'avg_ccti': round(completed.aggregate(
                avg=Avg('computed_index_wmtrl_to_fcomp_ccti')
            )['avg'] or 0, 2),
            'avg_comp_index': round(completed.aggregate(
                avg=Avg('computed_index_comp')
            )['avg'] or 0, 2),
            
            # Exclusions
            'ccti_exclusions': queryset.filter(for_ccti_exclusion=True).count(),
            'apt_exclusions': queryset.filter(for_apt_exclusion=True).count(),
            
            # Special categories
            'vip_projects': queryset.filter(vip=True).count(),
            'with_backjob': queryset.filter(with_back_job=True).count(),
            'encoded_in_eam': queryset.filter(encoded_in_eam=True).count(),
        }
        
        return Response(metrics)
    
    @action(detail=False, methods=['get'])
    def monthly_trends(self, request):
        """
        GET /api/work-orders/monthly_trends/
        
        Monthly trends for work order creation and completion
        """
        end_date = timezone.now().date()
        start_date = end_date - timedelta(days=365)
        
        queryset = self.filter_queryset(self.get_queryset())
        
        # Work orders received by month
        received_by_month = queryset.filter(
            date_received_jacket_ps__range=[start_date, end_date]
        ).annotate(
            month=TruncMonth('date_received_jacket_ps')
        ).values('month').annotate(
            count=Count('id'),
            vip_count=Count('id', filter=Q(vip=True))
        ).order_by('month')
        
        # Work orders completed by month
        completed_by_month = queryset.filter(
            date_comp__range=[start_date, end_date]
        ).annotate(
            month=TruncMonth('date_comp')
        ).values('month').annotate(
            count=Count('id'),
            avg_days=Avg('days_comp')
        ).order_by('month')
        
        return Response({
            'received_by_month': list(received_by_month),
            'completed_by_month': list(completed_by_month)
        })
    
    # ============================================================
    # FILTERING & GROUPING ENDPOINTS
    # ============================================================
    
    @action(detail=False, methods=['get'])
    def by_municipality(self, request):
        """
        GET /api/work-orders/by_municipality/
        
        Group work orders by municipality with statistics
        """
        queryset = self.filter_queryset(self.get_queryset())
        
        municipality_stats = queryset.values('municipality').annotate(
            total_wo=Count('id'),
            completed=Count('id', filter=Q(date_comp__isnull=False)),
            vip_count=Count('id', filter=Q(vip=True)),
            avg_completion_days=Avg('days_comp'),
            overdue=Count('id', filter=Q(days_comp__gt=60)),
            with_backjob=Count('id', filter=Q(with_back_job=True))
        ).order_by('-total_wo')
        
        return Response(list(municipality_stats))
    
    @action(detail=False, methods=['get'])
    def by_assigned(self, request):
        """
        GET /api/work-orders/by_assigned/
        
        Group work orders by assigned crew/person
        """
        queryset = self.filter_queryset(self.get_queryset())
        
        assigned_stats = queryset.values('assigned').annotate(
            total_wo=Count('id'),
            completed=Count('id', filter=Q(date_comp__isnull=False)),
            in_progress=Count('id', filter=Q(date_comp__isnull=True)),
            avg_completion_days=Avg('days_comp'),
            vip_count=Count('id', filter=Q(vip=True))
        ).order_by('-total_wo')
        
        return Response(list(assigned_stats))
    
    @action(detail=False, methods=['get'])
    def by_status(self, request):
        """
        GET /api/work-orders/by_status/
        
        Detailed statistics for each status
        """
        queryset = self.filter_queryset(self.get_queryset())
        
        status_stats = queryset.values('status').annotate(
            count=Count('id'),
            vip_count=Count('id', filter=Q(vip=True)),
            avg_days=Avg('days_comp'),
            with_exclusion=Count('id', filter=Q(for_ccti_exclusion=True) | Q(for_apt_exclusion=True))
        ).order_by('-count')
        
        return Response(list(status_stats))
    
    @action(detail=False, methods=['get'])
    def overdue_projects(self, request):
        """
        GET /api/work-orders/overdue_projects/
        
        Get all overdue work orders
        """
        queryset = self.filter_queryset(self.get_queryset())
        
        overdue = queryset.filter(
            Q(days_comp__gt=60) | 
            Q(date_received_jacket_ps__lt=timezone.now().date() - timedelta(days=90), date_comp__isnull=True)
        ).order_by('-days_comp')
        
        serializer = WorkOrderListSerializer(overdue, many=True)
        return Response(serializer.data)
    
    @action(detail=False, methods=['get'])
    def vip_projects(self, request):
        """
        GET /api/work-orders/vip_projects/
        
        Get all VIP work orders
        """
        queryset = self.filter_queryset(self.get_queryset())
        vip = queryset.filter(vip=True).order_by('-date_received_jacket_ps')
        
        serializer = WorkOrderListSerializer(vip, many=True)
        return Response(serializer.data)
    
    # ============================================================
    # TIMELINE & TRACKING ENDPOINTS
    # ============================================================
    
    @action(detail=True, methods=['get'])
    def timeline(self, request, pk=None):
        """
        GET /api/work-orders/{id}/timeline/
        
        Detailed timeline for specific work order
        """
        work_order = self.get_object()
        serializer = WorkOrderTimelineSerializer(work_order)
        return Response(serializer.data)
    
    @action(detail=True, methods=['post'])
    def update_milestone(self, request, pk=None):
        """
        POST /api/work-orders/{id}/update_milestone/
        
        Update specific milestone date
        
        Body: {"milestone": "date_wmtrl", "date": "2026-01-15"}
        """
        work_order = self.get_object()
        milestone = request.data.get('milestone')
        date_value = request.data.get('date')
        
        # List of valid date milestones
        allowed_milestones = [
            'date_received_jacket_ps', 'date_received_awarding_wo',
            'date_wmtrl', 'date_sched', 'date_received_by_vc',
            'actual_date_completed_on_site', 'date_fcomp', 'date_comp',
            'date_received_by_contractor', 'date_corrected', 'date_audit',
            'date_material_balancing', 'date_printed_pole_tag'
        ]
        
        if milestone not in allowed_milestones:
            return Response(
                {'error': f'Invalid milestone. Must be one of: {", ".join(allowed_milestones)}'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            setattr(work_order, milestone, date_value)
            work_order.save()
            
            serializer = WorkOrderSerializer(work_order)
            return Response({
                'message': f'{milestone} updated successfully',
                'work_order': serializer.data
            })
        except Exception as e:
            return Response(
                {'error': str(e)},
                status=status.HTTP_400_BAD_REQUEST
            )
    
    # ============================================================
    # EXCEL IMPORT/EXPORT ENDPOINTS - IMPROVED
    # ============================================================
    
    @action(detail=False, methods=['get'])
    def export_excel(self, request):
        """
        GET /api/work-orders/export_excel/
        
        Export work orders to Excel with improved column mapping
        """
        queryset = self.filter_queryset(self.get_queryset())
        
        # Prepare data with exact column names matching the model
        data = []
        for wo in queryset:
            data.append({
                # Basic Info
                'WO No': wo.wo_no,
                'Vendor ID': wo.vendor_id,
                'Project ID': wo.project_id,
                'VIP': 'Yes' if wo.vip else 'No',
                'Description': wo.description or '',
                'Location': wo.location or '',
                'Municipality': wo.municipality or '',
                'Area of Responsibility': wo.area_of_responsibility or '',
                
                # Dates - Receipt
                'Date Received Jacket (PS)': wo.date_received_jacket_ps,
                'Date Received Awarding WO': wo.date_received_awarding_wo,
                
                # Dates - Work Progress
                'Date WMTRL': wo.date_wmtrl,
                'Date Sched': wo.date_sched,
                'Date Received by VC': wo.date_received_by_vc,
                'Actual Date Completed on Site': wo.actual_date_completed_on_site,
                'Date FCOMP': wo.date_fcomp,
                'Date COMP': wo.date_comp,
                
                # Durations (APT/SPT)
                'Days WMTRL to FCOMP (APT)': wo.days_wmtrl_to_fcomp_apt,
                'Days Sched to FCOMP': wo.days_sched_to_fcomp,
                'Days COMP': wo.days_comp,
                'Date Needed (0.75) WMTRL to FCOMP': wo.date_needed_wmtrl_to_fcomp_075,
                'Date Needed (0.95) FCOMP': wo.date_needed_fcomp_095,
                'Date Needed (50 days) WMTRL to FCOMP': wo.date_needed_wmtrl_to_fcomp_50,
                'Computed Index WMTRL to FCOMP (CCTI)': float(wo.computed_index_wmtrl_to_fcomp_ccti) if wo.computed_index_wmtrl_to_fcomp_ccti else None,
                'Computed Index COMP': float(wo.computed_index_comp) if wo.computed_index_comp else None,
                
                # SPT Values
                'SPT M': wo.spt_m,
                'SPT L': wo.spt_l,
                'Duration 0.75 Days': wo.duration_075_days,
                'Duration 0.95 Days': wo.duration_095_days,
                'Target Days': wo.target_days,
                'SPT M for COMP': wo.spt_m_for_comp,
                'Duration COMP Days': wo.duration_comp_days,
                'Target Days COMP': wo.target_days_comp,
                'Date Needed to COMP': wo.date_needed_to_comp,
                'Ageing Days Since FCOMP': wo.ageing_days_since_fcomp,
                
                # Remarks & Status
                'Vendor Remarks': wo.vendor_remarks or '',
                'C1 Remarks': wo.c1_remarks or '',
                'Assigned': wo.assigned or '',
                'Status': wo.status or '',
                
                # Exclusions
                'Exclusion Reason': wo.exclusion_reason or '',
                'For CCTI Exclusion': 'Yes' if wo.for_ccti_exclusion else 'No',
                'Encoded in EAM': 'Yes' if wo.encoded_in_eam else 'No',
                'Validated by DCSAM': 'Yes' if wo.validated_by_dcsam else 'No',
                'For APT Exclusion': 'Yes' if wo.for_apt_exclusion else 'No',
                'Exclusion Start Date': wo.exclusion_start_date,
                'Exclusion Duration (Days)': wo.exclusion_duration_days,
                'Exclusion End Date': wo.exclusion_end_date,
                
                # COC
                'Remarks Follow Up By': wo.remarks_follow_up_by or '',
                'Remarks 2': wo.remarks_2 or '',
                'Date Needed Submit COC': wo.date_needed_submit_coc,
                'Ageing Submission COC': wo.ageing_submission_coc,
                'Date Completed from COC': wo.date_completed_from_coc,
                'Actual Received COC': wo.actual_received_coc,
                
                # Audit / Backjob
                'Date Audit': wo.date_audit,
                'Audit By': wo.audit_by or '',
                'With Back Job': 'Yes' if wo.with_back_job else 'No',
                'Backjob Tagged in EAM': 'Yes' if wo.backjob_tagged_eam else 'No',
                
                # Contractor / Correction
                'Date Received by Contractor': wo.date_received_by_contractor,
                'Date Corrected': wo.date_corrected,
                'Date Material Balancing': wo.date_material_balancing,
                'Material Balancing By': wo.material_balancing_by or '',
                'Yes/No Flag': 'Yes' if wo.yes_no_flag else 'No',
                'Emailed to Meter': 'Yes' if wo.emailed_to_meter else 'No',
                'DT Correction Method': wo.dt_correction_method or '',
                'TLN': wo.tln or '',
                'With Pole Replacement': 'Yes' if wo.with_pole_replacement else 'No',
                'Actual Field Status': wo.actual_field_status or '',
                'Remarks 3': wo.remarks_3 or '',
                'ABF Printed By': wo.abf_printed_by or '',
                'Date Printed Pole Tag Form': wo.date_printed_pole_tag_form,
                'Pole TLN Tags': wo.pole_tln_tags or '',
                
                # APT / CCTI with Exclusion
                'Exclusion Days (APT)': wo.exclusion_days_apt,
                'APT with Exclusion': wo.apt_with_exclusion,
                'Exclusion Days (CCTI)': wo.exclusion_days_ccti,
                'Duration CCTI with Exclusion': wo.duration_ccti_with_exclusion,
                'CCTI with Exclusion': float(wo.ccti_with_exclusion) if wo.ccti_with_exclusion else None,
                
                # Performance
                'E2E PRDI': float(wo.e2e_prdi) if wo.e2e_prdi else None,
                'Current CCTI with Exclusion': float(wo.current_ccti_with_exclusion) if wo.current_ccti_with_exclusion else None,
                'Current CCTI': float(wo.current_ccti) if wo.current_ccti else None,
                'Final CCTI Less Than FCOMP': float(wo.final_ccti_less_than_fcomp) if wo.final_ccti_less_than_fcomp else None,
                'PRDI': wo.prdi or '',
                'Days Ageing': wo.days_ageing,
                'Rev/Non-Rev': wo.rev_non_rev or '',
                'Age Bracket': wo.age_bracket or '',
                
                # NTC
                'NTC Date Created': wo.ntc_date_created,
                'NTC Amount': float(wo.ntc_amount) if wo.ntc_amount else None,
                'NTC': wo.ntc or '',
                'NTC Date Received by Contractor': wo.ntc_date_received_by_contractor,
                'NTC Date Completed': wo.ntc_date_completed,
                'NTC Running Days': wo.ntc_running_days,
                
                # NOV / Debit
                'NOV Debit Memo Date Created': wo.nov_debit_memo_date_created,
                'NOV Amount': float(wo.nov_amount) if wo.nov_amount else None,
                'NOV Date Received by Contractor': wo.nov_date_received_by_contractor,
                
                # Supervisor
                'Ext': wo.ext or '',
                'Updated Supv': 'Yes' if wo.updated_supv else 'No',
                'Supv Name': wo.supv_name or '',
                'Status as of 2025-04-04': wo.status_as_of_2025_04_04 or '',
                'Diff Days WMTRL to Sched (2025)': wo.diff_days_wmtrl_to_sched_2025,
                'Filter Flag': wo.filter_flag or '',
                'Supervisor Full Name': wo.supervisor_full_name or '',
                
                # Timestamps
                'Created At': wo.created_at,
                'Updated At': wo.updated_at,
            })
        
        # Create DataFrame
        df = pd.DataFrame(data)
        
        # Create Excel file in memory
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Work Orders', index=False)
            
            # ✅ FIX: Proper column width adjustment using openpyxl's get_column_letter
            from openpyxl.utils import get_column_letter
            
            worksheet = writer.sheets['Work Orders']
            for idx, col in enumerate(df.columns, start=1):
                max_length = max(
                    df[col].astype(str).apply(len).max(),
                    len(col)
                )
                # Use openpyxl's built-in function to get proper column letter
                col_letter = get_column_letter(idx)
                worksheet.column_dimensions[col_letter].width = min(max_length + 2, 50)
        
        output.seek(0)
        
        # Create response
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=work_orders_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        
        return response
    
    

    @action(detail=False, methods=['post'])
    def import_excel(self, request):
        """
        POST /api/work-orders/import_excel/
        
        Import work orders from Excel file with complete column mapping
        Automatically assigns vendor_id from authenticated user
        """
        if 'file' not in request.FILES:
            return Response(
                {'error': 'No file provided'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        file = request.FILES['file']
        vendor_id = None
        
        if request.user and request.user.is_authenticated:
            # ✅ GET VENDOR_ID FROM AUTHENTICATED USER
            if hasattr(request.user, 'vendor_id'):
                vendor_id = request.user.vendor_id
            elif hasattr(request.user, 'user_id'):
                # Try to get vendor_id from User model
                vendor_id = request.user.user_id
            
            if not vendor_id:
                return Response(
                    {'error': 'Unable to determine vendor ID from authenticated user'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            logger.info(f"📥 Import started by vendor: {vendor_id}")
        
        try:
            # Read Excel file
            df = pd.read_excel(file)
            
            logger.info(f"Excel columns found: {df.columns.tolist()}")
            
            imported = 0
            updated = 0
            errors = []
            
            # COMPLETE Column mapping (Excel column name -> Model field name)
            column_mapping = {
                # Basic Info
                'WO No': 'wo_no',
                'VIP': 'vip',
                'Description': 'description',
                'Location': 'location',
                'Municipality': 'municipality',
                'Area of Responsibility': 'area_of_responsibility',
                
                # NOTE: Removed 'Vendor ID' from mapping - will be auto-assigned
                'Project ID': 'project_id',
                
                # Dates - Receipt
                'Date Received Jacket (PS)': 'date_received_jacket_ps',
                'Date Received Awarding WO': 'date_received_awarding_wo',
                
                # Dates - Work Progress
                'Date WMTRL': 'date_wmtrl',
                'Date Sched': 'date_sched',
                'Date Received by VC': 'date_received_by_vc',
                'Actual Date Completed on Site': 'actual_date_completed_on_site',
                'Date FCOMP': 'date_fcomp',
                'Date COMP': 'date_comp',
                
                # Durations (APT/SPT)
                'Days WMTRL to FCOMP (APT)': 'days_wmtrl_to_fcomp_apt',
                'Days Sched to FCOMP': 'days_sched_to_fcomp',
                'Days COMP': 'days_comp',
                'Date Needed (0.75) WMTRL to FCOMP': 'date_needed_wmtrl_to_fcomp_075',
                'Date Needed (0.95) FCOMP': 'date_needed_fcomp_095',
                'Date Needed (50 days) WMTRL to FCOMP': 'date_needed_wmtrl_to_fcomp_50',
                'Computed Index WMTRL to FCOMP (CCTI)': 'computed_index_wmtrl_to_fcomp_ccti',
                'Computed Index COMP': 'computed_index_comp',
                
                # SPT Values
                'SPT M': 'spt_m',
                'SPT L': 'spt_l',
                'Duration 0.75 Days': 'duration_075_days',
                'Duration 0.95 Days': 'duration_095_days',
                'Target Days': 'target_days',
                'SPT M for COMP': 'spt_m_for_comp',
                'Duration COMP Days': 'duration_comp_days',
                'Target Days COMP': 'target_days_comp',
                'Date Needed to COMP': 'date_needed_to_comp',
                'Ageing Days Since FCOMP': 'ageing_days_since_fcomp',
                
                # Remarks & Status
                'Vendor Remarks': 'vendor_remarks',
                'C1 Remarks': 'c1_remarks',
                'Assigned': 'assigned',
                'Status': 'status',
                
                # Exclusions
                'Exclusion Reason': 'exclusion_reason',
                'For CCTI Exclusion': 'for_ccti_exclusion',
                'Encoded in EAM': 'encoded_in_eam',
                'Validated by DCSAM': 'validated_by_dcsam',
                'For APT Exclusion': 'for_apt_exclusion',
                'Exclusion Start Date': 'exclusion_start_date',
                'Exclusion Duration (Days)': 'exclusion_duration_days',
                'Exclusion End Date': 'exclusion_end_date',
                
                # COC
                'Remarks Follow Up By': 'remarks_follow_up_by',
                'Remarks 2': 'remarks_2',
                'Date Needed Submit COC': 'date_needed_submit_coc',
                'Ageing Submission COC': 'ageing_submission_coc',
                'Date Completed from COC': 'date_completed_from_coc',
                'Actual Received COC': 'actual_received_coc',
                
                # Audit / Backjob
                'Date Audit': 'date_audit',
                'Audit By': 'audit_by',
                'With Back Job': 'with_back_job',
                'Backjob Tagged in EAM': 'backjob_tagged_eam',
                
                # Contractor / Correction
                'Date Received by Contractor': 'date_received_by_contractor',
                'Date Corrected': 'date_corrected',
                'Date Material Balancing': 'date_material_balancing',
                'Material Balancing By': 'material_balancing_by',
                'Yes/No Flag': 'yes_no_flag',
                'Emailed to Meter': 'emailed_to_meter',
                'DT Correction Method': 'dt_correction_method',
                'TLN': 'tln',
                'With Pole Replacement': 'with_pole_replacement',
                'Actual Field Status': 'actual_field_status',
                'Remarks 3': 'remarks_3',
                'ABF Printed By': 'abf_printed_by',
                'Date Printed Pole Tag Form': 'date_printed_pole_tag_form',
                'Pole TLN Tags': 'pole_tln_tags',
                
                # APT / CCTI with Exclusion
                'Exclusion Days (APT)': 'exclusion_days_apt',
                'APT with Exclusion': 'apt_with_exclusion',
                'Exclusion Days (CCTI)': 'exclusion_days_ccti',
                'Duration CCTI with Exclusion': 'duration_ccti_with_exclusion',
                'CCTI with Exclusion': 'ccti_with_exclusion',
                
                # Performance
                'E2E PRDI': 'e2e_prdi',
                'Current CCTI with Exclusion': 'current_ccti_with_exclusion',
                'Current CCTI': 'current_ccti',
                'Final CCTI Less Than FCOMP': 'final_ccti_less_than_fcomp',
                'PRDI': 'prdi',
                'Days Ageing': 'days_ageing',
                'Rev/Non-Rev': 'rev_non_rev',
                'Age Bracket': 'age_bracket',
                
                # NTC
                'NTC Date Created': 'ntc_date_created',
                'NTC Amount': 'ntc_amount',
                'NTC': 'ntc',
                'NTC Date Received by Contractor': 'ntc_date_received_by_contractor',
                'NTC Date Completed': 'ntc_date_completed',
                'NTC Running Days': 'ntc_running_days',
                
                # NOV / Debit
                'NOV Debit Memo Date Created': 'nov_debit_memo_date_created',
                'NOV Amount': 'nov_amount',
                'NOV Date Received by Contractor': 'nov_date_received_by_contractor',
                
                # Supervisor
                'Ext': 'ext',
                'Updated Supv': 'updated_supv',
                'Supv Name': 'supv_name',
                'Status as of 2025-04-04': 'status_as_of_2025_04_04',
                'Diff Days WMTRL to Sched (2025)': 'diff_days_wmtrl_to_sched_2025',
                'Filter Flag': 'filter_flag',
                'Supervisor Full Name': 'supervisor_full_name',
            }
            
            for index, row in df.iterrows():
                try:
                    # Skip rows without WO number
                    if pd.isna(row.get('WO No')):
                        continue
                    
                    work_order_data = {}
                    
                    # Map columns
                    for excel_col, model_field in column_mapping.items():
                        if excel_col in df.columns:
                            value = row.get(excel_col)
                            
                            # Skip NaN values
                            if pd.isna(value):
                                continue
                            
                            # Convert Yes/No to boolean
                            if isinstance(value, str) and value.lower() in ['yes', 'no']:
                                value = value.lower() == 'yes'
                            
                            # Convert dates
                            if 'date' in model_field.lower() and value:
                                try:
                                    value = pd.to_datetime(value).date()
                                except:
                                    value = None
                            
                            work_order_data[model_field] = value
                    
                    # ✅ AUTOMATICALLY ASSIGN VENDOR_ID
                    if vendor_id:
                        work_order_data['vendor_id'] = vendor_id
                        
                        logger.info(f"Row {index + 2}: Assigning vendor_id={vendor_id} to WO {work_order_data.get('wo_no')}")
                        
                    # Create or update work order
                    wo_no = work_order_data.get('wo_no')
                    if wo_no:
                        obj, created = WorkOrder.objects.update_or_create(
                            wo_no=wo_no,
                            defaults=work_order_data
                        )
                        
                        if created:
                            imported += 1
                            logger.info(f"✅ Created WO {wo_no} with vendor_id={vendor_id}")
                        else:
                            updated += 1
                            logger.info(f"✅ Updated WO {wo_no} with vendor_id={vendor_id}")
                
                except Exception as e:
                    errors.append({
                        'row': index + 2,
                        'wo_no': row.get('WO No', 'Unknown'),
                        'error': str(e)
                    })
                    logger.error(f"Error importing row {index + 2}: {e}")
            
            return Response({
                'message': f'Successfully processed {imported + updated} work orders for vendor {vendor_id}',
                'vendor_id': vendor_id,
                'imported': imported,
                'updated': updated,
                'errors': errors,
                'error_count': len(errors)
            })
        
        except Exception as e:
            logger.error(f"Excel import failed: {e}")
            return Response(
                {'error': f'Failed to process file: {str(e)}'},
                status=status.HTTP_400_BAD_REQUEST
            )


    @action(detail=False, methods=['get'])
    def download_template(self, request):
        """
        GET /api/work-orders/download_template/
        
        Download Excel template with ALL columns for importing work orders
        """
        # Create empty DataFrame with ALL columns matching the export format
        template_data = {
            # Basic Info
            'WO No': [],
            'VIP': [],
            'Description': [],
            'Location': [],
            'Municipality': [],
            'Area of Responsibility': [],
            
            # Dates - Receipt
            'Date Received Jacket (PS)': [],
            'Date Received Awarding WO': [],
            
            # Dates - Work Progress
            'Date WMTRL': [],
            'Date Sched': [],
            'Date Received by VC': [],
            'Actual Date Completed on Site': [],
            'Date FCOMP': [],
            'Date COMP': [],
            
            # Durations (APT/SPT)
            'Days WMTRL to FCOMP (APT)': [],
            'Days Sched to FCOMP': [],
            'Days COMP': [],
            'Date Needed (0.75) WMTRL to FCOMP': [],
            'Date Needed (0.95) FCOMP': [],
            'Date Needed (50 days) WMTRL to FCOMP': [],
            'Computed Index WMTRL to FCOMP (CCTI)': [],
            'Computed Index COMP': [],
            
            # SPT Values
            'SPT M': [],
            'SPT L': [],
            'Duration 0.75 Days': [],
            'Duration 0.95 Days': [],
            'Target Days': [],
            'SPT M for COMP': [],
            'Duration COMP Days': [],
            'Target Days COMP': [],
            'Date Needed to COMP': [],
            'Ageing Days Since FCOMP': [],
            
            # Remarks & Status
            'Vendor Remarks': [],
            'C1 Remarks': [],
            'Assigned': [],
            'Status': [],
            
            # Exclusions
            'Exclusion Reason': [],
            'For CCTI Exclusion': [],
            'Encoded in EAM': [],
            'Validated by DCSAM': [],
            'For APT Exclusion': [],
            'Exclusion Start Date': [],
            'Exclusion Duration (Days)': [],
            'Exclusion End Date': [],
            
            # COC
            'Remarks Follow Up By': [],
            'Remarks 2': [],
            'Date Needed Submit COC': [],
            'Ageing Submission COC': [],
            'Date Completed from COC': [],
            'Actual Received COC': [],
            
            # Audit / Backjob
            'Date Audit': [],
            'Audit By': [],
            'With Back Job': [],
            'Backjob Tagged in EAM': [],
            
            # Contractor / Correction
            'Date Received by Contractor': [],
            'Date Corrected': [],
            'Date Material Balancing': [],
            'Material Balancing By': [],
            'Yes/No Flag': [],
            'Emailed to Meter': [],
            'DT Correction Method': [],
            'TLN': [],
            'With Pole Replacement': [],
            'Actual Field Status': [],
            'Remarks 3': [],
            'ABF Printed By': [],
            'Date Printed Pole Tag Form': [],
            'Pole TLN Tags': [],
            
            # APT / CCTI with Exclusion
            'Exclusion Days (APT)': [],
            'APT with Exclusion': [],
            'Exclusion Days (CCTI)': [],
            'Duration CCTI with Exclusion': [],
            'CCTI with Exclusion': [],
            
            # Performance
            'E2E PRDI': [],
            'Current CCTI with Exclusion': [],
            'Current CCTI': [],
            'Final CCTI Less Than FCOMP': [],
            'PRDI': [],
            'Days Ageing': [],
            'Rev/Non-Rev': [],
            'Age Bracket': [],
            
            # NTC
            'NTC Date Created': [],
            'NTC Amount': [],
            'NTC': [],
            'NTC Date Received by Contractor': [],
            'NTC Date Completed': [],
            'NTC Running Days': [],
            
            # NOV / Debit
            'NOV Debit Memo Date Created': [],
            'NOV Amount': [],
            'NOV Date Received by Contractor': [],
            
            # Supervisor
            'Ext': [],
            'Updated Supv': [],
            'Supv Name': [],
            'Status as of 2025-04-04': [],
            'Diff Days WMTRL to Sched (2025)': [],
            'Filter Flag': [],
            'Supervisor Full Name': [],
        }
        
        df = pd.DataFrame(template_data)
        
        # Create Excel file
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Template', index=False)
            
            # Add instructions sheet with detailed guidance
            instructions = pd.DataFrame({
                'Field': [
                    'WO No',
                    'Vendor ID / Project ID',
                    'VIP',
                    'Dates',
                    'Boolean Fields',
                    'Numbers',
                    'Text Fields',
                ],
                'Instructions': [
                    'REQUIRED - Must be unique. This is the primary identifier.',
                    'Should be valid IDs from the system. Leave blank if not available.',
                    'Use "Yes" or "No" for VIP and all boolean fields (Exclusions, Flags, etc.)',
                    'Use YYYY-MM-DD format (e.g., 2026-02-06) for all date fields.',
                    'Use "Yes" or "No" for: VIP, For CCTI Exclusion, For APT Exclusion, With Back Job, Encoded in EAM, etc.',
                    'Use numbers without commas. Decimals allowed for amounts and indices.',
                    'Description, Location, Remarks, etc. can contain any text. Leave empty if not available.',
                ]
            })
            instructions.to_excel(writer, sheet_name='Instructions', index=False)
            
            # Auto-adjust column widths for both sheets
            for sheet_name in writer.sheets:
                worksheet = writer.sheets[sheet_name]
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
        
        output.seek(0)
        
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=work_orders_template.xlsx'
        
        return response
    
    # ============================================================
    # BULK OPERATIONS
    # ============================================================
    
    def list(self, request, *args, **kwargs):
        """Override list to debug and handle filtering properly"""
        # Get the queryset (already filtered by get_queryset)
        queryset = self.get_queryset()
        
        print(f"🎯 LIST METHOD - Initial queryset count: {queryset.count()}")
        print(f"🎯 First item: {queryset.first()}")
        
        # Apply additional filters (search, ordering) but NOT project_id again
        # Since project_id is not in filterset_fields, this should be safe
        queryset = self.filter_queryset(queryset)
        
        print(f"🎯 After filter_queryset count: {queryset.count()}")
        
        # Let DRF handle pagination
        page = self.paginate_queryset(queryset)
        
        print(f"🎯 After pagination: {len(page) if page else 'None'}")
        
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            print(f"🎯 Serialized data length: {len(serializer.data)}")
            return self.get_paginated_response(serializer.data)

        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)
            
    @action(detail=False, methods=['post'])
    def bulk_update_status(self, request):
        """
        POST /api/work-orders/bulk_update_status/
        
        Update status for multiple work orders
        
        Body: {"work_order_ids": [1, 2, 3], "status": "COMPLETED"}
        """
        work_order_ids = request.data.get('work_order_ids', [])
        new_status = request.data.get('status')
        
        if not work_order_ids or not new_status:
            return Response(
                {'error': 'work_order_ids and status are required'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        updated = WorkOrder.objects.filter(
            id__in=work_order_ids
        ).update(status=new_status)
        
        return Response({
            'message': f'Updated {updated} work orders',
            'updated_count': updated
        })
    
    @action(detail=False, methods=['delete'])
    def bulk_delete(self, request):
        """
        DELETE /api/work-orders/bulk_delete/
        
        Delete multiple work orders
        
        Body: {"work_order_ids": [1, 2, 3]}
        """
        work_order_ids = request.data.get('work_order_ids', [])
        
        if not work_order_ids:
            return Response(
                {'error': 'work_order_ids is required'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        deleted_count, _ = WorkOrder.objects.filter(
            id__in=work_order_ids
        ).delete()
        
        return Response({
            'message': f'Deleted {deleted_count} work orders',
            'deleted_count': deleted_count
        })




class WorkOrderDocumentViewSet(viewsets.ModelViewSet):
    queryset = WorkOrderDocument.objects.all()
    serializer_class = WorkOrderDocumentSerializer
    permission_classes = [AllowAny]
    parser_classes = [MultiPartParser, FormParser]

    def create(self, request, *args, **kwargs):
        """Handle document upload with proper user assignment"""
        try:
            data = request.data.copy()
            
            # Get the uploaded_by value from request
            uploaded_by_id = data.get('uploaded_by')
            
            # Validate and convert to integer
            if uploaded_by_id:
                try:
                    uploaded_by_id = int(uploaded_by_id)
                    # Verify user exists
                    User.objects.get(user_id=uploaded_by_id)
                    data['uploaded_by_id'] = uploaded_by_id
                except (User.DoesNotExist, ValueError, TypeError):
                    return Response(
                        {'error': f'Invalid user ID: {uploaded_by_id}'}, 
                        status=status.HTTP_400_BAD_REQUEST
                    )
                # Remove the old field name
                data.pop('uploaded_by', None)
            
            # Validate work_order exists
            work_order_id = data.get('work_order')
            if work_order_id:
                try:
                    WorkOrder.objects.get(id=work_order_id)
                except WorkOrder.DoesNotExist:
                    return Response(
                        {'error': f'Work order with ID {work_order_id} not found'}, 
                        status=status.HTTP_400_BAD_REQUEST
                    )
            
            # Create serializer with modified data
            serializer = self.get_serializer(data=data)
            serializer.is_valid(raise_exception=True)
            
            # Save the document
            self.perform_create(serializer)
            
            headers = self.get_success_headers(serializer.data)
            return Response(
                serializer.data, 
                status=status.HTTP_201_CREATED, 
                headers=headers
            )
            
        except Exception as e:
            return Response(
                {'error': str(e)}, 
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    def get_queryset(self):
        """Filter queryset based on query parameters"""
        queryset = super().get_queryset()

        # Optional filters
        work_order_id = self.request.query_params.get('work_order')
        document_type = self.request.query_params.get('document_type')
        uploaded_by_id = self.request.query_params.get('uploaded_by_id')
        is_approved = self.request.query_params.get('is_approved')

        if work_order_id:
            queryset = queryset.filter(work_order_id=work_order_id)

        if document_type:
            queryset = queryset.filter(document_type=document_type)
        
        if uploaded_by_id:
            queryset = queryset.filter(uploaded_by_id=uploaded_by_id)
        
        if is_approved is not None:
            queryset = queryset.filter(is_approved=is_approved.lower() == 'true')

        return queryset


# ============================================
# CREW MONITORING VIEWSETS
# ============================================

class CrewTypeViewSet(viewsets.ModelViewSet):
    queryset = CrewType.objects.all()
    serializer_class = CrewTypeSerializer
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['crew_code', 'crew_name']
    ordering = ['crew_code']


class DailyCrewMonitoringViewSet(viewsets.ModelViewSet):
    queryset = DailyCrewMonitoring.objects.all()
    serializer_class = DailyCrewMonitoringSerializer
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_fields = ['crew_type', 'monitoring_date']
    ordering = ['-monitoring_date']
    
    @action(detail=False, methods=['get'])
    def monthly_summary(self, request):
        """Get monthly summary for all crews"""
        month_param = request.query_params.get('month')
        
        if month_param:
            try:
                month_date = datetime.strptime(month_param, '%Y-%m-%d').date()
            except ValueError:
                return Response({'error': 'Invalid date format. Use YYYY-MM-DD'}, 
                              status=status.HTTP_400_BAD_REQUEST)
        else:
            month_date = timezone.now().date().replace(day=1)
        
        # Get next month for filtering
        if month_date.month == 12:
            next_month = month_date.replace(year=month_date.year + 1, month=1)
        else:
            next_month = month_date.replace(month=month_date.month + 1)
        
        summary = DailyCrewMonitoring.objects.filter(
            monitoring_date__gte=month_date,
            monitoring_date__lt=next_month
        ).values(
            'crew_type__crew_code',
            'crew_type__crew_name'
        ).annotate(
            total_productivity=Sum('weighted_productivity'),
            total_peso_value=Sum('monthly_peso_value'),
            average_daily_productivity=Avg('weighted_productivity'),
            days_recorded=Count('monitoring_date')
        )
        
        return Response(summary)
    
    @action(detail=False, methods=['get'])
    def crew_comparison(self, request):
        """Compare productivity across all crews"""
        month_param = request.query_params.get('month')
        
        if month_param:
            month_date = datetime.strptime(month_param, '%Y-%m-%d').date()
        else:
            month_date = timezone.now().date().replace(day=1)
        
        if month_date.month == 12:
            next_month = month_date.replace(year=month_date.year + 1, month=1)
        else:
            next_month = month_date.replace(month=month_date.month + 1)
        
        comparison = DailyCrewMonitoring.objects.filter(
            monitoring_date__gte=month_date,
            monitoring_date__lt=next_month
        ).values('crew_type__crew_code').annotate(
            total_productivity=Sum('weighted_productivity'),
            avg_productivity=Avg('weighted_productivity'),
            total_value=Sum('daily_peso_value')
        ).order_by('-total_productivity')
        
        return Response(comparison)


# ============================================
# QI MONITORING VIEWSETS
# ============================================

class QIWeeklyAccomplishmentViewSet(viewsets.ModelViewSet):
    queryset = QIWeeklyAccomplishment.objects.all()
    serializer_class = QIWeeklyAccomplishmentSerializer
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_fields = ['qi_user', 'week_start_date', 'target_met']
    ordering = ['-week_start_date']
    
    @action(detail=False, methods=['get'])
    def current_week(self, request):
        """Get current week's accomplishments for all QIs"""
        today = timezone.now().date()
        week_start = today - timedelta(days=today.weekday())
        
        current_week = QIWeeklyAccomplishment.objects.filter(
            week_start_date=week_start
        )
        
        serializer = self.get_serializer(current_week, many=True)
        return Response(serializer.data)
    
    @action(detail=False, methods=['get'])
    def qi_performance(self, request):
        """Get performance statistics for all QIs"""
        qi_id = request.query_params.get('qi_user')
        
        queryset = self.get_queryset()
        if qi_id:
            queryset = queryset.filter(qi_user_id=qi_id)
        
        stats = queryset.aggregate(
            total_weeks=Count('id'),
            total_inspections=Sum('total_inspections'),
            avg_weekly_inspections=Avg('total_inspections'),
            weeks_target_met=Count('id', filter=Q(target_met=True))
        )
        
        if stats['total_weeks']:
            stats['target_achievement_rate'] = (
                stats['weeks_target_met'] / stats['total_weeks'] * 100
            )
        else:
            stats['target_achievement_rate'] = 0
        
        return Response(stats)


class QIMonthlyAccomplishmentViewSet(viewsets.ModelViewSet):
    queryset = QIMonthlyAccomplishment.objects.all()
    serializer_class = QIMonthlyAccomplishmentSerializer
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_fields = ['qi_user', 'month', 'target_met']
    ordering = ['-month']
    
    @action(detail=False, methods=['get'])
    def current_month(self, request):
        """Get current month's accomplishments"""
        current_month = timezone.now().date().replace(day=1)
        
        accomplishments = QIMonthlyAccomplishment.objects.filter(
            month=current_month
        )
        
        serializer = self.get_serializer(accomplishments, many=True)
        return Response(serializer.data)
    
    @action(detail=False, methods=['get'])
    def yearly_summary(self, request):
        """Get yearly summary for a QI"""
        qi_id = request.query_params.get('qi_user')
        year = request.query_params.get('year', timezone.now().year)
        
        if not qi_id:
            return Response({'error': 'qi_user parameter required'}, 
                          status=status.HTTP_400_BAD_REQUEST)
        
        start_date = datetime(int(year), 1, 1).date()
        end_date = datetime(int(year), 12, 31).date()
        
        yearly_data = QIMonthlyAccomplishment.objects.filter(
            qi_user_id=qi_id,
            month__gte=start_date,
            month__lte=end_date
        ).order_by('month')
        
        serializer = self.get_serializer(yearly_data, many=True)
        
        # Calculate totals
        totals = yearly_data.aggregate(
            total_inspections=Sum('total_inspections'),
            total_target=Sum('target_inspections'),
            months_target_met=Count('id', filter=Q(target_met=True))
        )
        
        return Response({
            'monthly_data': serializer.data,
            'yearly_totals': totals
        })


# ============================================
# PCA VIEWSETS
# ============================================

class PCAGoalViewSet(viewsets.ModelViewSet):
    queryset = PCAGoal.objects.all()
    serializer_class = PCAGoalSerializer
    ordering = ['-month']


class PCASummaryViewSet(viewsets.ModelViewSet):
    queryset = PCASummary.objects.all()
    serializer_class = PCASummarySerializer
    ordering = ['-month']
    
    @action(detail=False, methods=['get'])
    def current_month(self, request):
        """Get current month's PCA summary"""
        current_month = timezone.now().date().replace(day=1)
        
        try:
            summary = PCASummary.objects.get(month=current_month)
            serializer = self.get_serializer(summary)
            return Response(serializer.data)
        except PCASummary.DoesNotExist:
            return Response({'message': 'No summary for current month'}, 
                          status=status.HTTP_404_NOT_FOUND)
    
    @action(detail=False, methods=['post'])
    def generate_summary(self, request):
        """Generate PCA summary for a specific month"""
        month_param = request.data.get('month')
        
        if not month_param:
            return Response({'error': 'month parameter required'}, 
                          status=status.HTTP_400_BAD_REQUEST)
        
        try:
            month_date = datetime.strptime(month_param, '%Y-%m-%d').date()
        except ValueError:
            return Response({'error': 'Invalid date format. Use YYYY-MM-DD'}, 
                          status=status.HTTP_400_BAD_REQUEST)
        
        # Calculate statistics from WorkOrder model
        if month_date.month == 12:
            next_month = month_date.replace(year=month_date.year + 1, month=1)
        else:
            next_month = month_date.replace(month=month_date.month + 1)
        
        # Get work orders for the month
        monthly_wos = WorkOrder.objects.filter(
            date_received_by_vc__gte=month_date,
            date_received_by_vc__lt=next_month
        )
        
        ytd_energized = WorkOrder.objects.filter(
            date_received_by_vc__year=month_date.year,
            date_received_by_vc__lte=month_date
        ).count()
        
        completed_count = monthly_wos.filter(status='AUDITED').count()
        cancelled_count = monthly_wos.filter(status='CANCELLED').count()
        
        # Get or create summary
        summary, created = PCASummary.objects.update_or_create(
            month=month_date,
            defaults={
                'ytd_energized': ytd_energized,
                'completed_count': completed_count,
                'cancelled_count': cancelled_count,
                'new_work_orders_count': monthly_wos.count()
            }
        )
        
        serializer = self.get_serializer(summary)
        return Response(serializer.data)


# ============================================
# VENDOR PRODUCTIVITY VIEWSETS
# ============================================

class VendorProductivityMonthlyViewSet(viewsets.ModelViewSet):
    queryset = VendorProductivityMonthly.objects.all()
    serializer_class = VendorProductivityMonthlySerializer
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_fields = ['vendor', 'month']
    ordering = ['-month']
    
    @action(detail=False, methods=['get'])
    def comparison(self, request):
        """Compare all vendors for a specific month"""
        month_param = request.query_params.get('month')
        
        if not month_param:
            month_date = timezone.now().date().replace(day=1)
        else:
            month_date = datetime.strptime(month_param, '%Y-%m-%d').date()
        
        comparison = VendorProductivityMonthly.objects.filter(
            month=month_date
        ).select_related('vendor').order_by('-productivity_percentage')
        
        serializer = self.get_serializer(comparison, many=True)
        return Response(serializer.data)
    
    @action(detail=False, methods=['get'])
    def vendor_trend(self, request):
        """Get productivity trend for a specific vendor"""
        vendor_id = request.query_params.get('vendor')
        months = int(request.query_params.get('months', 6))
        
        if not vendor_id:
            return Response({'error': 'vendor parameter required'}, 
                          status=status.HTTP_400_BAD_REQUEST)
        
        end_date = timezone.now().date().replace(day=1)
        start_date = end_date - timedelta(days=30 * months)
        
        trend_data = VendorProductivityMonthly.objects.filter(
            vendor_id=vendor_id,
            month__gte=start_date,
            month__lte=end_date
        ).order_by('month')
        
        serializer = self.get_serializer(trend_data, many=True)
        return Response(serializer.data)


# ============================================
# AGEING ANALYSIS VIEWSETS
# ============================================

class AgeingAnalysisViewSet(viewsets.ModelViewSet):
    queryset = AgeingAnalysis.objects.all()
    serializer_class = AgeingAnalysisSerializer
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_fields = ['analysis_date', 'age_bracket', 'supervisor', 'crew']
    ordering = ['-analysis_date', '-age_in_days']
    
    @action(detail=False, methods=['get'])
    def current_analysis(self, request):
        """Get current ageing analysis"""
        latest_date = AgeingAnalysis.objects.latest('analysis_date').analysis_date
        current = AgeingAnalysis.objects.filter(analysis_date=latest_date)
        
        serializer = self.get_serializer(current, many=True)
        return Response(serializer.data)
    
    @action(detail=False, methods=['get'])
    def summary_by_bracket(self, request):
        """Get summary statistics by age bracket"""
        analysis_date = request.query_params.get('date')
        
        if analysis_date:
            date_filter = datetime.strptime(analysis_date, '%Y-%m-%d').date()
        else:
            date_filter = AgeingAnalysis.objects.latest('analysis_date').analysis_date
        
        summary = AgeingAnalysis.objects.filter(
            analysis_date=date_filter
        ).values('age_bracket').annotate(
            count=Count('id'),
            exclusion_duration=Sum('work_order__exclusion_duration')
        ).order_by('age_bracket')
        
        return Response(summary)
    
    @action(detail=False, methods=['post'])
    def generate_analysis(self, request):
        """Generate ageing analysis for current work orders"""
        analysis_date = timezone.now().date()
        
        # Get all open work orders
        open_wos = WorkOrder.objects.filter(
            status__in=['NEW', 'FOR AUDIT', 'NO COC']
        )
        
        created_count = 0
        
        for wo in open_wos:
            if wo.date_received_by_vc:
                age_days = (analysis_date - wo.date_received_by_vc).days
                age_months = age_days // 30
                
                # Determine age bracket
                if age_months <= 3:
                    age_bracket = '0-3'
                elif age_months <= 6:
                    age_bracket = '4-6'
                elif age_months <= 9:
                    age_bracket = '7-9'
                else:
                    age_bracket = '10+'
                
                AgeingAnalysis.objects.create(
                    analysis_date=analysis_date,
                    work_order=wo,
                    age_bracket=age_bracket,
                    age_in_days=age_days,
                    age_in_months=age_months,
                    supervisor=wo.supervisor_full_name,
                    crew=wo.assigned_crew,
                    status_at_analysis=wo.status
                )
                created_count += 1
        
        return Response({
            'message': f'Generated ageing analysis for {created_count} work orders',
            'analysis_date': analysis_date
        })


# ============================================
# BACKJOB MONITORING VIEWSETS
# ============================================

class BackjobMonitoringViewSet(viewsets.ModelViewSet):
    queryset = BackjobMonitoring.objects.all()
    serializer_class = BackjobMonitoringSerializer
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['status', 'work_order', 'assigned_to', 'is_overdue']
    search_fields = ['issue_description', 'work_order__wo_no']
    ordering = ['-reported_date']
    
    @action(detail=False, methods=['get'])
    def pending_backjobs(self, request):
        """Get all pending backjobs"""
        pending = BackjobMonitoring.objects.filter(
            status__in=['PENDING', 'IN_PROGRESS']
        ).order_by('-days_pending')
        
        serializer = self.get_serializer(pending, many=True)
        return Response(serializer.data)
    
    @action(detail=False, methods=['get'])
    def overdue_backjobs(self, request):
        """Get all overdue backjobs"""
        overdue = BackjobMonitoring.objects.filter(
            is_overdue=True,
            status__in=['PENDING', 'IN_PROGRESS']
        ).order_by('-days_pending')
        
        serializer = self.get_serializer(overdue, many=True)
        return Response(serializer.data)
    
    @action(detail=True, methods=['post'])
    def resolve(self, request, pk=None):
        """Mark a backjob as resolved"""
        backjob = self.get_object()
        
        backjob.status = 'RESOLVED'
        backjob.actual_resolution_date = timezone.now().date()
        backjob.resolution_notes = request.data.get('resolution_notes', '')
        backjob.save()
        
        serializer = self.get_serializer(backjob)
        return Response(serializer.data)
    
    @action(detail=False, methods=['get'])
    def statistics(self, request):
        """Get backjob statistics"""
        stats = {
            'total': BackjobMonitoring.objects.count(),
            'pending': BackjobMonitoring.objects.filter(status='PENDING').count(),
            'in_progress': BackjobMonitoring.objects.filter(status='IN_PROGRESS').count(),
            'resolved': BackjobMonitoring.objects.filter(status='RESOLVED').count(),
            'overdue': BackjobMonitoring.objects.filter(is_overdue=True).count(),
            'avg_resolution_days': BackjobMonitoring.objects.filter(
                status='RESOLVED'
            ).aggregate(avg=Avg('days_pending'))['avg']
        }
        
        return Response(stats)
    
    
from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from datetime import datetime, timedelta
from dateutil.relativedelta import relativedelta
from .models import KPISnapshot, KPITarget
from .serializers import KPISnapshotSerializer, KPITargetSerializer, KPIDashboardSerializer
from .kpi_service import KPICalculationService

class KPISnapshotViewSet(viewsets.ModelViewSet):
    queryset = KPISnapshot.objects.all()
    serializer_class = KPISnapshotSerializer
    
    @action(detail=False, methods=['post'])
    def calculate_and_save(self, request):
        """Calculate KPIs for a period and save snapshots"""
        period_start = request.data.get('period_start')
        period_end = request.data.get('period_end')
        
        if not period_start or not period_end:
            return Response(
                {'error': 'period_start and period_end are required'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            period_start = datetime.strptime(period_start, '%Y-%m-%d').date()
            period_end = datetime.strptime(period_end, '%Y-%m-%d').date()
        except ValueError:
            return Response(
                {'error': 'Invalid date format. Use YYYY-MM-DD'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Calculate all KPIs
        kpis = KPICalculationService.calculate_all_kpis(period_start, period_end)
        
        # Save snapshots
        snapshots_created = []
        
        kpi_mapping = {
            'ccti': 'CCTI',
            'pca_conversion': 'PCA_CONVERSION',
            'ageing_completion': 'AGEING_COMPLETION',
            'termination_apt': 'TERM_APT',
            'prdi': 'PRDI',
            'cost_settlement': 'COST_SETTLEMENT',
            'quality_index': 'QUALITY_INDEX',
            'capability_utilization': 'CAPABILITY_UTIL'
        }
        
        for key, kpi_type in kpi_mapping.items():
            kpi_data = kpis.get(key, {})
            
            # Get or create target
            target = KPITarget.objects.filter(
                kpi_type=kpi_type,
                period_start__lte=period_start,
                period_end__gte=period_end,
                is_active=True
            ).first()
            
            snapshot = KPISnapshot.objects.create(
                kpi_type=kpi_type,
                period_start=period_start,
                period_end=period_end,
                kpi_value=kpi_data.get('value', 0),
                target_value=target.target_value if target else None,
                numerator=kpi_data.get('numerator'),
                denominator=kpi_data.get('denominator'),
                sample_size=kpi_data.get('sample_size'),
                calculation_details=kpi_data.get('details', {}),
                calculated_by=request.user
            )
            
            snapshots_created.append(snapshot)
        
        serializer = self.get_serializer(snapshots_created, many=True)
        return Response(serializer.data, status=status.HTTP_201_CREATED)


class KPITargetViewSet(viewsets.ModelViewSet):
    queryset = KPITarget.objects.all()
    serializer_class = KPITargetSerializer


class KPIDashboardViewSet(viewsets.ViewSet):
    """
    ViewSet for KPI Dashboard data
    """
    
    @action(detail=False, methods=['get'])
    def current_period(self, request):
        """Get KPI dashboard for current period"""
        # Default to current month
        today = datetime.now().date()
        period_start = today.replace(day=1)
        period_end = (period_start + relativedelta(months=1)) - timedelta(days=1)
        
        # Allow custom period
        if request.query_params.get('period_start'):
            period_start = datetime.strptime(
                request.query_params['period_start'], '%Y-%m-%d'
            ).date()
        
        if request.query_params.get('period_end'):
            period_end = datetime.strptime(
                request.query_params['period_end'], '%Y-%m-%d'
            ).date()
        
        # Calculate all KPIs
        kpis = KPICalculationService.calculate_all_kpis(period_start, period_end)
        
        # Get targets
        targets = {}
        for kpi_type_key in ['CCTI', 'PCA_CONVERSION', 'AGEING_COMPLETION', 'TERM_APT', 
                            'PRDI', 'COST_SETTLEMENT', 'QUALITY_INDEX', 'CAPABILITY_UTIL']:
            target = KPITarget.objects.filter(
                kpi_type=kpi_type_key,
                period_start__lte=period_start,
                period_end__gte=period_end,
                is_active=True
            ).first()
            
            targets[kpi_type_key] = {
                'value': float(target.target_value) if target else None,
                'green': float(target.threshold_green) if target and target.threshold_green else None,
                'yellow': float(target.threshold_yellow) if target and target.threshold_yellow else None,
                'red': float(target.threshold_red) if target and target.threshold_red else None
            }
        
        # Add targets to KPI data
        for key, kpi_type in [('ccti', 'CCTI'), ('pca_conversion', 'PCA_CONVERSION'),
                              ('ageing_completion', 'AGEING_COMPLETION'), ('termination_apt', 'TERM_APT'),
                              ('prdi', 'PRDI'), ('cost_settlement', 'COST_SETTLEMENT'),
                              ('quality_index', 'QUALITY_INDEX'), ('capability_utilization', 'CAPABILITY_UTIL')]:
            if key in kpis:
                kpis[key]['target'] = targets.get(kpi_type, {})
        
        # Get historical trends (last 6 months)
        historical_trends = self._get_historical_trends(period_start)
        
        # Prepare chart data
        chart_data = self._prepare_chart_data(kpis, historical_trends)
        
        dashboard_data = {
            'period_start': period_start,
            'period_end': period_end,
            'total_kpis': 8,
            'ccti': kpis.get('ccti', {}),
            'pca_conversion': kpis.get('pca_conversion', {}),
            'ageing_completion': kpis.get('ageing_completion', {}),
            'pai_adherence': {},  # Placeholder - needs SAIDI data
            'termination_apt': kpis.get('termination_apt', {}),
            'termination_resolution': {},  # Placeholder
            'prdi': kpis.get('prdi', {}),
            'cost_settlement': kpis.get('cost_settlement', {}),
            'quality_index': kpis.get('quality_index', {}),
            'capability_utilization': kpis.get('capability_utilization', {}),
            'historical_trends': historical_trends,
            'chart_data': chart_data
        }
        
        serializer = KPIDashboardSerializer(dashboard_data)
        return Response(serializer.data)
    
    def _get_historical_trends(self, current_period_start, months=6):
        """Get historical KPI trends"""
        trends = {}
        
        for i in range(months):
            month_start = current_period_start - relativedelta(months=i)
            month_end = (month_start + relativedelta(months=1)) - timedelta(days=1)
            
            snapshots = KPISnapshot.objects.filter(
                period_start__gte=month_start,
                period_end__lte=month_end
            )
            
            month_key = month_start.strftime('%Y-%m')
            trends[month_key] = {}
            
            for snapshot in snapshots:
                trends[month_key][snapshot.kpi_type] = {
                    'value': float(snapshot.kpi_value),
                    'target': float(snapshot.target_value) if snapshot.target_value else None
                }
        
        return trends
    
    def _prepare_chart_data(self, current_kpis, historical_trends):
        """Prepare data formatted for charts"""
        return {
            'kpi_summary': {
                'labels': ['CCTI', 'PCA Conv.', 'Quality', 'Cost Settlement'],
                'current': [
                    current_kpis.get('ccti', {}).get('value', 0),
                    current_kpis.get('pca_conversion', {}).get('value', 0),
                    current_kpis.get('quality_index', {}).get('value', 0),
                    current_kpis.get('cost_settlement', {}).get('value', 0)
                ],
                'target': [
                    current_kpis.get('ccti', {}).get('target', {}).get('value', 0),
                    current_kpis.get('pca_conversion', {}).get('target', {}).get('value', 0),
                    current_kpis.get('quality_index', {}).get('target', {}).get('value', 0),
                    current_kpis.get('cost_settlement', {}).get('target', {}).get('value', 0)
                ]
            },
            'trend_data': historical_trends
        }
    

    @action(detail=False, methods=['get'])
    def kpi_detail(self, request):
        """Get detailed breakdown for a specific KPI"""
        kpi_type = request.query_params.get('type')
        period_start = request.query_params.get('period_start')
        period_end = request.query_params.get('period_end')

        if not all([kpi_type, period_start, period_end]):
            return Response(
                {'error': 'type, period_start, and period_end are required'},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            period_start = datetime.strptime(period_start, '%Y-%m-%d').date()
            period_end = datetime.strptime(period_end, '%Y-%m-%d').date()
        except ValueError:
            return Response(
                {'error': 'Invalid date format. Use YYYY-MM-DD'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Calculate specific KPI
        kpi_methods = {
            'CCTI': KPICalculationService.calculate_ccti,
            'PCA_CONVERSION': KPICalculationService.calculate_pca_conversion_rate,
            'AGEING_COMPLETION': KPICalculationService.calculate_ageing_pca_completion,
            'TERM_APT': KPICalculationService.calculate_termination_apt,
            'PRDI': KPICalculationService.calculate_prdi,
            'COST_SETTLEMENT': KPICalculationService.calculate_cost_settlement,
            'QUALITY_INDEX': KPICalculationService.calculate_quality_index,
            'CAPABILITY_UTIL': KPICalculationService.calculate_capability_utilization
        }

        if kpi_type not in kpi_methods:
            return Response(
                {'error': f'Invalid KPI type. Choose from: {", ".join(kpi_methods.keys())}'},
                status=status.HTTP_400_BAD_REQUEST
            )

        kpi_data = kpi_methods[kpi_type](period_start, period_end)

        return Response({
            'kpi_type': kpi_type,
            'period_start': period_start,
            'period_end': period_end,
            'data': kpi_data
        })



from django.core.mail import EmailMultiAlternatives
from django.conf import settings
from django.http import HttpResponse

def test_email_view(request):
    recipient_list = ["recipient@example.com"]

    # Example for "Good" notification
    if send_custom_email(
        subject="✅ Report Status: GOOD",
        message="All systems are running smoothly.",
        recipient_list=recipient_list,
        status="good"
    ):
        return HttpResponse("Good email sent successfully!")

    # Example for "Bad" notification
    if send_custom_email(
        subject="⚠️ Report Status: BAD",
        message="There are issues detected in the system.",
        recipient_list=recipient_list,
        status="bad"
    ):
        return HttpResponse("Bad email sent successfully!")

    return HttpResponse("Failed to send email.")

def send_custom_email(subject, message, recipient_list, status="good", from_email=None, fail_silently=False):
    """
    Sends a styled HTML email with a notification-like report.
    Status can be "good" or "bad".
    """
    if from_email is None:
        from_email = settings.DEFAULT_FROM_EMAIL

    # Define color schemes for good and bad notifications
    colors = {
        "good": {
            "bg": "#e6ffed",
            "border": "#4CAF50",
            "text": "#2e7d32",
            "emoji": "✅"
        },
        "bad": {
            "bg": "#ffe6e6",
            "border": "#f44336",
            "text": "#b71c1c",
            "emoji": "⚠️"
        }
    }

    color = colors.get(status, colors["good"])

    html_content = f"""
    <html>
    <body style="font-family: Arial, sans-serif; background-color: #f5f5f5; padding: 20px;">
        <div style="
            max-width: 600px;
            margin: auto;
            background-color: {color['bg']};
            border-left: 6px solid {color['border']};
            padding: 20px;
            border-radius: 8px;
            box-shadow: 0 2px 4px rgba(0,0,0,0.1);
        ">
            <h2 style="color: {color['text']}; margin-top: 0;">{color['emoji']} Notification Report</h2>
            <p style="color: #333; font-size: 16px;">{message}</p>
            <hr style="border: none; border-top: 1px solid #ddd; margin: 20px 0;">
            <p style="font-size: 12px; color: #666;">This is an automated message from your system.</p>
        </div>
    </body>
    </html>
    """

    try:
        email = EmailMultiAlternatives(subject, message, from_email, recipient_list)
        email.attach_alternative(html_content, "text/html")
        email.send(fail_silently=fail_silently)
        return True
    except Exception as e:
        print(f"Error sending email: {e}")
        return False

    



from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import AllowAny, IsAuthenticated
from rest_framework.response import Response
from rest_framework import status
from .kpi_email_service import KPIEmailService


@api_view(['POST'])
def send_daily_kpi_email(request):
    """
    Manually trigger daily KPI email
    Can be called from a cron job or scheduled task
    """
    recipient_email = request.data.get('recipient_email', KPIEmailService.DEFAULT_RECIPIENT)
    
    result = KPIEmailService.send_daily_kpi_email(recipient_email)
    
    if result['success']:
        return Response(result, status=status.HTTP_200_OK)
    else:
        return Response(result, status=status.HTTP_400_BAD_REQUEST)


@api_view(['GET'])
@permission_classes([AllowAny])
def check_daily_email_status(request):
    """Check if daily email has been sent today"""
    from datetime import date
    
    log = EmailNotificationLog.objects.filter(
        notification_type='KPI_DAILY',
        notification_date=date.today()
    ).first()
    
    if log:
        return Response({
            'sent_today': True,
            'sent_at': log.sent_at,
            'status': log.status,
            'recipient': log.recipient_email
        })
    else:
        return Response({
            'sent_today': False,
            'message': 'No email sent today yet'
        })


@api_view(['POST'])
@permission_classes([AllowAny])
def auto_send_daily_email(request):
    """
    Auto-send endpoint - can be called by external cron job
    This endpoint will check if email was sent and send if not
    
    Usage: Call this endpoint once per day via cron job or task scheduler
    Example: curl -X POST http://your-domain.com/api/auto-send-daily-email/
    """
    result = KPIEmailService.send_daily_kpi_email()
    return Response(result)

from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import AllowAny, IsAuthenticated
from rest_framework.response import Response
from rest_framework import status, viewsets
from rest_framework.decorators import action
from .daily_action_email_service import DailyActionEmailService
from .email_notification_service import EmailNotificationService
from .models import User, EmailNotificationLog
from datetime import date


@api_view(['POST'])
@permission_classes([AllowAny])
def send_my_daily_email(request):
    """Send daily action email to the authenticated user"""
    
    user = request.user
    
    result = DailyActionEmailService.send_daily_action_email(user)
    
    if result['success']:
        return Response(result, status=status.HTTP_200_OK)
    else:
        return Response(result, status=status.HTTP_400_BAD_REQUEST)

@api_view(['GET'])
def notify_new_work_order_view(request, wo_id):
    work_order = get_object_or_404(WorkOrder, pk=wo_id)

    EmailNotificationService.notify_new_work_order(work_order)

    return Response({
        "success": True,
        "message": f"Email notifications sent for WO {work_order.wo_no}"
    })


@api_view(['GET'])
@permission_classes([AllowAny])  # Protect this in production
def send_notifsss(request):
    """
    Send daily action emails to all active users
    Should be called by cron job or scheduler
    """
    
    results = EmailNotificationService()
    
    return Response(results, status=status.HTTP_200_OK)

@api_view(['POST'])
@permission_classes([AllowAny])  # Protect this in production
def send_daily_emails_to_all(request):
    """
    Send daily action emails to all active users
    Should be called by cron job or scheduler
    """
    
    results = DailyActionEmailService.send_daily_emails_to_all_users()
    
    return Response(results, status=status.HTTP_200_OK)


@api_view(['POST'])
@permission_classes([AllowAny])
def send_daily_email_to_user(request, user_id):
    """Send daily action email to a specific user (admin only)"""
    
    if not request.user.is_staff:
        return Response(
            {'error': 'Admin permission required'},
            status=status.HTTP_403_FORBIDDEN
        )
    
    try:
        user = User.objects.get(user_id=user_id)
        result = DailyActionEmailService.send_daily_action_email(user)
        
        if result['success']:
            return Response(result, status=status.HTTP_200_OK)
        else:
            return Response(result, status=status.HTTP_400_BAD_REQUEST)
            
    except User.DoesNotExist:
        return Response(
            {'error': 'User not found'},
            status=status.HTTP_404_NOT_FOUND
        )


@api_view(['GET'])
@permission_classes([AllowAny])
def check_my_actions(request):
    """Get today's pending actions for the authenticated user (without sending email)"""
    
    user = request.user
    role_code = user.role.role_name if user.role else 'USER'
    
    pending_actions = DailyActionEmailService.get_pending_actions_for_role(role_code)
    capabilities = DailyActionEmailService.get_role_capabilities(role_code)
    
    return Response({
        'user': {
            'username': user.username,
            'email': user.email,
            'role': role_code
        },
        'pending_actions': pending_actions,
        'capabilities': capabilities
    })


@api_view(['GET'])
@permission_classes([AllowAny])
def check_daily_email_logs(request):
    """Check email logs for today"""
    
    logs = EmailNotificationLog.objects.filter(
        notification_type='DAILY_ACTION',
        notification_date=date.today()
    ).values(
        'recipient_email', 'status', 'sent_at', 'error_message'
    )
    
    return Response({
        'date': date.today(),
        'total': logs.count(),
        'logs': list(logs)
    })


class DailyActionEmailViewSet(viewsets.ViewSet):
    """ViewSet for daily action email management"""
    
    @action(detail=False, methods=['post'])
    def send_to_me(self, request):
        """Send daily action email to authenticated user"""
        result = DailyActionEmailService.send_daily_action_email(request.user)
        return Response(result)
    
    @action(detail=False, methods=['post'])
    def send_to_all(self, request):
        """Send daily action emails to all users (admin only)"""
        if not request.user.is_staff:
            return Response(
                {'error': 'Admin permission required'},
                status=status.HTTP_403_FORBIDDEN
            )
        
        results = DailyActionEmailService.send_daily_emails_to_all_users()
        return Response(results)
    
    @action(detail=False, methods=['get'])
    def my_actions(self, request):
        """Get pending actions for authenticated user"""
        user = request.user
        role_code = user.role.role_name if user.role else 'USER'
        
        pending_actions = DailyActionEmailService.get_pending_actions_for_role(role_code)
        capabilities = DailyActionEmailService.get_role_capabilities(role_code)
        
        return Response({
            'user': {
                'username': user.username,
                'role': role_code
            },
            'pending_actions': pending_actions,
            'capabilities': capabilities
        })
    
    @action(detail=False, methods=['get'])
    def email_logs(self, request):
        """Get email logs"""
        date_param = request.query_params.get('date', date.today())
        
        logs = EmailNotificationLog.objects.filter(
            notification_type='DAILY_ACTION',
            notification_date=date_param
        ).values(
            'recipient_email', 'status', 'sent_at', 'error_message'
        )
        
        return Response({
            'date': date_param,
            'total': logs.count(),
            'sent': logs.filter(status='SENT').count(),
            'failed': logs.filter(status='FAILED').count(),
            'logs': list(logs)
        })
    
    @action(detail=False, methods=['post'])
    def test_email(self, request):
        """Send test email to authenticated user"""
        user = request.user
        
        # Override email if provided
        test_email = request.data.get('email', user.email)
        
        # Create a temporary user object for testing
        from copy import copy
        test_user = copy(user)
        test_user.email = test_email
        
        result = DailyActionEmailService.send_daily_action_email(test_user)
        return Response(result)
    


from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from django.db.models import Count, Sum, Q
from datetime import date, timedelta


@api_view(['GET'])
@permission_classes([AllowAny])
def user_dashboard(request):
    """
    Get role-specific dashboard data
    """
    user = request.user
    role_name = user.role.role_name if user.role else None
    
    dashboard_data = {
        'user_info': {
            'name': user.get_full_name(),
            'role': role_name,
            'email': user.email
        }
    }
    
    # VENDOR REPRESENTATIVE Dashboard
    if role_name == 'Vendor Representative':
        vendor = Vendor.objects.filter(email=user.email).first()
        if vendor:
            dashboard_data['vendor_data'] = {
                'vendor_name': vendor.vendor_name,
                'active_projects': Project.objects.filter(
                    vendor=vendor,
                    status__status_name__in=['In Progress', 'Planning']
                ).count(),
                'completed_projects': Project.objects.filter(
                    vendor=vendor,
                    status__status_name='Completed'
                ).count(),
                'pending_documents': DocumentCompliance.objects.filter(
                    project__vendor=vendor,
                    is_submitted=False
                ).count(),
                'total_penalties': float(Penalty.objects.filter(
                    vendor=vendor
                ).exclude(penalty_status='Waived').aggregate(
                    total=Sum('penalty_amount')
                )['total'] or 0),
                'pending_invoices': Invoice.objects.filter(
                    vendor=vendor
                ).exclude(payment_status='Paid').count()
            }
    
    # CLERK Dashboard
    elif role_name == 'Clerk':
        dashboard_data['clerk_data'] = {
            'pending_uploads': DocumentCompliance.objects.filter(
                is_submitted=False
            ).count(),
            'documents_uploaded_today': ProjectDocument.objects.filter(
                uploaded_by=user,
                upload_date__date=date.today()
            ).count(),
            'pending_approvals': ProjectDocument.objects.filter(
                approval_status='Pending'
            ).count(),
            'overdue_documents': DocumentCompliance.objects.filter(
                is_overdue=True,
                is_submitted=False
            ).count()
        }
    
    # ENGINEERING AIDE Dashboard
    elif role_name == 'Engineering Aide':
        dashboard_data['aide_data'] = {
            'active_workflows': ProjectWorkflow.objects.filter(
                status='In Progress'
            ).count(),
            'pending_stages': ProjectWorkflow.objects.filter(
                completion_date__isnull=True
            ).count(),
            'document_compliance_rate': _calculate_compliance_rate(),
            'upcoming_deadlines': ProjectWorkflow.objects.filter(
                due_date__gte=date.today(),
                due_date__lte=date.today() + timedelta(days=7),
                completion_date__isnull=True
            ).count()
        }
    
    # QUALITY INSPECTOR Dashboard
    elif role_name == 'Quality Inspector':
        dashboard_data['qi_data'] = {
            'pending_inspections': QIInspection.objects.filter(
                assigned_qi=user,
                is_completed=False
            ).count(),
            'completed_today': QIInspection.objects.filter(
                assigned_qi=user,
                inspection_date=date.today(),
                is_completed=True
            ).count(),
            'daily_target': _get_qi_daily_target(user),
            'monthly_progress': _get_qi_monthly_progress(user),
            'overdue_inspections': QIInspection.objects.filter(
                assigned_qi=user,
                scheduled_date__lt=date.today(),
                is_completed=False
            ).count()
        }
    
    # ENGINEER Dashboard
    elif role_name == 'Engineer':
        dashboard_data['engineer_data'] = {
            'assigned_projects': Project.objects.filter(
                assigned_engineer=user
            ).exclude(status__status_name__in=['Completed', 'Cancelled']).count(),
            'pending_approvals': ProjectDocument.objects.filter(
                approval_status='Pending'
            ).count(),
            'sla_breaches': SLATracking.objects.filter(
                project__assigned_engineer=user,
                is_breached=True
            ).count(),
            'vendor_performance': _get_vendor_performance_summary()
        }
    
    # WO SUPERVISOR Dashboard
    elif role_name == 'WO Supervisor':
        dashboard_data['supervisor_data'] = {
            'total_projects': Project.objects.count(),
            'active_projects': Project.objects.exclude(
                status__status_name__in=['Completed', 'Cancelled']
            ).count(),
            'delayed_projects': Project.objects.filter(is_delayed=True).count(),
            'pending_penalties': Penalty.objects.filter(
                penalty_status='Draft'
            ).count(),
            'sla_breaches': SLATracking.objects.filter(
                is_breached=True,
                status='Breached'
            ).count(),
            'qi_workload': _get_qi_workload_summary(),
            'escalations': Escalation.objects.filter(status='Open').count()
        }
    
    # TEAM LEADER Dashboard
    elif role_name == 'Team Leader':
        dashboard_data['leader_data'] = {
            'organization_overview': {
                'total_projects': Project.objects.count(),
                'on_time_percentage': _calculate_on_time_percentage(),
                'avg_completion_days': _calculate_avg_completion_days(),
                'total_contract_value': float(Project.objects.aggregate(
                    total=Sum('contract_value')
                )['total'] or 0)
            },
            'pending_approvals': {
                'penalties': Penalty.objects.filter(penalty_status='Draft').count(),
                'documents': ProjectDocument.objects.filter(approval_status='Pending').count(),
                'projects': Project.objects.filter(status__status_name='Pending Approval').count()
            },
            'trends': _get_performance_trends(),
            'ai_suggestions': []  # Placeholder for AI suggestions
        }
    
    # SECTOR MANAGER Dashboard
    elif role_name == 'Sector Manager':
        dashboard_data['sector_manager_data'] = {
            'kpi_summary': _get_kpi_summary(),
            'sector_comparison': _get_sector_comparison(),
            'financial_overview': {
                'total_contract_value': float(Project.objects.aggregate(
                    total=Sum('contract_value')
                )['total'] or 0),
                'total_billed': float(Invoice.objects.filter(
                    payment_status='Paid'
                ).aggregate(total=Sum('net_amount'))['total'] or 0),
                'total_penalties': float(Penalty.objects.exclude(
                    penalty_status='Waived'
                ).aggregate(total=Sum('penalty_amount'))['total'] or 0)
            },
            'vendor_rankings': _get_vendor_rankings()
        }
    
    # SYSTEM ADMINISTRATOR Dashboard
    elif role_name == 'System Administrator':
        dashboard_data['admin_data'] = {
            'system_health': {
                'total_users': User.objects.count(),
                'active_users': User.objects.filter(is_active=True).count(),
                'active_sessions': UserSession.objects.filter(is_active=True).count()
            },
            'recent_activities': _get_recent_audit_logs(),
            'system_settings': SystemSetting.objects.count(),
            'backup_status': 'OK'  # Placeholder
        }
    
    return Response(dashboard_data)


# Helper functions
def _calculate_compliance_rate():
    total = DocumentCompliance.objects.count()
    if total == 0:
        return 0
    submitted = DocumentCompliance.objects.filter(is_submitted=True).count()
    return round((submitted / total) * 100, 2)


def _get_qi_daily_target(user):
    target = QIDailyTarget.objects.filter(
        qi_user=user,
        target_date=date.today()
    ).first()
    if target:
        return {
            'target': target.target_audits,
            'actual': target.actual_audits,
            'met': target.target_met
        }
    return {'target': 0, 'actual': 0, 'met': False}


def _get_qi_monthly_progress(user):
    month_start = date.today().replace(day=1)
    targets = QIDailyTarget.objects.filter(
        qi_user=user,
        target_date__gte=month_start
    )
    total_target = targets.aggregate(total=Sum('target_audits'))['total'] or 0
    total_actual = targets.aggregate(total=Sum('actual_audits'))['total'] or 0
    return {
        'total_target': total_target,
        'total_actual': total_actual,
        'percentage': round((total_actual / total_target * 100) if total_target > 0 else 0, 2)
    }


def _get_vendor_performance_summary():
    return Vendor.objects.filter(is_active=True).annotate(
        total_projects=Count('projects'),
        delayed=Count('projects', filter=Q(projects__is_delayed=True))
    ).values('vendor_name', 'compliance_score', 'total_projects', 'delayed')[:5]


def _get_qi_workload_summary():
    return QIInspection.objects.filter(
        is_completed=False
    ).values('assigned_qi__first_name', 'assigned_qi__last_name').annotate(
        pending_count=Count('id')
    )


def _calculate_on_time_percentage():
    total = Project.objects.filter(
        status__status_name='Completed'
    ).count()
    if total == 0:
        return 0
    on_time = Project.objects.filter(
        status__status_name='Completed',
        is_delayed=False
    ).count()
    return round((on_time / total) * 100, 2)


def _calculate_avg_completion_days():
    avg = Project.objects.filter(
        status__status_name='Completed',
        delay_days__isnull=False
    ).aggregate(avg=Avg('delay_days'))['avg']
    return round(float(avg or 0), 2)


def _get_performance_trends():
    # Last 6 months trend
    six_months_ago = date.today() - timedelta(days=180)
    monthly_data = Project.objects.filter(
        created_at__gte=six_months_ago
    ).annotate(
        month=TruncMonth('created_at')
    ).values('month').annotate(
        total=Count('project_id'),
        completed=Count('project_id', filter=Q(status__status_name='Completed'))
    ).order_by('month')
    return list(monthly_data)


def _get_kpi_summary():
    # Return recent KPI snapshots
    return KPISnapshot.objects.all().order_by('-period_end')[:8].values(
        'kpi_type', 'kpi_value', 'target_value', 'period_end'
    )


def _get_sector_comparison():
    return Sector.objects.annotate(
        total_projects=Count('projects'),
        completed=Count('projects', filter=Q(projects__status__status_name='Completed')),
        delayed=Count('projects', filter=Q(projects__is_delayed=True))
    ).values('sector_name', 'total_projects', 'completed', 'delayed')


def _get_vendor_rankings():
    return Vendor.objects.filter(is_active=True).order_by('-compliance_score')[:10].values(
        'vendor_name', 'compliance_score'
    )


def _get_recent_audit_logs():
    return SystemAuditLog.objects.all().order_by('-created_at')[:20].values(
        'user__username', 'action_type', 'status', 'created_at'
    )
    

class VendorPortalViewSet(viewsets.ViewSet):
    """
    Vendor Representative Portal - specific endpoints for vendors
    """
    permission_classes = [AllowAny]
    
    @action(detail=False, methods=['get'])
    def my_projects(self, request):
        """Get projects for current vendor"""
        vendor = Vendor.objects.filter(email=request.user.email).first()
        if not vendor:
            return Response({'error': 'Vendor not found'}, status=404)
        
        projects = Project.objects.filter(vendor=vendor).select_related(
            'status', 'sector', 'assigned_engineer'
        )
        serializer = ProjectListSerializer(projects, many=True)
        return Response(serializer.data)
    
    @action(detail=False, methods=['get'])
    def pending_documents(self, request):
        """Get pending document submissions"""
        vendor = Vendor.objects.filter(email=request.user.email).first()
        if not vendor:
            return Response({'error': 'Vendor not found'}, status=404)
        
        pending = DocumentCompliance.objects.filter(
            project__vendor=vendor,
            is_submitted=False
        ).select_related('project', 'doc_type')
        
        serializer = DocumentComplianceSerializer(pending, many=True)
        return Response(serializer.data)
    
    @action(detail=False, methods=['post'])
    def upload_document(self, request):
        """Upload compliance document"""
        vendor = Vendor.objects.filter(email=request.user.email).first()
        if not vendor:
            return Response({'error': 'Vendor not found'}, status=404)
        
        project_id = request.data.get('project_id')
        project = Project.objects.filter(project_id=project_id, vendor=vendor).first()
        
        if not project:
            return Response({'error': 'Project not found'}, status=404)
        
        serializer = ProjectDocumentSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save(
                project=project,
                uploaded_by=request.user
            )
            return Response(serializer.data, status=201)
        return Response(serializer.errors, status=400)
    
    @action(detail=False, methods=['get'])
    def payment_summary(self, request):
        """Get payment and billing summary"""
        vendor = Vendor.objects.filter(email=request.user.email).first()
        if not vendor:
            return Response({'error': 'Vendor not found'}, status=404)
        
        summary = {
            'total_invoiced': float(Invoice.objects.filter(
                vendor=vendor
            ).aggregate(total=Sum('invoice_amount'))['total'] or 0),
            'total_paid': float(Invoice.objects.filter(
                vendor=vendor,
                payment_status='Paid'
            ).aggregate(total=Sum('net_amount'))['total'] or 0),
            'outstanding': float(Invoice.objects.filter(
                vendor=vendor
            ).exclude(payment_status='Paid').aggregate(
                total=Sum('net_amount')
            )['total'] or 0),
            'penalties': float(Penalty.objects.filter(
                vendor=vendor
            ).exclude(penalty_status='Waived').aggregate(
                total=Sum('penalty_amount')
            )['total'] or 0)
        }
        return Response(summary)
    
    @action(detail=False, methods=['post'])
    def submit_dispute(self, request):
        """Submit a dispute"""
        vendor = Vendor.objects.filter(email=request.user.email).first()
        if not vendor:
            return Response({'error': 'Vendor not found'}, status=404)
        
        serializer = VendorDisputeSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save(vendor=vendor)
            return Response(serializer.data, status=201)
        return Response(serializer.errors, status=400)



class QIMobileViewSet(viewsets.ViewSet):
    """
    Quality Inspector Mobile-Friendly Endpoints
    """
    permission_classes = [AllowAny]
    
    @action(detail=False, methods=['get'])
    def today_schedule(self, request):
        """Get today's inspection schedule"""
        inspections = QIInspection.objects.filter(
            assigned_qi=request.user,
            scheduled_date=date.today()
        ).select_related('project', 'inspection_type')
        
        serializer = QIInspectionSerializer(inspections, many=True)
        return Response(serializer.data)
    
    @action(detail=False, methods=['post'])
    def complete_inspection(self, request, pk=None):
        """Complete an inspection with mobile data"""
        inspection_id = request.data.get('inspection_id')
        
        try:
            inspection = QIInspection.objects.get(
                id=inspection_id,
                assigned_qi=request.user
            )
            
            inspection.is_completed = True
            inspection.inspection_date = date.today()
            inspection.inspection_result = request.data.get('result')
            inspection.findings = request.data.get('findings')
            inspection.location_coordinates = request.data.get('coordinates')
            inspection.save()
            
            # Update daily target
            target, _ = QIDailyTarget.objects.get_or_create(
                qi_user=request.user,
                target_date=date.today()
            )
            target.actual_audits += 1
            target.save()
            
            return Response({'status': 'completed'})
        except QIInspection.DoesNotExist:
            return Response({'error': 'Inspection not found'}, status=404)
    
    @action(detail=False, methods=['get'])
    def daily_progress(self, request):
        """Get today's progress"""
        target = QIDailyTarget.objects.filter(
            qi_user=request.user,
            target_date=date.today()
        ).first()
        
        if target:
            return Response({
                'target': target.target_audits,
                'actual': target.actual_audits,
                'remaining': target.target_audits - target.actual_audits,
                'percentage': round((target.actual_audits / target.target_audits * 100) if target.target_audits > 0 else 0, 2),
                'target_met': target.target_met
            })
        return Response({
            'target': 0,
            'actual': 0,
            'remaining': 0,
            'percentage': 0,
            'target_met': False
        })
    
    @action(detail=False, methods=['post'])
    def log_missed_target(self, request):
        """Log reason for missing target"""
        target = QIDailyTarget.objects.filter(
            qi_user=request.user,
            target_date=date.today()
        ).first()
        
        if target:
            target.reason_not_met = request.data.get('reason')
            target.reason_category = request.data.get('category')
            target.save()
            return Response({'status': 'saved'})
        else:
            return Response({'error': 'No target found for today'}, status=404)


class ClerkViewSet(viewsets.ViewSet):
    """
    Clerk Portal - Document management and basic communications
    """
    permission_classes = [AllowAny]
    
    @action(detail=False, methods=['get'])
    def pending_documents(self, request):
        """Get documents pending upload or approval"""
        pending_upload = DocumentCompliance.objects.filter(
            is_submitted=False
        ).select_related('project', 'doc_type').order_by('due_date')
        
        pending_approval = ProjectDocument.objects.filter(
            approval_status='Pending'
        ).select_related('project', 'doc_type').order_by('-upload_date')
        
        return Response({
            'pending_upload': DocumentComplianceSerializer(pending_upload, many=True).data,
            'pending_approval': ProjectDocumentSerializer(pending_approval, many=True).data,
            'stats': {
                'total_pending_upload': pending_upload.count(),
                'total_pending_approval': pending_approval.count(),
                'overdue': DocumentCompliance.objects.filter(
                    is_overdue=True,
                    is_submitted=False
                ).count()
            }
        })
    
    @action(detail=False, methods=['post'])
    def bulk_upload_documents(self, request):
        """Bulk upload documents for multiple projects"""
        documents_data = request.data.get('documents', [])
        uploaded = []
        errors = []
        
        for doc_data in documents_data:
            try:
                project_id = doc_data.get('project_id')
                project = Project.objects.get(project_id=project_id)
                
                document = ProjectDocument.objects.create(
                    project=project,
                    doc_type_id=doc_data.get('doc_type_id'),
                    document_name=doc_data.get('document_name'),
                    document_path=doc_data.get('document_path'),
                    uploaded_by=request.user
                )
                uploaded.append(document.id)
                
                # Update compliance status
                DocumentCompliance.objects.filter(
                    project=project,
                    doc_type_id=doc_data.get('doc_type_id')
                ).update(
                    is_submitted=True,
                    submission_date=timezone.now()
                )
                
            except Exception as e:
                errors.append({
                    'project_id': doc_data.get('project_id'),
                    'error': str(e)
                })
        
        return Response({
            'uploaded_count': len(uploaded),
            'uploaded_ids': uploaded,
            'errors': errors
        })
    
    @action(detail=False, methods=['post'])
    def send_reminder(self, request):
        """Send document submission reminder to vendor"""
        project_id = request.data.get('project_id')
        doc_type_id = request.data.get('doc_type_id')
        message = request.data.get('message', '')
        
        try:
            project = Project.objects.get(project_id=project_id)
            doc_type = DocumentType.objects.get(doc_type_id=doc_type_id)
            
            # Create notification
            notification = Notification.objects.create(
                recipient_user=None,
                recipient_email=project.vendor.email if project.vendor else None,
                notification_type='Email',
                subject=f'Document Reminder: {doc_type.doc_type_name}',
                message=message or f'Please submit {doc_type.doc_type_name} for project {project.project_code}',
                related_project=project,
                status='Pending'
            )
            
            # Log the action
            ChangeLog.objects.create(
                table_name='notifications',
                record_id=notification.id,
                change_type='INSERT',
                changed_by=request.user,
                change_reason=f'Reminder sent by clerk for {doc_type.doc_type_name}'
            )
            
            return Response({
                'status': 'sent',
                'notification_id': notification.id
            })
            
        except Project.DoesNotExist:
            return Response({'error': 'Project not found'}, status=404)
        except DocumentType.DoesNotExist:
            return Response({'error': 'Document type not found'}, status=404)
    
    @action(detail=False, methods=['get'])
    def my_upload_history(self, request):
        """Get clerk's document upload history"""
        days = int(request.query_params.get('days', 30))
        start_date = date.today() - timedelta(days=days)
        
        uploads = ProjectDocument.objects.filter(
            uploaded_by=request.user,
            upload_date__gte=start_date
        ).select_related('project', 'doc_type').order_by('-upload_date')
        
        # Statistics
        stats = {
            'total_uploaded': uploads.count(),
            'approved': uploads.filter(approval_status='Approved').count(),
            'pending': uploads.filter(approval_status='Pending').count(),
            'rejected': uploads.filter(approval_status='Rejected').count(),
            'by_type': uploads.values('doc_type__doc_type_name').annotate(
                count=Count('id')
            ).order_by('-count')
        }
        
        return Response({
            'uploads': ProjectDocumentSerializer(uploads[:50], many=True).data,
            'stats': stats
        })
    
    @action(detail=False, methods=['post'])
    def schedule_reminders(self, request):
        """Schedule automatic reminders for upcoming deadlines"""
        days_ahead = int(request.data.get('days_ahead', 7))
        target_date = date.today() + timedelta(days=days_ahead)
        
        upcoming = DocumentCompliance.objects.filter(
            due_date__lte=target_date,
            due_date__gte=date.today(),
            is_submitted=False
        ).select_related('project', 'doc_type')
        
        scheduled = []
        for doc_compliance in upcoming:
            if doc_compliance.project.vendor:
                notification = Notification.objects.create(
                    recipient_email=doc_compliance.project.vendor.email,
                    notification_type='Email',
                    subject=f'Upcoming Deadline: {doc_compliance.doc_type.doc_type_name}',
                    message=f'Document {doc_compliance.doc_type.doc_type_name} is due on {doc_compliance.due_date}',
                    related_project=doc_compliance.project,
                    status='Pending'
                )
                scheduled.append(notification.id)
        
        return Response({
            'scheduled_count': len(scheduled),
            'notification_ids': scheduled
        })
    
    @action(detail=False, methods=['get'])
    def missing_documents_report(self, request):
        """Generate report of missing documents"""
        vendor_id = request.query_params.get('vendor_id')
        sector_id = request.query_params.get('sector_id')
        
        missing = DocumentCompliance.objects.filter(
            is_submitted=False
        ).select_related('project', 'doc_type')
        
        if vendor_id:
            missing = missing.filter(project__vendor_id=vendor_id)
        if sector_id:
            missing = missing.filter(project__sector_id=sector_id)
        
        # Group by vendor
        by_vendor = {}
        for doc in missing:
            vendor_name = doc.project.vendor.vendor_name if doc.project.vendor else 'Unknown'
            if vendor_name not in by_vendor:
                by_vendor[vendor_name] = []
            by_vendor[vendor_name].append({
                'project_code': doc.project.project_code,
                'document_type': doc.doc_type.doc_type_name,
                'due_date': doc.due_date,
                'is_overdue': doc.is_overdue,
                'overdue_days': doc.overdue_days
            })
        
        return Response({
            'total_missing': missing.count(),
            'overdue_count': missing.filter(is_overdue=True).count(),
            'by_vendor': by_vendor
        })
        
    @action(detail=False, methods=['get'], url_path='coc-checklist')
    def coc_checklist(self, request):
        """Get work orders needing COC review/processing"""
        
        # Filter work orders that are energized but need COC processing
        queryset = WorkOrder.objects.all(
        ).filter(
            status__in=['INPRG', 'SCHED']
        ).filter(
            models.Q(date_received_by_vc__isnull=False) &
            (
                models.Q(date_sched__isnull=True) |
                models.Q(date_received_jacket_ps__isnull=True)
            )
        ).order_by('-date_received_by_vc')
        
        # Apply filters
        status = request.query_params.get('status')
        vendor_id = request.query_params.get('vendor')
        crew = request.query_params.get('crew')
        needs_attention = request.query_params.get('needs_attention')
        
        if status:
            queryset = queryset.filter(status=status)
        if vendor_id:
            queryset = queryset.filter(vendor_id=vendor_id)
        if crew:
            queryset = queryset.filter(assigned_crew=crew)
        if needs_attention == 'true':
            # Filter items needing immediate attention
            from django.utils import timezone
            seven_days_ago = timezone.now().date() - timezone.timedelta(days=7)
            three_days_ago = timezone.now().date() - timezone.timedelta(days=3)
            
            queryset = queryset.filter(
                models.Q(date_received_by_vc__lte=seven_days_ago, date_sched__isnull=True) |
                models.Q(date_sched__lte=three_days_ago, date_received_jacket_ps__isnull=True)
            )
        
        serializer = COCChecklistSerializer(queryset, many=True)
        
        # Calculate statistics
        stats = {
            'total': queryset.count(),
            'awaiting_coc': queryset.filter(date_sched__isnull=True).count(),
            'awaiting_audit': queryset.filter(
                date_sched__isnull=False,
                date_received_jacket_ps__isnull=True
            ).count(),
            'needs_attention': sum(1 for item in serializer.data if item['needs_attention']),
        }
        
        return Response({
            'results': serializer.data,
            'stats': stats
        })
    
    @action(detail=True, methods=['post'], url_path='mark-coc-received')
    def mark_coc_received(self, request, pk=None):
        """Mark that COC has been received for a work order"""
        try:
            work_order = WorkOrder.objects.get(id=pk)
            
            date_received = request.data.get('date_sched')
            remarks = request.data.get('clerk_remarks', '')
            
            if date_received:
                from datetime import datetime
                work_order.date_sched = datetime.strptime(date_received, '%Y-%m-%d').date()
            else:
                from django.utils import timezone
                work_order.date_sched = timezone.now().date()
            
            if remarks:
                work_order.clerk_remarks = remarks
            
            work_order.save()
            
            serializer = COCChecklistSerializer(work_order)
            return Response({
                'message': 'COC receipt date marked successfully',
                'data': serializer.data
            })
        except WorkOrder.DoesNotExist:
            return Response(
                {'error': 'Work order not found'},
                status=status.HTTP_404_NOT_FOUND
            )
        except Exception as e:
            return Response(
                {'error': str(e)},
                status=status.HTTP_400_BAD_REQUEST
            )
    
    @action(detail=True, methods=['post'], url_path='send-for-audit')
    def send_for_audit(self, request, pk=None):
        """Mark work order as sent for audit"""
        try:
            work_order = WorkOrder.objects.get(id=pk)
            
            if not work_order.date_sched:
                return Response(
                    {'error': 'Cannot send for audit without COC received date'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            date_received_jacket_ps = request.data.get('date_received_jacket_ps')
            remarks = request.data.get('clerk_remarks', '')
            
            if date_received_jacket_ps:
                from datetime import datetime
                work_order.date_received_jacket_ps = datetime.strptime(date_received_jacket_ps, '%Y-%m-%d').date()
            else:
                from django.utils import timezone
                work_order.date_received_jacket_ps = timezone.now().date()
            
            work_order.status = 'FOR AUDIT'
            
            if remarks:
                if work_order.clerk_remarks:
                    work_order.clerk_remarks += f"\n{remarks}"
                else:
                    work_order.clerk_remarks = remarks
            
            work_order.save()
            
            serializer = COCChecklistSerializer(work_order)
            return Response({
                'message': 'Work order sent for audit successfully',
                'data': serializer.data
            })
        except WorkOrder.DoesNotExist:
            return Response(
                {'error': 'Work order not found'},
                status=status.HTTP_404_NOT_FOUND
            )
        except Exception as e:
            return Response(
                {'error': str(e)},
                status=status.HTTP_400_BAD_REQUEST
            )
    
    @action(detail=False, methods=['post'], url_path='bulk-mark-coc')
    def bulk_mark_coc(self, request):
        """Bulk mark COC received for multiple work orders"""
        ids = request.data.get('ids', [])
        date_received = request.data.get('date_sched')
        
        if not ids:
            return Response(
                {'error': 'No work order IDs provided'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            from datetime import datetime
            from django.utils import timezone
            
            if date_received:
                coc_date = datetime.strptime(date_received, '%Y-%m-%d').date()
            else:
                coc_date = timezone.now().date()
            
            updated = WorkOrder.objects.filter(
                id__in=ids
            ).update(date_sched=coc_date)
            
            return Response({
                'message': f'Successfully marked COC received for {updated} work orders',
                'updated_count': updated
            })
        except Exception as e:
            return Response(
                {'error': str(e)},
                status=status.HTTP_400_BAD_REQUEST
            )
    


class EngineeringAideViewSet(viewsets.ViewSet):
    """
    Engineering Aide Portal - Workflow coordination and monitoring
    """
    permission_classes = [AllowAny]
    
    @action(detail=False, methods=['get'])
    def workflow_overview(self, request):
        """Get overview of all active workflows"""
        active_workflows = ProjectWorkflow.objects.filter(
            is_current_stage=True
        ).select_related('project', 'stage', 'assigned_user')
        
        # Group by stage
        by_stage = {}
        for workflow in active_workflows:
            stage_name = workflow.stage.stage_name
            if stage_name not in by_stage:
                by_stage[stage_name] = {
                    'count': 0,
                    'projects': []
                }
            by_stage[stage_name]['count'] += 1
            by_stage[stage_name]['projects'].append({
                'project_code': workflow.project.project_code,
                'project_name': workflow.project.project_name,
                'assigned_to': workflow.assigned_user.get_full_name() if workflow.assigned_user else None,
                'due_date': workflow.due_date,
                'status': workflow.status
            })
        
        return Response({
            'total_active': active_workflows.count(),
            'by_stage': by_stage,
            'blocked': active_workflows.filter(status='Blocked').count(),
            'overdue': active_workflows.filter(
                due_date__lt=date.today(),
                completion_date__isnull=True
            ).count()
        })
    
    @action(detail=False, methods=['get'])
    def workflow_visualization(self, request):
        """Get data for workflow visualization"""
        project_id = request.query_params.get('project_id')
        
        if not project_id:
            return Response({'error': 'project_id required'}, status=400)
        
        workflows = ProjectWorkflow.objects.filter(
            project_id=project_id
        ).select_related('stage', 'assigned_user').order_by('stage__stage_order')
        
        timeline = []
        for wf in workflows:
            timeline.append({
                'stage_name': wf.stage.stage_name,
                'stage_order': wf.stage.stage_order,
                'status': wf.status,
                'start_date': wf.start_date,
                'due_date': wf.due_date,
                'completion_date': wf.completion_date,
                'assigned_to': wf.assigned_user.get_full_name() if wf.assigned_user else None,
                'is_current': wf.is_current_stage,
                'notes': wf.notes
            })
        
        return Response({
            'project_id': project_id,
            'timeline': timeline,
            'total_stages': len(timeline),
            'completed_stages': len([w for w in timeline if w['status'] == 'Completed']),
            'current_stage': next((w for w in timeline if w['is_current']), None)
        })
    
    @action(detail=False, methods=['get'])
    def document_compliance_summary(self, request):
        """Get document compliance summary across all projects"""
        total_required = DocumentCompliance.objects.count()
        submitted = DocumentCompliance.objects.filter(is_submitted=True).count()
        approved = DocumentCompliance.objects.filter(is_approved=True).count()
        overdue = DocumentCompliance.objects.filter(is_overdue=True, is_submitted=False).count()
        
        # By document type
        by_type = DocumentCompliance.objects.values(
            'doc_type__doc_type_name'
        ).annotate(
            total=Count('id'),
            submitted=Count('id', filter=Q(is_submitted=True)),
            approved=Count('id', filter=Q(is_approved=True)),
            overdue=Count('id', filter=Q(is_overdue=True, is_submitted=False))
        )
        
        return Response({
            'overall': {
                'total_required': total_required,
                'submitted': submitted,
                'approved': approved,
                'overdue': overdue,
                'compliance_rate': round((submitted / total_required * 100) if total_required > 0 else 0, 2)
            },
            'by_type': list(by_type)
        })
    
    @action(detail=False, methods=['post'])
    def send_workflow_notification(self, request):
        """Send notification to workflow stage assignee"""
        workflow_id = request.data.get('workflow_id')
        message = request.data.get('message')
        
        try:
            workflow = ProjectWorkflow.objects.get(id=workflow_id)
            
            if not workflow.assigned_user:
                return Response({'error': 'No user assigned to this workflow stage'}, status=400)
            
            notification = Notification.objects.create(
                recipient_user=workflow.assigned_user,
                notification_type='In-App',
                subject=f'Workflow Update: {workflow.stage.stage_name}',
                message=message,
                related_project=workflow.project,
                status='Pending'
            )
            
            return Response({
                'status': 'sent',
                'notification_id': notification.id,
                'recipient': workflow.assigned_user.get_full_name()
            })
            
        except ProjectWorkflow.DoesNotExist:
            return Response({'error': 'Workflow not found'}, status=404)
    
    @action(detail=False, methods=['get'])
    def upcoming_deadlines(self, request):
        """Get upcoming workflow deadlines"""
        days_ahead = int(request.query_params.get('days', 7))
        end_date = date.today() + timedelta(days=days_ahead)
        
        upcoming = ProjectWorkflow.objects.filter(
            due_date__gte=date.today(),
            due_date__lte=end_date,
            completion_date__isnull=True
        ).select_related('project', 'stage', 'assigned_user').order_by('due_date')
        
        calendar_data = []
        for workflow in upcoming:
            days_until = (workflow.due_date - date.today()).days
            calendar_data.append({
                'project_code': workflow.project.project_code,
                'project_name': workflow.project.project_name,
                'stage_name': workflow.stage.stage_name,
                'due_date': workflow.due_date,
                'days_until': days_until,
                'assigned_to': workflow.assigned_user.get_full_name() if workflow.assigned_user else None,
                'priority': workflow.project.priority,
                'urgency': 'critical' if days_until <= 2 else 'high' if days_until <= 5 else 'normal'
            })
        
        return Response({
            'total_upcoming': len(calendar_data),
            'deadlines': calendar_data
        })
    
    @action(detail=False, methods=['get'])
    def summary_report(self, request):
        """Generate summary report for aide"""
        report_type = request.query_params.get('type', 'daily')
        
        if report_type == 'daily':
            start_date = date.today()
        elif report_type == 'weekly':
            start_date = date.today() - timedelta(days=7)
        else:  # monthly
            start_date = date.today() - timedelta(days=30)
        
        # Workflows completed
        completed_workflows = ProjectWorkflow.objects.filter(
            completion_date__gte=start_date,
            status='Completed'
        ).count()
        
        # Documents uploaded
        documents_uploaded = ProjectDocument.objects.filter(
            upload_date__gte=start_date
        ).count()
        
        # Notifications sent
        notifications_sent = Notification.objects.filter(
            created_at__gte=start_date
        ).count()
        
        # Active issues
        blocked_workflows = ProjectWorkflow.objects.filter(
            status='Blocked'
        ).count()
        
        overdue_deadlines = ProjectWorkflow.objects.filter(
            due_date__lt=date.today(),
            completion_date__isnull=True
        ).count()
        
        return Response({
            'period': report_type,
            'start_date': start_date,
            'end_date': date.today(),
            'metrics': {
                'completed_workflows': completed_workflows,
                'documents_uploaded': documents_uploaded,
                'notifications_sent': notifications_sent,
                'blocked_workflows': blocked_workflows,
                'overdue_deadlines': overdue_deadlines
            }
        })


class EngineerViewSet(viewsets.ViewSet):
    """
    Engineer/Design Engineer Portal - Technical review and approvals
    """
    permission_classes = [AllowAny]
    
    @action(detail=False, methods=['get'])
    def my_projects(self, request):
        """Get projects assigned to engineer"""
        projects = Project.objects.all()
        
        # Categorize by status
        categorized = {
            'active': [],
            'pending_review': [],
            'delayed': [],
            'sla_risk': []
        }
        
        for project in projects:
            project_data = ProjectListSerializer(project).data
            
            if project.is_delayed:
                categorized['delayed'].append(project_data)
            
            # Check SLA risk
            sla_at_risk = SLATracking.objects.filter(
                project=project,
                is_breached=False,
                due_date__lte=date.today() + timedelta(days=3)
            ).exists()
            if sla_at_risk:
                categorized['sla_risk'].append(project_data)
            
            # Check pending approvals
            pending_docs = ProjectDocument.objects.filter(
                project=project,
                approval_status='Pending'
            ).exists()
            if pending_docs:
                categorized['pending_review'].append(project_data)
            else:
                categorized['active'].append(project_data)
        
        return Response({
            'total': projects.count(),
            'categorized': categorized
        })
    
    @action(detail=False, methods=['get'])
    def pending_approvals(self, request):
        """Get items pending engineer approval"""
        # Documents pending approval
        documents = ProjectDocument.objects.filter(
            approval_status='Pending',
            project__assigned_engineer=request.user
        ).select_related('project', 'doc_type', 'uploaded_by')
        
        # Inspections pending approval
        inspections = QIInspection.objects.filter(
            is_completed=True,
            project__assigned_engineer=request.user
        ).select_related('project', 'inspection_type', 'assigned_qi')
        
        # Projects pending final approval
        projects = Project.objects.filter(
            assigned_engineer=request.user,
            status__status_name='Pending Approval'
        ).select_related('vendor', 'sector')
        
        return Response({
            'documents': {
                'count': documents.count(),
                'items': ProjectDocumentSerializer(documents, many=True).data
            },
            'inspections': {
                'count': inspections.count(),
                'items': QIInspectionSerializer(inspections, many=True).data
            },
            'projects': {
                'count': projects.count(),
                'items': ProjectListSerializer(projects, many=True).data
            }
        })
    
    @action(detail=False, methods=['post'])
    def approve_document(self, request):
        """Approve a document"""
        document_id = request.data.get('document_id')
        comments = request.data.get('comments', '')
        
        try:
            document = ProjectDocument.objects.get(id=document_id)
            
            # Check if engineer is assigned to project
            if document.project.assigned_engineer != request.user:
                return Response({'error': 'Not authorized'}, status=403)
            
            document.approval_status = 'Approved'
            document.approved_by = request.user
            document.approval_date = timezone.now()
            document.notes = comments
            document.save()
            
            # Update compliance
            DocumentCompliance.objects.filter(
                project=document.project,
                doc_type=document.doc_type
            ).update(
                is_approved=True,
                approval_date=timezone.now()
            )
            
            return Response({
                'status': 'approved',
                'document_id': document_id
            })
            
        except ProjectDocument.DoesNotExist:
            return Response({'error': 'Document not found'}, status=404)
    
    @action(detail=False, methods=['post'])
    def reject_document(self, request):
        """Reject a document"""
        document_id = request.data.get('document_id')
        reason = request.data.get('reason', '')
        
        try:
            document = ProjectDocument.objects.get(id=document_id)
            
            if document.project.assigned_engineer != request.user:
                return Response({'error': 'Not authorized'}, status=403)
            
            document.approval_status = 'Rejected'
            document.rejection_reason = reason
            document.save()
            
            # Send notification to uploader
            Notification.objects.create(
                recipient_user=document.uploaded_by,
                notification_type='In-App',
                subject='Document Rejected',
                message=f'Document {document.document_name} was rejected. Reason: {reason}',
                related_project=document.project,
                status='Pending'
            )
            
            return Response({
                'status': 'rejected',
                'document_id': document_id
            })
            
        except ProjectDocument.DoesNotExist:
            return Response({'error': 'Document not found'}, status=404)
    
    @action(detail=False, methods=['get'])
    def sla_compliance(self, request):
        """Get SLA compliance for engineer's projects"""
        projects = Project.objects.filter(assigned_engineer=request.user)
        
        sla_data = SLATracking.objects.filter(
            project__in=projects
        ).values('status').annotate(count=Count('sla_tracking_id'))
        
        # At-risk SLAs
        at_risk = SLATracking.objects.filter(
            project__in=projects,
            is_breached=False,
            due_date__lte=date.today() + timedelta(days=3),
            completion_date__isnull=True
        ).select_related('project', 'sla_rule')
        
        return Response({
            'summary': list(sla_data),
            'at_risk': SLATrackingSerializer(at_risk, many=True).data,
            'breach_rate': self._calculate_breach_rate(projects)
        })
    
    def _calculate_breach_rate(self, projects):
        total = SLATracking.objects.filter(project__in=projects).count()
        if total == 0:
            return 0
        breached = SLATracking.objects.filter(
            project__in=projects,
            is_breached=True
        ).count()
        return round((breached / total * 100), 2)
    
    @action(detail=False, methods=['get'])
    def vendor_performance(self, request):
        """Get vendor performance analytics for engineer's projects"""
        projects = Project.objects.filter(assigned_engineer=request.user)
        vendors = Vendor.objects.filter(projects__in=projects).distinct()
        
        performance_data = []
        for vendor in vendors:
            vendor_projects = projects.filter(vendor=vendor)
            
            total = vendor_projects.count()
            delayed = vendor_projects.filter(is_delayed=True).count()
            completed = vendor_projects.filter(status__status_name='Completed').count()
            
            penalties = Penalty.objects.filter(
                vendor=vendor,
                project__in=vendor_projects
            ).exclude(penalty_status='Waived').aggregate(
                total=Sum('penalty_amount')
            )['total'] or 0
            
            performance_data.append({
                'vendor_code': vendor.vendor_code,
                'vendor_name': vendor.vendor_name,
                'total_projects': total,
                'completed': completed,
                'delayed': delayed,
                'on_time_rate': round(((total - delayed) / total * 100) if total > 0 else 0, 2),
                'total_penalties': float(penalties),
                'compliance_score': float(vendor.compliance_score)
            })
        
        return Response({
            'vendors': performance_data,
            'summary': {
                'total_vendors': len(performance_data),
                'avg_on_time_rate': round(
                    sum(v['on_time_rate'] for v in performance_data) / len(performance_data)
                    if performance_data else 0, 2
                )
            }
        })
    
    @action(detail=False, methods=['post'])
    def use_chatbot(self, request):
        """Use AI chatbot for historical queries"""
        question = request.data.get('question')
        
        if not question:
            return Response({'error': 'question required'}, status=400)
        
        # Use the existing chatbot service
        from .chatbot_service import chatbot_service
        answer = chatbot_service.answer(question)
        
        # Log the query
        SystemAuditLog.objects.create(
            user=request.user,
            action_type='AI_CHATBOT_QUERY',
            action_description=f'Question: {question[:100]}',
            status='Success'
        )
        
        return Response({
            'question': question,
            'answer': answer
        })

class WOSupervisorViewSet(viewsets.ViewSet):
    """
    Work Order Supervisor Portal - Full operational oversight
    """
    permission_classes = [AllowAny]
    
    @action(detail=False, methods=['get'])
    def full_dashboard(self, request):
        """Comprehensive supervisor dashboard"""
        # Project tracking
        total_projects = Project.objects.count()
        active = Project.objects.exclude(
            status__status_name__in=['Completed', 'Cancelled', 'Billed']
        ).count()
        delayed = Project.objects.filter(is_delayed=True).count()
        
        # Penalties
        pending_penalties = Penalty.objects.filter(penalty_status='Draft').count()
        total_penalty_amount = float(Penalty.objects.exclude(
            penalty_status='Waived'
        ).aggregate(total=Sum('penalty_amount'))['total'] or 0)
        
        # SLA tracking
        sla_breaches = SLATracking.objects.filter(
            is_breached=True,
            status='Breached'
        ).count()
        sla_at_risk = SLATracking.objects.filter(
            is_breached=False,
            due_date__lte=date.today() + timedelta(days=3),
            completion_date__isnull=True
        ).count()
        
        # QI workload
        qi_workload = self._get_qi_workload_distribution()
        
        # Escalations
        open_escalations = Escalation.objects.filter(status='Open').count()
        
        # Financial
        billing_stats = self._get_billing_stats()
        
        return Response({
            'projects': {
                'total': total_projects,
                'active': active,
                'delayed': delayed,
                'completion_rate': round(
                    ((total_projects - active) / total_projects * 100) if total_projects > 0 else 0, 2
                )
            },
            'penalties': {
                'pending': pending_penalties,
                'total_amount': total_penalty_amount
            },
            'sla': {
                'breaches': sla_breaches,
                'at_risk': sla_at_risk
            },
            'qi_workload': qi_workload,
            'escalations': open_escalations,
            'billing': billing_stats
        })
    
    def _get_qi_workload_distribution(self):
        """Get QI workload distribution"""
        qi_users = User.objects.filter(role__role_name='Quality Inspector')
        workload = []
        
        for qi in qi_users:
            pending = QIInspection.objects.filter(
                assigned_qi=qi,
                is_completed=False
            ).count()
            
            completed_today = QIInspection.objects.filter(
                assigned_qi=qi,
                inspection_date=date.today(),
                is_completed=True
            ).count()
            
            # Get target
            target = QIDailyTarget.objects.filter(
                qi_user=qi,
                target_date=date.today()
            ).first()
            
            workload.append({
                'qi_name': qi.get_full_name(),
                'pending_inspections': pending,
                'completed_today': completed_today,
                'daily_target': target.target_audits if target else 0,
                'target_met': target.target_met if target else False
            })
        
        return workload
    
    def _get_billing_stats(self):
        """Get billing statistics"""
        return {
            'total_invoiced': float(Invoice.objects.aggregate(
                total=Sum('invoice_amount')
            )['total'] or 0),
            'total_paid': float(Invoice.objects.filter(
                payment_status='Paid'
            ).aggregate(total=Sum('net_amount'))['total'] or 0),
            'outstanding': float(Invoice.objects.exclude(
                payment_status='Paid'
            ).aggregate(total=Sum('net_amount'))['total'] or 0),
            'overdue': Invoice.objects.filter(
                payment_status__in=['Unpaid', 'Partially Paid'],
                due_date__lt=date.today()
            ).count()
        }
    
    @action(detail=False, methods=['get'])
    def manage_penalties(self, request):
        """Manage penalties and violations"""
        status_filter = request.query_params.get('status', 'all')
        
        penalties = Penalty.objects.all().select_related(
            'project', 'vendor', 'penalty_rule'
        )
        
        if status_filter != 'all':
            penalties = penalties.filter(penalty_status=status_filter)
        
        # Group by vendor
        by_vendor = penalties.values(
            'vendor__vendor_name'
        ).annotate(
            total_penalties=Count('id'),
            total_amount=Sum('penalty_amount')
        ).order_by('-total_amount')
        
        return Response({
            'penalties': PenaltySerializer(penalties, many=True).data,
            'by_vendor': list(by_vendor),
            'summary': {
                'total': penalties.count(),
                'draft': penalties.filter(penalty_status='Draft').count(),
                'issued': penalties.filter(penalty_status='Issued').count(),
                'paid': penalties.filter(penalty_status='Paid').count(),
                'waived': penalties.filter(penalty_status='Waived').count()
            }})
    @action(detail=False, methods=['post'])
    def manage_escalation(self, request):
        """Create or manage escalation"""
        action = request.data.get('action')  # 'create', 'assign', 'resolve'
        
        if action == 'create':
            return self._create_escalation(request)
        elif action == 'assign':
            return self._assign_escalation(request)
        elif action == 'resolve':
            return self._resolve_escalation(request)
        else:
            return Response({'error': 'Invalid action'}, status=400)

    def _create_escalation(self, request):
        """Create new escalation"""
        project_id = request.data.get('project_id')
        rule_id = request.data.get('escalation_rule_id')
        reason = request.data.get('reason')
        escalate_to_user_id = request.data.get('escalate_to_user_id')
        
        try:
            project = Project.objects.get(project_id=project_id)
            rule = EscalationRule.objects.get(id=rule_id)
            escalate_to = User.objects.get(id=escalate_to_user_id)
            
            escalation = Escalation.objects.create(
                project=project,
                escalation_rule=rule,
                escalated_from_user=request.user,
                escalated_to_user=escalate_to,
                escalation_reason=reason,
                status='Open'
            )
            
            # Send notification
            Notification.objects.create(
                recipient_user=escalate_to,
                notification_type='In-App',
                subject=f'Escalation: {project.project_code}',
                message=reason,
                related_project=project,
                status='Pending'
            )
            
            return Response({
                'status': 'created',
                'escalation_id': escalation.id
            })
            
        except (Project.DoesNotExist, EscalationRule.DoesNotExist, User.DoesNotExist) as e:
            return Response({'error': str(e)}, status=404)

    def _assign_escalation(self, request):
        """Assign escalation to user"""
        escalation_id = request.data.get('escalation_id')
        user_id = request.data.get('user_id')
        
        try:
            escalation = Escalation.objects.get(id=escalation_id)
            user = User.objects.get(id=user_id)
            
            escalation.escalated_to_user = user
            escalation.status = 'In Progress'
            escalation.save()
            
            return Response({'status': 'assigned'})
            
        except (Escalation.DoesNotExist, User.DoesNotExist) as e:
            return Response({'error': str(e)}, status=404)

    def _resolve_escalation(self, request):
        """Resolve escalation"""
        escalation_id = request.data.get('escalation_id')
        resolution = request.data.get('resolution')
        
        try:
            escalation = Escalation.objects.get(id=escalation_id)
            
            escalation.status = 'Resolved'
            escalation.resolution = resolution
            escalation.resolved_by = request.user
            escalation.resolution_date = timezone.now()
            escalation.save()
            
            return Response({'status': 'resolved'})
            
        except Escalation.DoesNotExist:
            return Response({'error': 'Escalation not found'}, status=404)

    @action(detail=False, methods=['get'])
    def predictive_analytics(self, request):
        """Get predictive analytics for supervisor"""
        # Use ML service for predictions
        from .ml_service import ml_service
        
        # Get projects at risk
        active_projects = Project.objects.exclude(
            status__status_name__in=['Completed', 'Cancelled']
        )
        
        at_risk_projects = []
        for project in active_projects:
            # Simple risk calculation based on delays and SLA
            risk_score = 0
            
            if project.is_delayed:
                risk_score += 30
            
            sla_breaches = SLATracking.objects.filter(
                project=project,
                is_breached=True
            ).count()
            risk_score += (sla_breaches * 20)
            
            pending_docs = DocumentCompliance.objects.filter(
                project=project,
                is_overdue=True,
                is_submitted=False
            ).count()
            risk_score += (pending_docs * 10)
            
            if risk_score >= 50:
                at_risk_projects.append({
                    'project_code': project.project_code,
                    'project_name': project.project_name,
                    'risk_score': min(risk_score, 100),
                    'risk_level': 'Critical' if risk_score >= 70 else 'High',
                    'factors': {
                        'is_delayed': project.is_delayed,
                        'sla_breaches': sla_breaches,
                        'overdue_documents': pending_docs
                    }
                })
        
        return Response({
            'at_risk_projects': at_risk_projects,
            'total_at_risk': len(at_risk_projects)
        })



class TeamLeaderViewSet(viewsets.ViewSet):
    """
    Team Leader Portal - Strategic management and approvals
    """
    permission_classes = [AllowAny]
    
    @action(detail=False, methods=['get'])
    def organization_overview(self, request):
        """Get organization-wide monitoring dashboard (TV Mode)"""
        # Real-time metrics
        total_projects = Project.objects.count()
        active_projects = Project.objects.exclude(
            status__status_name__in=['Completed', 'Cancelled', 'Billed']
        ).count()
        delayed_projects = Project.objects.filter(is_delayed=True).count()
        
        # Team performance
        team_performance = self._get_team_performance()
        
        # Vendor performance
        vendor_rankings = self._get_vendor_rankings()
        
        # Financial overview
        financial = {
            'total_contract_value': float(Project.objects.aggregate(
                total=Sum('contract_value')
            )['total'] or 0),
            'total_penalties': float(Penalty.objects.exclude(
                penalty_status='Waived'
            ).aggregate(total=Sum('penalty_amount'))['total'] or 0),
            'outstanding_payments': float(Invoice.objects.exclude(
                payment_status='Paid'
            ).aggregate(total=Sum('net_amount'))['total'] or 0)
        }
        
        # Recent alerts
        alerts = self._get_critical_alerts()
        
        return Response({
            'overview': {
                'total_projects': total_projects,
                'active_projects': active_projects,
                'delayed_projects': delayed_projects,
                'on_time_percentage': round(
                    ((total_projects - delayed_projects) / total_projects * 100) 
                    if total_projects > 0 else 0, 2
                )
            },
            'team_performance': team_performance,
            'vendor_rankings': vendor_rankings,
            'financial': financial,
            'alerts': alerts
        })
    
    def _get_team_performance(self):
        """Get team member performance metrics"""
        engineers = User.objects.filter(role__role_name='Engineer')
        qi_users = User.objects.filter(role__role_name='Quality Inspector')
        
        performance = {
            'engineers': [],
            'quality_inspectors': []
        }
        
        # Engineer metrics
        for engineer in engineers:
            assigned = Project.objects.filter(assigned_engineer=engineer)
            completed = assigned.filter(status__status_name='Completed')
            delayed = assigned.filter(is_delayed=True)
            
            performance['engineers'].append({
                'name': engineer.get_full_name(),
                'assigned_projects': assigned.count(),
                'completed': completed.count(),
                'delayed': delayed.count(),
                'on_time_rate': round(
                    ((assigned.count() - delayed.count()) / assigned.count() * 100)
                    if assigned.count() > 0 else 0, 2
                )
            })
        
        # QI metrics
        for qi in qi_users:
            inspections = QIInspection.objects.filter(assigned_qi=qi)
            completed = inspections.filter(is_completed=True)
            
            # Monthly target achievement
            current_month = date.today().replace(day=1)
            monthly = QIMonthlyAccomplishment.objects.filter(
                qi_user=qi,
                month=current_month
            ).first()
            
            performance['quality_inspectors'].append({
                'name': qi.get_full_name(),
                'total_inspections': inspections.count(),
                'completed': completed.count(),
                'monthly_target': monthly.target_inspections if monthly else 0,
                'monthly_actual': monthly.total_inspections if monthly else 0,
                'target_met': monthly.target_met if monthly else False
            })
        
        return performance
    
    def _get_vendor_rankings(self):
        """Get vendor performance rankings"""
        vendors = Vendor.objects.filter(is_active=True).annotate(
            total_projects=Count('projects'),
            delayed_projects=Count('projects', filter=Q(projects__is_delayed=True)),
            total_penalties=Sum('penalties__penalty_amount', 
                               filter=~Q(penalties__penalty_status='Waived'))
        ).order_by('-compliance_score')[:10]
        
        rankings = []
        for vendor in vendors:
            on_time = vendor.total_projects - vendor.delayed_projects
            rankings.append({
                'vendor_name': vendor.vendor_name,
                'compliance_score': float(vendor.compliance_score),
                'total_projects': vendor.total_projects,
                'on_time_rate': round(
                    (on_time / vendor.total_projects * 100) 
                    if vendor.total_projects > 0 else 0, 2
                ),
                'total_penalties': float(vendor.total_penalties or 0)
            })
        
        return rankings
    
    def _get_critical_alerts(self):
        """Get critical alerts requiring attention"""
        alerts = []
        
        # SLA breaches
        sla_breaches = SLATracking.objects.filter(
            is_breached=True,
            status='Breached'
        ).count()
        if sla_breaches > 0:
            alerts.append({
                'type': 'SLA_BREACH',
                'severity': 'critical',
                'count': sla_breaches,
                'message': f'{sla_breaches} SLA breaches require attention'
            })
        
        # Pending penalty approvals
        pending_penalties = Penalty.objects.filter(penalty_status='Draft').count()
        if pending_penalties > 0:
            alerts.append({
                'type': 'PENDING_PENALTY',
                'severity': 'high',
                'count': pending_penalties,
                'message': f'{pending_penalties} penalties pending approval'
            })
        
        # Overdue documents
        overdue_docs = DocumentCompliance.objects.filter(
            is_overdue=True,
            is_submitted=False
        ).count()
        if overdue_docs > 5:
            alerts.append({
                'type': 'OVERDUE_DOCUMENTS',
                'severity': 'medium',
                'count': overdue_docs,
                'message': f'{overdue_docs} documents overdue'
            })
        
        # Open escalations
        open_escalations = Escalation.objects.filter(status='Open').count()
        if open_escalations > 0:
            alerts.append({
                'type': 'ESCALATION',
                'severity': 'high',
                'count': open_escalations,
                'message': f'{open_escalations} open escalations'
            })
        
        return alerts
    
    @action(detail=False, methods=['get'])
    def pending_approvals(self, request):
        """Get all items pending team leader approval"""
        # Penalties
        penalties = Penalty.objects.filter(
            penalty_status='Draft'
        ).select_related('project', 'vendor', 'penalty_rule')
        
        # High-value invoices
        invoices = Invoice.objects.filter(
            approval_status='Pending',
            invoice_amount__gte=100000  # High-value threshold
        ).select_related('project', 'vendor')
        
        # Projects
        projects = Project.objects.filter(
            status__status_name='Pending Approval'
        ).select_related('vendor', 'sector', 'assigned_engineer')
        
        return Response({
            'penalties': {
                'count': penalties.count(),
                'items': PenaltySerializer(penalties, many=True).data,
                'total_amount': float(penalties.aggregate(
                    total=Sum('penalty_amount')
                )['total'] or 0)
            },
            'invoices': {
                'count': invoices.count(),
                'items': InvoiceSerializer(invoices, many=True).data
            },
            'projects': {
                'count': projects.count(),
                'items': ProjectListSerializer(projects, many=True).data
            }
        })
    
    @action(detail=False, methods=['post'])
    def approve_penalty(self, request):
        """Approve or reject penalty"""
        penalty_id = request.data.get('penalty_id')
        action = request.data.get('action')  # 'approve' or 'reject'
        reason = request.data.get('reason', '')
        
        try:
            penalty = Penalty.objects.get(id=penalty_id)
            
            if action == 'approve':
                penalty.penalty_status = 'Issued'
                penalty.approved_by = request.user
                penalty.approval_date = timezone.now()
                penalty.issue_date = date.today()
                penalty.save()
                
                # Send notification to vendor
                if penalty.vendor:
                    Notification.objects.create(
                        recipient_email=penalty.vendor.email,
                        notification_type='Email',
                        subject=f'Penalty Issued: {penalty.project.project_code}',
                        message=f'A penalty of {penalty.penalty_amount} has been issued.',
                        related_project=penalty.project,
                        status='Pending'
                    )
                
                return Response({'status': 'approved', 'penalty_id': penalty_id})
                
            elif action == 'reject':
                penalty.penalty_status = 'Waived'
                penalty.waiver_reason = reason
                penalty.waived_by = request.user
                penalty.waiver_date = timezone.now()
                penalty.save()
                
                return Response({'status': 'rejected', 'penalty_id': penalty_id})
            
            else:
                return Response({'error': 'Invalid action'}, status=400)
                
        except Penalty.DoesNotExist:
            return Response({'error': 'Penalty not found'}, status=404)
    
    @action(detail=False, methods=['get'])
    def performance_trends(self, request):
        """Get historical performance trends"""
        months = int(request.query_params.get('months', 12))
        start_date = date.today() - timedelta(days=30 * months)
        
        # Monthly project completion trend
        monthly_data = Project.objects.filter(
            created_at__gte=start_date
        ).annotate(
            month=TruncMonth('created_at')
        ).values('month').annotate(
            total=Count('project_id'),
            completed=Count('project_id', filter=Q(status__status_name='Completed')),
            delayed=Count('project_id', filter=Q(is_delayed=True))
        ).order_by('month')
        
        # Calculate trends
        trend_data = list(monthly_data)
        for item in trend_data:
            if item['total'] > 0:
                item['on_time_percentage'] = round(
                    ((item['total'] - item['delayed']) / item['total'] * 100), 2
                )
                item['completion_rate'] = round(
                    (item['completed'] / item['total'] * 100), 2
                )
        
        return Response({
            'period_months': months,
            'monthly_trends': trend_data
        })
    
    @action(detail=False, methods=['get'])
    def comparison_report(self, request):
        """Compare current vs previous period performance"""
        # Current month
        current_month_start = date.today().replace(day=1)
        if current_month_start.month == 1:
            previous_month_start = current_month_start.replace(
                year=current_month_start.year - 1, month=12
            )
        else:
            previous_month_start = current_month_start.replace(
                month=current_month_start.month - 1
            )
        
        current_stats = self._get_period_stats(current_month_start, date.today())
        previous_stats = self._get_period_stats(previous_month_start, current_month_start)
        
        # Calculate changes
        comparison = {
            'projects_completed': {
                'current': current_stats['completed'],
                'previous': previous_stats['completed'],
                'change': current_stats['completed'] - previous_stats['completed'],
                'change_percentage': round(
                    ((current_stats['completed'] - previous_stats['completed']) / 
                     previous_stats['completed'] * 100) 
                    if previous_stats['completed'] > 0 else 0, 2
                )
            },
            'on_time_rate': {
                'current': current_stats['on_time_rate'],
                'previous': previous_stats['on_time_rate'],
                'change': current_stats['on_time_rate'] - previous_stats['on_time_rate']
            },
            'total_penalties': {
                'current': current_stats['penalties'],
                'previous': previous_stats['penalties'],
                'change': current_stats['penalties'] - previous_stats['penalties']
            }
        }
        
        return Response(comparison)
    
    def _get_period_stats(self, start_date, end_date):
        """Get statistics for a specific period"""
        projects = Project.objects.filter(
            created_at__gte=start_date,
            created_at__lt=end_date
        )
        
        completed = projects.filter(status__status_name='Completed').count()
        delayed = projects.filter(is_delayed=True).count()
        total = projects.count()
        
        penalties = float(Penalty.objects.filter(
            created_at__gte=start_date,
            created_at__lt=end_date
        ).exclude(penalty_status='Waived').aggregate(
            total=Sum('penalty_amount')
        )['total'] or 0)
        
        return {
            'completed': completed,
            'delayed': delayed,
            'total': total,
            'on_time_rate': round(
                ((total - delayed) / total * 100) if total > 0 else 0, 2
            ),
            'penalties': penalties
        }
    
    @action(detail=False, methods=['get'])
    def ai_suggestions(self, request):
        """Get AI-powered improvement suggestions"""
        suggestions = []
        
        # Analyze delay patterns
        delay_factors = ProjectDelay.objects.values(
            'factor__factor_name'
        ).annotate(
            occurrence=Count('delay_id')
        ).order_by('-occurrence')[:5]
        
        for factor in delay_factors:
            suggestions.append({
                'category': 'DELAY_REDUCTION',
                'priority': 'high',
                'title': f'Address {factor["factor__factor_name"]}',
                'description': f'This factor caused {factor["occurrence"]} delays. Consider implementing preventive measures.',
                'potential_impact': 'Could reduce delays by 15-20%'
            })
        
        # Analyze vendor performance
        underperforming = Vendor.objects.filter(
            is_active=True,
            compliance_score__lt=70
        ).count()
        
        if underperforming > 0:
            suggestions.append({
                'category': 'VENDOR_MANAGEMENT',
                'priority': 'medium',
                'title': 'Review Underperforming Vendors',
                'description': f'{underperforming} vendors have compliance scores below 70%.',
                'potential_impact': 'Could improve overall project completion rate'
            })
        
        # Check QI workload imbalance
        qi_workload = QIInspection.objects.filter(
            is_completed=False
        ).values('assigned_qi').annotate(
            pending=Count('id')
        ).order_by('-pending')
        
        if qi_workload and qi_workload[0]['pending'] > 20:
            suggestions.append({
                'category': 'RESOURCE_ALLOCATION',
                'priority': 'medium',
                'title': 'Rebalance QI Workload',
                'description': 'Some QIs have high pending inspection counts.',
                'potential_impact': 'Could improve inspection completion time'
            })
        
        return Response({
            'suggestions': suggestions,
            'generated_at': timezone.now()
        })
    
    @action(detail=False, methods=['post'])
    def manage_user_access(self, request):
        """Manage user access and permissions"""
        action = request.data.get('action')  # 'activate', 'deactivate', 'change_role'
        user_id = request.data.get('user_id')
        
        try:
            user = User.objects.get(user_id=user_id)
            
            if action == 'activate':
                user.is_active = True
                user.save()
                return Response({'status': 'activated', 'user_id': user_id})
                
            elif action == 'deactivate':
                user.is_active = False
                user.save()
                
                # Deactivate sessions
                UserSession.objects.filter(
                    user=user,
                    is_active=True
                ).update(
                    is_active=False,
                    logout_time=timezone.now()
                )
                
                return Response({'status': 'deactivated', 'user_id': user_id})
                
            elif action == 'change_role':
                new_role_id = request.data.get('role_id')
                new_role = UserRole.objects.get(role_id=new_role_id)
                user.role = new_role
                user.save()
                
                return Response({
                    'status': 'role_changed',
                    'user_id': user_id,
                    'new_role': new_role.role_name
                })
            
            else:
                return Response({'error': 'Invalid action'}, status=400)
                
        except User.DoesNotExist:
            return Response({'error': 'User not found'}, status=404)
        except UserRole.DoesNotExist:
            return Response({'error': 'Role not found'}, status=404)


class SectorManagerViewSet(viewsets.ViewSet):
    """
    Sector Manager Portal - Executive KPI dashboard and strategic oversight
    """
    permission_classes = [AllowAny]
    
    @action(detail=False, methods=['get'])
    def executive_dashboard(self, request):
        """Get executive KPI dashboard"""
        # Get manager's sectors
        managed_sectors = Sector.objects.filter(sector_manager=request.user)
        
        # KPI Summary
        kpi_summary = self._get_kpi_summary()
        
        # Sector comparison
        sector_comparison = self._get_sector_comparison(managed_sectors)
        
        # Financial performance
        financial = self._get_financial_performance(managed_sectors)
        
        # Strategic metrics
        strategic = self._get_strategic_metrics(managed_sectors)
        
        return Response({
            'kpi_summary': kpi_summary,
            'sector_comparison': sector_comparison,
            'financial_performance': financial,
            'strategic_metrics': strategic
        })
    
    def _get_kpi_summary(self):
        """Get latest KPI snapshot summary"""
        latest_period = KPISnapshot.objects.latest('period_end').period_end
        
        kpis = KPISnapshot.objects.filter(
            period_end=latest_period
        ).values('kpi_type', 'kpi_value', 'target_value')
        
        kpi_data = {}
        for kpi in kpis:
            achievement = 0
            if kpi['target_value']:
                achievement = round(
                    (kpi['kpi_value'] / kpi['target_value'] * 100), 2
                )
            
            kpi_data[kpi['kpi_type']] = {
                'value': float(kpi['kpi_value']),
                'target': float(kpi['target_value']) if kpi['target_value'] else None,
                'achievement': achievement,
                'status': 'green' if achievement >= 100 else 'yellow' if achievement >= 80 else 'red'
            }
        
        return kpi_data
    
    def _get_sector_comparison(self, sectors):
        """Compare performance across sectors"""
        comparison = []
        
        for sector in sectors:
            projects = Project.objects.filter(sector=sector)
            total = projects.count()
            completed = projects.filter(status__status_name='Completed').count()
            delayed = projects.filter(is_delayed=True).count()
            
            contract_value = float(projects.aggregate(
                total=Sum('contract_value')
            )['total'] or 0)
            
            comparison.append({
                'sector_name': sector.sector_name,
                'total_projects': total,
                'completed': completed,
                'delayed': delayed,
                'on_time_rate': round(
                    ((total - delayed) / total * 100) if total > 0 else 0, 2
                ),
                'completion_rate': round(
                    (completed / total * 100) if total > 0 else 0, 2
                ),
                'contract_value': contract_value
            })
        
        return comparison
    
    def _get_financial_performance(self, sectors):
        """Get financial performance metrics"""
        projects = Project.objects.filter(sector__in=sectors)
        
        return {
            'total_contract_value': float(projects.aggregate(
                total=Sum('contract_value')
            )['total'] or 0),
            'completed_value': float(projects.filter(
                status__status_name='Completed'
            ).aggregate(total=Sum('contract_value'))['total'] or 0),
            'billed_amount': float(Invoice.objects.filter(
                project__sector__in=sectors
            ).aggregate(total=Sum('invoice_amount'))['total'] or 0),
            'collected_amount': float(Invoice.objects.filter(
                project__sector__in=sectors,
                payment_status='Paid'
            ).aggregate(total=Sum('net_amount'))['total'] or 0),
            'total_penalties': float(Penalty.objects.filter(
                project__sector__in=sectors
            ).exclude(penalty_status='Waived').aggregate(
                total=Sum('penalty_amount')
            )['total'] or 0),
            'outstanding_receivables': float(Invoice.objects.filter(
                project__sector__in=sectors
            ).exclude(payment_status='Paid').aggregate(
                total=Sum('net_amount')
            )['total'] or 0)
        }
    
    def _get_strategic_metrics(self, sectors):
        """Get strategic performance indicators"""
        projects = Project.objects.filter(sector__in=sectors)
        
        # Vendor diversity
        active_vendors = Vendor.objects.filter(
            projects__sector__in=sectors,
            is_active=True
        ).distinct().count()
        
        # Average project duration
        completed = projects.filter(
            status__status_name='Completed',
            start_date__isnull=False,
            completion_date__isnull=False
        )
        
        avg_duration = 0
        if completed.exists():
            durations = [
                (p.completion_date - p.start_date).days 
                for p in completed if p.completion_date and p.start_date
            ]
            if durations:
                avg_duration = round(sum(durations) / len(durations), 2)
        
        # Quality metrics
        total_inspections = QIInspection.objects.filter(
            project__sector__in=sectors
        ).count()
        passed_inspections = QIInspection.objects.filter(
            project__sector__in=sectors,
            inspection_result='Pass'
        ).count()
        
        return {
            'active_vendors': active_vendors,
            'avg_project_duration_days': avg_duration,
            'quality_pass_rate': round(
                (passed_inspections / total_inspections * 100) 
                if total_inspections > 0 else 0, 2
            ),
            'sla_compliance_rate': self._calculate_sla_compliance(projects)
        }
    
    def _calculate_sla_compliance(self, projects):
        """Calculate SLA compliance rate"""
        total_sla = SLATracking.objects.filter(project__in=projects).count()
        if total_sla == 0:
            return 0
        
        met = SLATracking.objects.filter(
            project__in=projects,
            status='Met'
        ).count()
        
        return round((met / total_sla * 100), 2)
    
    @action(detail=False, methods=['get'])
    def sector_trends(self, request):
        """Get sector performance trends over time"""
        sector_id = request.query_params.get('sector_id')
        months = int(request.query_params.get('months', 12))
        
        if not sector_id:
            return Response({'error': 'sector_id required'}, status=400)
        
        start_date = date.today() - timedelta(days=30 * months)
        
        monthly_data = Project.objects.filter(
            sector_id=sector_id,
            created_at__gte=start_date
        ).annotate(
            month=TruncMonth('created_at')
        ).values('month').annotate(
            total=Count('project_id'),
            completed=Count('project_id', filter=Q(status__status_name='Completed')),
            delayed=Count('project_id', filter=Q(is_delayed=True)),
            contract_value=Sum('contract_value')
        ).order_by('month')
        
        return Response({
            'sector_id': sector_id,
            'period_months': months,
            'monthly_trends': list(monthly_data)
        })
    
    @action(detail=False, methods=['get'])
    def vendor_rankings(self, request):
        """Get vendor performance rankings"""
        sector_id = request.query_params.get('sector_id')
        
        vendors = Vendor.objects.filter(is_active=True)
        
        if sector_id:
            vendors = vendors.filter(projects__sector_id=sector_id).distinct()
        
        rankings = []
        for vendor in vendors:
            projects = Project.objects.filter(vendor=vendor)
            if sector_id:
                projects = projects.filter(sector_id=sector_id)
            
            total = projects.count()
            if total == 0:
                continue
            
            completed = projects.filter(status__status_name='Completed').count()
            delayed = projects.filter(is_delayed=True).count()
            
            penalties = float(Penalty.objects.filter(
                vendor=vendor,
                project__in=projects
            ).exclude(penalty_status='Waived').aggregate(
                total=Sum('penalty_amount')
            )['total'] or 0)
            
            rankings.append({
                'vendor_name': vendor.vendor_name,
                'vendor_code': vendor.vendor_code,
                'compliance_score': float(vendor.compliance_score),
                'total_projects': total,
                'completion_rate': round((completed / total * 100), 2),
                'on_time_rate': round(((total - delayed) / total * 100), 2),
                'total_penalties': penalties,
                'avg_penalty_per_project': round(penalties / total, 2) if total > 0 else 0
            })
        
        # Sort by compliance score
        rankings.sort(key=lambda x: x['compliance_score'], reverse=True)
        
        return Response({
            'rankings': rankings,
            'total_vendors': len(rankings)
        })
    
    @action(detail=False, methods=['get'])
    def strategic_recommendations(self, request):
        """Get AI-powered strategic recommendations"""
        managed_sectors = Sector.objects.filter(sector_manager=request.user)
        projects = Project.objects.filter(sector__in=managed_sectors)
        
        recommendations = []
        
        # Analyze resource allocation
        sector_workload = projects.values('sector__sector_name').annotate(
            count=Count('project_id')
        )
        
        max_workload = max([s['count'] for s in sector_workload]) if sector_workload else 0
        min_workload = min([s['count'] for s in sector_workload]) if sector_workload else 0
        
        if max_workload - min_workload > 10:
            recommendations.append({
                'category': 'RESOURCE_ALLOCATION',
                'priority': 'high',
                'title': 'Rebalance Sector Workload',
                'description': 'Significant workload imbalance detected across sectors.',
                'potential_impact': 'Could improve overall efficiency by 10-15%',
                'action_items': [
                    'Review project distribution',
                    'Consider reallocating resources',
                    'Evaluate sector capacity'
                ]
            })
        
        # Vendor concentration risk
        vendor_concentration = projects.values('vendor').annotate(
            count=Count('project_id')
        ).order_by('-count')[:3]
        
        if vendor_concentration:
            top_vendor_share = (vendor_concentration[0]['count'] / projects.count() * 100)
            if top_vendor_share > 40:
                recommendations.append({
                    'category': 'VENDOR_DIVERSIFICATION',
                    'priority': 'medium',
                    'title': 'Reduce Vendor Concentration',
                    'description': f'Top vendor handles {top_vendor_share:.1f}% of projects.',
                    'potential_impact': 'Reduced dependency risk',
                    'action_items': [
                        'Evaluate alternative vendors',
                        'Develop contingency plans',
                        'Consider vendor development programs'
                    ]
                })
        
        # Financial performance
        financial = self._get_financial_performance(managed_sectors)
        collection_rate = (
            financial['collected_amount'] / financial['billed_amount'] * 100
            if financial['billed_amount'] > 0 else 0
        )
        
        if collection_rate < 80:
            recommendations.append({
                'category': 'FINANCIAL_MANAGEMENT',
                'priority': 'high',
                'title': 'Improve Collection Rate',
                'description': f'Collection rate at {collection_rate:.1f}% is below target.',
                'potential_impact': 'Could improve cash flow significantly',
                'action_items': [
                    'Review billing processes',
                    'Implement stricter payment terms',
                    'Consider early payment incentives'
                ]
            })
        
        return Response({
            'recommendations': recommendations,
            'generated_at': timezone.now()
        })

class SystemAdministratorViewSet(viewsets.ViewSet):
    """
    System Administrator Portal - Full system management
    """
    permission_classes = [AllowAny]
    
    @action(detail=False, methods=['get'])
    def system_health(self, request):
        """Get comprehensive system health metrics"""
        # User statistics
        total_users = User.objects.count()
        active_users = User.objects.filter(is_active=True).count()
        active_sessions = UserSession.objects.filter(is_active=True).count()
        
        # Database statistics
        db_stats = {
            'projects': Project.objects.count(),
            'vendors': Vendor.objects.count(),
            'documents': ProjectDocument.objects.count(),
            'inspections': QIInspection.objects.count(),
            'work_orders': WorkOrder.objects.count()
        }
        
        # Recent activity
        recent_logins = UserSession.objects.filter(
            login_time__gte=timezone.now() - timedelta(hours=24)
        ).count()
        
        # System errors
        recent_errors = SystemAuditLog.objects.filter(
            status='Failed',
            created_at__gte=timezone.now() - timedelta(hours=24)).count()
        return Response({
        'users': {
            'total': total_users,
            'active': active_users,
            'active_sessions': active_sessions,
            'recent_logins_24h': recent_logins
        },
        'database': db_stats,
        'system_status': 'healthy' if recent_errors == 0 else 'warning',
        'recent_errors_24h': recent_errors,
        'uptime': 'Available',  # Placeholder
        'last_backup': 'N/A'  # Placeholder
        })

    @action(detail=False, methods=['get'])
    def user_management(self, request):
        """Get user management overview"""
        users = User.objects.all().select_related('role')
        
        # Group by role
        by_role = users.values('role__role_name').annotate(
            count=Count('user_id'),
            active=Count('user_id', filter=Q(is_active=True))
        )
        
        # Recent user activity
        recent_activity = UserSession.objects.filter(
            login_time__gte=timezone.now() - timedelta(days=7)
        ).values('user__username', 'user__role__role_name').annotate(
            login_count=Count('session_id')
        ).order_by('-login_count')[:10]
        
        return Response({
            'total_users': users.count(),
            'by_role': list(by_role),
            'recent_activity': list(recent_activity),
            'inactive_users': users.filter(is_active=False).count()
        })

    @action(detail=False, methods=['post'])
    def create_user(self, request):
        """Create new user account"""
        serializer = RegisterUserSerializer(data=request.data)
        
        if serializer.is_valid():
            user = serializer.save()
            
            # Log action
            SystemAuditLog.objects.create(
                user=request.user,
                action_type='USER_CREATED',
                action_description=f'Created user: {user.username}',
                entity_type='User',
                entity_id=user.user_id,
                status='Success'
            )
            
            return Response({
                'status': 'created',
                'user_id': user.user_id,
                'username': user.username
            }, status=201)
        
        return Response(serializer.errors, status=400)

    @action(detail=False, methods=['post'])
    def manage_user(self, request):
        """Manage user account (activate, deactivate, reset password)"""
        action = request.data.get('action')
        user_id = request.data.get('user_id')
        
        try:
            user = User.objects.get(user_id=user_id)
            
            if action == 'activate':
                user.is_active = True
                user.save()
                action_desc = f'Activated user: {user.username}'
                
            elif action == 'deactivate':
                user.is_active = False
                user.save()
                
                # End all sessions
                UserSession.objects.filter(
                    user=user,
                    is_active=True
                ).update(
                    is_active=False,
                    logout_time=timezone.now()
                )
                action_desc = f'Deactivated user: {user.username}'
                
            elif action == 'reset_password':
                new_password = request.data.get('new_password', 'TempPassword123')
                user.set_password(new_password)
                user.save()
                action_desc = f'Reset password for: {user.username}'
                
            elif action == 'delete':
                username = user.username
                user.delete()
                action_desc = f'Deleted user: {username}'
                
            else:
                return Response({'error': 'Invalid action'}, status=400)
            
            # Log action
            SystemAuditLog.objects.create(
                user=request.user,
                action_type='USER_MANAGEMENT',
                action_description=action_desc,
                entity_type='User',
                entity_id=user_id,
                status='Success'
            )
            
            return Response({'status': action})
            
        except User.DoesNotExist:
            return Response({'error': 'User not found'}, status=404)

    @action(detail=False, methods=['get'])
    def audit_logs(self, request):
        """Get system audit logs"""
        days = int(request.query_params.get('days', 7))
        action_type = request.query_params.get('action_type')
        user_id = request.query_params.get('user_id')
        
        start_date = timezone.now() - timedelta(days=days)
        
        logs = SystemAuditLog.objects.filter(
            created_at__gte=start_date
        ).select_related('user')
        
        if action_type:
            logs = logs.filter(action_type=action_type)
        
        if user_id:
            logs = logs.filter(user_id=user_id)
        
        logs = logs.order_by('-created_at')[:500]  # Limit to 500 records
        
        serializer = SystemAuditLogSerializer(logs, many=True)
        return Response({
            'logs': serializer.data,
            'total': logs.count()
        })

    @action(detail=False, methods=['get'])
    def system_settings(self, request):
        """Get system settings"""
        settings = SystemSetting.objects.all()
        serializer = SystemSettingSerializer(settings, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=['post'])
    def update_setting(self, request):
        """Update system setting"""
        setting_key = request.data.get('setting_key')
        setting_value = request.data.get('setting_value')
        
        try:
            setting = SystemSetting.objects.get(setting_key=setting_key)
            
            if not setting.is_editable:
                return Response({'error': 'Setting is not editable'}, status=403)
            
            old_value = setting.setting_value
            setting.setting_value = setting_value
            setting.save()
            
            # Log change
            ChangeLog.objects.create(
                table_name='system_settings',
                record_id=setting.setting_id,
                change_type='UPDATE',
                field_name='setting_value',
                old_value=old_value,
                new_value=setting_value,
                changed_by=request.user
            )
            
            return Response({'status': 'updated', 'setting_key': setting_key})
            
        except SystemSetting.DoesNotExist:
            return Response({'error': 'Setting not found'}, status=404)

    @action(detail=False, methods=['post'])
    def database_backup(self, request):
        """Trigger database backup (placeholder)"""
        # This would typically trigger an actual backup process
        # For now, just log the action
        
        SystemAuditLog.objects.create(
            user=request.user,
            action_type='DATABASE_BACKUP',
            action_description='Database backup initiated',
            status='Success'
        )
        
        return Response({
            'status': 'backup_initiated',
            'timestamp': timezone.now(),
            'message': 'Database backup process started'
        })

    @action(detail=False, methods=['get'])
    def security_report(self, request):
        """Get security report"""
        days = int(request.query_params.get('days', 7))
        start_date = timezone.now() - timedelta(days=days)
        
        # Failed login attempts
        failed_logins = SystemAuditLog.objects.filter(
            action_type='LOGIN_ATTEMPT',
            status='Failed',
            created_at__gte=start_date
        ).count()
        
        # Unauthorized access attempts
        unauthorized = SystemAuditLog.objects.filter(
            action_type__icontains='UNAUTHORIZED',
            created_at__gte=start_date
        ).count()
        
        # Suspicious activities
        suspicious = SystemAuditLog.objects.filter(
            status='Failed',
            created_at__gte=start_date
        ).values('user__username', 'action_type').annotate(
            count=Count('id')
        ).filter(count__gte=5)  # 5 or more failures
        
        # Active sessions by user
        session_stats = UserSession.objects.filter(
            is_active=True
        ).values('user__username', 'user__role__role_name').annotate(
            session_count=Count('session_id')
        )
        
        return Response({
            'period_days': days,
            'failed_logins': failed_logins,
            'unauthorized_attempts': unauthorized,
            'suspicious_activities': list(suspicious),
            'active_sessions_by_user': list(session_stats),
            'security_status': 'green' if failed_logins < 10 else 'yellow' if failed_logins < 50 else 'red'
        })

    @action(detail=False, methods=['get'])
    def performance_metrics(self, request):
        """Get system performance metrics"""
        # API response times (placeholder - would need actual monitoring)
        # Database query performance (placeholder)
        # Active connections (placeholder)
        
        return Response({
            'api_response_time_ms': 120,  # Placeholder
            'database_query_time_ms': 45,  # Placeholder
            'active_connections': UserSession.objects.filter(is_active=True).count(),
            'total_requests_24h': SystemAuditLog.objects.filter(
                created_at__gte=timezone.now() - timedelta(hours=24)
            ).count(),
            'error_rate': 0.5,  # Placeholder
            'system_load': 'normal'  # Placeholder
        })





class CalendarDashboardViewSet(viewsets.ViewSet):
    """Calendar and deadline dashboard"""
    permission_classes = [AllowAny]

    @action(detail=False, methods=['get'], url_path='upcoming-deadlines')
    def upcoming_deadlines(self, request):
        """
        Get all upcoming deadlines from various sources with enhanced filtering
        """
        try:
            today = timezone.now().date()
            
            # Get date range from query params (default 90 days)
            days_ahead = int(request.query_params.get('days', 90))
            end_date = today + timedelta(days=days_ahead)
            
            # Get filter parameters
            filter_type = request.query_params.get('type', 'all')  # all, project, deadline, sla, inspection
            priority = request.query_params.get('priority', None)
            
            deadlines = []
            
            # 1. Work Order Deadlines
            if filter_type in ['all', 'deadline']:
                work_orders = WorkOrder.objects.filter(
                    date_received_by_vc__range=[today, end_date]
                )
                
               
                for wo in work_orders:
                    days_remaining = (wo.date_received_by_vc - today).days if wo.date_received_by_vc else 0
                    deadlines.append({
                        'id': f'wo-{wo.id}',
                        'date': wo.date_received_by_vc.isoformat() if wo.date_received_by_vc else None,
                        'type': 'deadline',
                        'title': 'Work Order Completion',
                        'description': f"{wo.wo_no} - {wo.description or 'No Description'}",
                        
                        'status': wo.status,
                        'project_code': wo.wo_no,
                        'days_remaining': days_remaining,
                        'is_overdue': days_remaining < 0,
                        'assigned_to': wo.supervisor_full_name if wo.supervisor_full_name else None
                    })
            
            # 2. Project Completion Dates
            if filter_type in ['all', 'project']:
                projects = Project.objects.filter(
                    completion_date__range=[today, end_date],
                    status__status_name__in=['In Progress', 'Active']
                ).select_related('vendor', 'assigned_engineer', 'status')
                
                if priority:
                    projects = projects.filter(priority=priority)
                
                for proj in projects:
                    days_remaining = (proj.completion_date - today).days if proj.completion_date else 0
                    deadlines.append({
                        'id': f'proj-{proj.project_id}',
                        'date': proj.completion_date.isoformat() if proj.completion_date else None,
                        'type': 'project',
                        'title': 'Project Completion',
                        'description': f"{proj.project_code} - {proj.project_name}",
                        'priority': proj.priority,
                        'status': proj.status.status_name if proj.status else 'Unknown',
                        'project_code': proj.project_code,
                        'days_remaining': days_remaining,
                        'is_overdue': days_remaining < 0,
                        'assigned_to': proj.assigned_engineer.get_full_name() if proj.assigned_engineer else None
                    })
            
            # 3. SLA Tracking Deadlines
            if filter_type in ['all', 'sla']:
                sla_items = SLATracking.objects.filter(
                    due_date__range=[today, end_date],
                    status='Open'
                ).select_related('project', 'sla_rule')
                
                for sla in sla_items:
                    days_remaining = (sla.due_date - today).days if sla.due_date else 0
                    priority_level = 'Critical' if days_remaining <= 2 else 'High' if days_remaining <= 5 else 'Medium'
                    
                    deadlines.append({
                        'id': f'sla-{sla.sla_tracking_id}',
                        'date': sla.due_date.isoformat() if sla.due_date else None,
                        'type': 'sla',
                        'title': 'SLA Deadline',
                        'description': f"{sla.project.project_code} - {sla.project.project_name}",
                        'priority': priority_level,
                        'status': sla.status,
                        'project_code': sla.project.project_code,
                        'days_remaining': days_remaining,
                        'is_overdue': days_remaining < 0,
                        'assigned_to': None
                    })
            
            # 4. Document Compliance Deadlines
            if filter_type in ['all', 'deadline']:
                doc_compliance = DocumentCompliance.objects.filter(
                    due_date__range=[today, end_date],
                    is_submitted=False
                ).select_related('project', 'doc_type')
                
                for doc in doc_compliance:
                    days_remaining = (doc.due_date - today).days if doc.due_date else 0
                    priority_level = 'Critical' if days_remaining <= 2 else 'High'
                    
                    deadlines.append({
                        'id': f'doc-{doc.id}',
                        'date': doc.due_date.isoformat() if doc.due_date else None,
                        'type': 'deadline',
                        'title': f'Document: {doc.doc_type.doc_type_name}',
                        'description': f"{doc.project.project_code} - {doc.project.project_name}",
                        'priority': priority_level,
                        'status': 'Pending Submission',
                        'project_code': doc.project.project_code,
                        'days_remaining': days_remaining,
                        'is_overdue': days_remaining < 0,
                        'assigned_to': None
                    })
            
            # 5. QI Inspections
            if filter_type in ['all', 'inspection']:
                inspections = QIInspection.objects.filter(
                    scheduled_date__range=[today, end_date],
                    is_completed=False
                ).select_related('project', 'assigned_qi', 'inspection_type')
                
                for inspection in inspections:
                    days_remaining = (inspection.scheduled_date - today).days if inspection.scheduled_date else 0
                    
                    deadlines.append({
                        'id': f'qi-{inspection.inspection_id}',
                        'date': inspection.scheduled_date.isoformat() if inspection.scheduled_date else None,
                        'type': 'inspection',
                        'title': f'QI Inspection: {inspection.inspection_type.inspection_name}',
                        'description': f"{inspection.project.project_code} - {inspection.project.project_name}",
                        'priority': 'High',
                        'status': 'Scheduled',
                        'project_code': inspection.project.project_code,
                        'days_remaining': days_remaining,
                        'is_overdue': days_remaining < 0,
                        'assigned_to': inspection.assigned_qi.get_full_name() if inspection.assigned_qi else None
                    })
            
            # Sort by days remaining (most urgent first)
            deadlines.sort(key=lambda x: (x['is_overdue'], x['days_remaining']))
            
            return Response(deadlines)
            
        except Exception as e:
            return Response(
                {'error': str(e), 'detail': 'Failed to fetch upcoming deadlines'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    @action(detail=False, methods=['get'], url_path='calendar-stats')
    def calendar_stats(self, request):
        """Get statistics for calendar dashboard"""
        try:
            today = timezone.now().date()
            week_from_now = today + timedelta(days=7)
            
            # Count different types of events
            stats = {
                'total_events': 0,
                'overdue': 0,
                'this_week': 0,
                'by_type': {
                    'project': 0,
                    'deadline': 0,
                    'sla': 0,
                    'inspection': 0
                },
                'by_priority': {
                    'Critical': 0,
                    'High': 0,
                    'Medium': 0,
                    'Low': 0
                }
            }
            
            # Work Orders
            work_orders = WorkOrder.objects.filter(
                date_received_by_vc__isnull=False
            )
            
            for wo in work_orders:
                stats['total_events'] += 1
                stats['by_type']['deadline'] += 1
                
                if wo.date_received_by_vc < today:
                    stats['overdue'] += 1
                elif wo.date_received_by_vc <= week_from_now:
                    stats['this_week'] += 1
                
                
            
            # Projects
            projects = Project.objects.filter(
                completion_date__isnull=False,
                status__status_name__in=['In Progress', 'Active']
            )
            
            for proj in projects:
                stats['total_events'] += 1
                stats['by_type']['project'] += 1
                
                if proj.completion_date < today:
                    stats['overdue'] += 1
                elif proj.completion_date <= week_from_now:
                    stats['this_week'] += 1
                
                priority = proj.priority or 'Medium'
                stats['by_priority'][priority] = stats['by_priority'].get(priority, 0) + 1
            
            # SLA Deadlines
            sla_count = SLATracking.objects.filter(
                status='Open',
                due_date__isnull=False
            ).count()
            stats['by_type']['sla'] = sla_count
            stats['total_events'] += sla_count
            
            # QI Inspections
            inspection_count = QIInspection.objects.filter(
                is_completed=False,
                scheduled_date__isnull=False
            ).count()
            stats['by_type']['inspection'] = inspection_count
            stats['total_events'] += inspection_count
            
            return Response(stats)
            
        except Exception as e:
            return Response(
                {'error': str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


@api_view(['GET'])
@permission_classes([AllowAny])
def get_calendar_events(request):
    """Standalone endpoint for calendar events"""
    viewset = CalendarDashboardViewSet()
    viewset.request = request
    return viewset.upcoming_deadlines(request)


@api_view(['GET'])
@permission_classes([AllowAny])
def get_calendar_stats(request):
    """Standalone endpoint for calendar statistics"""
    viewset = CalendarDashboardViewSet()
    viewset.request = request
    return viewset.calendar_stats(request)

from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from django.utils import timezone
from datetime import datetime, timedelta

class VendorDailyActivityViewSet(viewsets.ModelViewSet):
    queryset = VendorDailyActivity.objects.all()
    serializer_class = VendorDailyActivitySerializer
    
    def get_queryset(self):
        queryset = super().get_queryset()
        
        # Filter by vendor if user is a vendor
        user = self.request.user
        if hasattr(user, 'role') and user.role and user.role.role_name == 'vendor':
            # Assuming vendor users are linked to a vendor
            vendor_id = self.request.query_params.get('vendor_id')
            if vendor_id:
                queryset = queryset.filter(vendor_id=vendor_id)
        
        # Filter by date
        date_param = self.request.query_params.get('date')
        if date_param:
            queryset = queryset.filter(activity_date=date_param)
        
        # Filter by status
        status_param = self.request.query_params.get('status')
        if status_param:
            queryset = queryset.filter(status=status_param)
        
        return queryset
    
    @action(detail=False, methods=['get'])
    def today_activities(self, request):
        """Get today's activities for the vendor"""
        today = timezone.now().date()
        vendor_id = request.query_params.get('vendor_id')
        
        activities = self.queryset.filter(
            activity_date=today,
            vendor_id=vendor_id
        )
        
        serializer = self.get_serializer(activities, many=True)
        return Response(serializer.data)
    
    @action(detail=False, methods=['get'])
    def weekly_summary(self, request):
        """Get weekly summary of activities"""
        vendor_id = request.query_params.get('vendor_id')
        end_date = timezone.now().date()
        start_date = end_date - timedelta(days=7)
        
        activities = self.queryset.filter(
            vendor_id=vendor_id,
            activity_date__gte=start_date,
            activity_date__lte=end_date
        )
        
        summary = {
            'total_activities': activities.count(),
            'signed_on': activities.filter(status='SIGNED_ON').count(),
            'in_progress': activities.filter(status='IN_PROGRESS').count(),
            'completed': activities.filter(status='COMPLETED').count(),
            'caution_count': activities.filter(has_caution=True).count(),
            'activities': VendorDailyActivitySerializer(activities, many=True).data
        }
        
        return Response(summary)
    
    @action(detail=True, methods=['post'])
    def upload_photo(self, request, pk=None):
        """Upload a photo for an activity"""
        activity = self.get_object()
        
        photo_type = request.data.get('photo_type', 'SIGN_ON')
        caption = request.data.get('caption', '')
        photo_file = request.FILES.get('photo')
        user_id = request.data.get('uploaded_by')
        
        if not photo_file:
            return Response(
                {'error': 'No photo file provided'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        photo = VendorActivityPhoto.objects.create(
            activity=activity,
            photo_type=photo_type,
            photo_file=photo_file,
            caption=caption,
            uploaded_by=user_id
        )
        
        serializer = VendorActivityPhotoSerializer(photo)
        return Response(serializer.data, status=status.HTTP_201_CREATED)
    
    @action(detail=True, methods=['patch'])
    def mark_completed(self, request, pk=None):
        """Mark activity as completed"""
        activity = self.get_object()
        
        activity.status = 'COMPLETED'
        activity.completed_at = timezone.now()
        activity.completion_notes = request.data.get('completion_notes', '')
        activity.save()
        
        serializer = self.get_serializer(activity)
        return Response(serializer.data)


class VendorActivityPhotoViewSet(viewsets.ModelViewSet):
    queryset = VendorActivityPhoto.objects.all()
    serializer_class = VendorActivityPhotoSerializer
    
    def get_queryset(self):
        queryset = super().get_queryset()
        
        activity_id = self.request.query_params.get('activity_id')
        if activity_id:
            queryset = queryset.filter(activity_id=activity_id)
        
        return queryset




class ClerkDocumentValidationViewSet(viewsets.ViewSet):
    """ViewSet for clerk document validation workflow"""
    permission_classes = [AllowAny]
    
    @action(detail=False, methods=['get'])
    def pending_projects(self, request):
        """Get projects with status 3 (Completed, awaiting docs)"""
        # Status 3 = Completed, awaiting document validation
        projects = Project.objects.filter(
            status_id=3
        ).select_related('vendor', 'status').prefetch_related('documents')
        
        serializer = ProjectValidationListSerializer(projects, many=True)
        return Response({
            'results': serializer.data,
            'count': projects.count()
        })
    
    @action(detail=True, methods=['get'])
    def project_documents(self, request, pk=None):
        """Get all documents for a specific project"""
        try:
            project = Project.objects.get(project_id=pk)
        except Project.DoesNotExist:
            return Response(
                {'error': 'Project not found'}, 
                status=status.HTTP_404_NOT_FOUND
            )
        
        documents = ProjectDocument.objects.filter(
            project=project
        ).select_related('doc_type', 'uploaded_by', 'approved_by')
        
        serializer = ProjectDocumentListSerializer(documents, many=True)
        return Response({
            'results': serializer.data,
            'count': documents.count()
        })
    
    @action(detail=True, methods=['patch'])
    def validate_document(self, request, pk=None):
        """Validate a single document (approve/reject)"""
        try:
            document = ProjectDocument.objects.get(document_id=pk)
        except ProjectDocument.DoesNotExist:
            return Response(
                {'error': 'Document not found'}, 
                status=status.HTTP_404_NOT_FOUND
            )
        
        serializer = ProjectDocumentValidationSerializer(
            document, 
            data=request.data, 
            partial=True,
            context={'request': request}
        )
        
        if serializer.is_valid():
            serializer.save()
            
            # Return updated document
            response_serializer = ProjectDocumentListSerializer(document)
            return Response(response_serializer.data)
        
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
    @action(detail=True, methods=['post'])
    def generate_confirmation(self, request, pk=None):
        """Generate confirmation number and update project status"""
        try:
            project = Project.objects.get(project_id=pk)
        except Project.DoesNotExist:
            return Response(
                {'error': 'Project not found'}, 
                status=status.HTTP_404_NOT_FOUND
            )
        
        # Check if all documents are approved
        pending_docs = project.documents.filter(
            Q(approval_status='Pending') | Q(approval_status='Rejected')
        ).count()
        
        if pending_docs > 0:
            return Response(
                {'error': 'All documents must be approved before generating confirmation'}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Generate confirmation number
        timestamp = int(timezone.now().timestamp())
        confirmation_number = f"CONF-{timestamp:X}-{project.project_id}"
        
        # Update project status to 4 (Awaiting Documents Review)
        project.status_id = 4
        project.save()
        
        # Create notification for vendor
        if project.vendor and project.vendor.user:
            Notification.objects.create(
                recipient_user=project.vendor.user,
                notification_type='Email',
                subject='Document Validation Complete',
                message=f'Your documents for project {project.project_code} have been validated. Confirmation Number: {confirmation_number}',
                related_project=project,
                status='Pending'
            )
        
        return Response({
            'confirmation_number': confirmation_number,
            'project_id': project.project_id,
            'project_code': project.project_code,
            'status': 'success'
        })
    
    @action(detail=False, methods=['get'])
    def validation_stats(self, request):
        """Get validation statistics"""
        today = timezone.now().date()
        
        # Projects pending validation
        pending = Project.objects.filter(status_id=3).count()
        
        # Documents validated today
        validated_today = ProjectDocument.objects.filter(
            approval_date__date=today
        ).count()
        
        # Documents with quality issues (rejected or large file size)
        issues = ProjectDocument.objects.filter(
            Q(approval_status='Rejected') |
            Q(file_size__gt=10485760)  # > 10MB
        ).filter(
            upload_date__date=today
        ).count()
        
        # Total documents pending
        total_pending = ProjectDocument.objects.filter(
            approval_status='Pending'
        ).count()
        
        stats = {
            'pending_validation': pending,
            'validated_today': validated_today,
            'issues_found': issues,
            'total_documents': total_pending
        }
        
        serializer = DocumentValidationStatsSerializer(stats)
        return Response(serializer.data)
    
    @action(detail=False, methods=['get'])
    def quality_check(self, request):
        """Get documents with quality issues"""
        # Find documents with quality issues
        documents = ProjectDocument.objects.filter(
            Q(file_size__gt=10485760) |  # > 10MB
            Q(approval_status='Rejected')
        ).select_related('project', 'doc_type')[:50]
        
        serializer = ProjectDocumentListSerializer(documents, many=True)
        return Response({
            'results': serializer.data,
            'count': documents.count()
        })
        
# Add these imports at the top
from django.db.models.signals import post_save, pre_save
from django.dispatch import receiver
from .email_notification_service import email_service

# ==================== SIGNAL HANDLERS FOR EMAIL NOTIFICATIONS ====================

# Store previous state for comparison
_work_order_cache = {}
_project_cache = {}

@receiver(pre_save, sender=WorkOrder)
def work_order_pre_save(sender, instance, **kwargs):
    """Cache old state before save"""
    if instance.pk:
        try:
            old = WorkOrder.objects.get(pk=instance.pk)
            _work_order_cache[instance.pk] = {
                'status': old.status,
            }
        except WorkOrder.DoesNotExist:
            pass

@receiver(post_save, sender=WorkOrder)
def work_order_post_save(sender, instance, created, **kwargs):
    """Send email notifications for work order changes"""
    
    if created:
        # New work order created
        email_service.notify_new_work_order(instance)
    else:
        # Work order updated - check for status change
        old_data = _work_order_cache.get(instance.pk)
        if old_data and old_data['status'] != instance.status:
            email_service.notify_work_order_status_change(
                instance, 
                old_data['status'], 
                instance.status
            )
        
        # Clean up cache
        if instance.pk in _work_order_cache:
            del _work_order_cache[instance.pk]





@receiver(pre_save, sender=Project)
def project_pre_save(sender, instance, **kwargs):
    """Cache old project state"""
    if instance.pk:
        try:
            old = Project.objects.get(pk=instance.pk)
            _project_cache[instance.pk] = {
                'status': old.status_id,
            }
        except Project.DoesNotExist:
            pass

@receiver(post_save, sender=Project)
def project_post_save(sender, instance, created, **kwargs):
    """Send email notifications for project changes"""
    
    if created:
        # New project created
        email_service.notify_new_project(instance)
    else:
        old_data = _project_cache.get(instance.pk)
        
        if old_data:
            # Check if project became delayed
            if not instance.is_delayed:
                email_service.notify_project_delay(instance)
            
            # Check if project marked as completed
            current_status = instance.status if instance.status else None
            if old_data['status'] != 3 and current_status == 3:
                email_service.notify_project_completion(instance)
        
        # Clean up cache
        if instance.pk in _project_cache:
            del _project_cache[instance.pk]


@receiver(post_save, sender=ProjectDocument)
def document_post_save(sender, instance, created, **kwargs):
    """Send email notifications for document changes"""
    
    if created:
        # New document uploaded
        email_service.notify_document_uploaded(instance)
    else:
        # Document updated - check approval status
        old_doc = ProjectDocument.objects.filter(pk=instance.pk).first()
        
        if instance.approval_status == 'Approved':
            email_service.notify_document_approved(instance)
        elif instance.approval_status == 'Rejected':
            email_service.notify_document_rejected(instance)


@receiver(post_save, sender=QIInspection)
def inspection_post_save(sender, instance, created, **kwargs):
    """Send email notifications for inspection changes"""
    
    if created:
        # New inspection scheduled
        email_service.notify_inspection_scheduled(instance)
    else:
        # Check if inspection completed
        if instance.is_completed:
            email_service.notify_inspection_completed(instance)


@receiver(post_save, sender=Penalty)
def penalty_post_save(sender, instance, created, **kwargs):
    """Send email notification when penalty issued"""
    
    # Only notify when penalty is officially issued
    if instance.penalty_status == 'Issued':
        email_service.notify_penalty_issued(instance)


@receiver(post_save, sender=SLATracking)
def sla_post_save(sender, instance, created, **kwargs):
    """Send email notification for SLA breach"""
    
    # Notify when SLA is breached
    if instance.is_breached and instance.status == 'Breached':
        email_service.notify_sla_breach(instance)
        
        
from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import AllowAny
from django.db.models import Q, Count
from django.utils import timezone
from datetime import date, timedelta
from .models import (
    InspectionFlag,
    InspectionChecklistItem,
    DefectReport,
    DefectCorrectionHistory,
    QIInspection
)
from .serializers import (
    InspectionFlagSerializer,
    InspectionChecklistItemSerializer,
    DefectReportSerializer,
    DefectReportCreateSerializer,
    DefectCorrectionHistorySerializer,
    AIDefectSuggestionSerializer
)


# ==================== INSPECTION CHECKLIST VIEWSET ====================
class InspectionChecklistItemViewSet(viewsets.ModelViewSet):
    """Manage inspection checklist items"""
    queryset = InspectionChecklistItem.objects.all()
    serializer_class = InspectionChecklistItemSerializer
    permission_classes = [AllowAny]
    
    def get_queryset(self):
        queryset = super().get_queryset()
        
        inspection_id = self.request.query_params.get('inspection')
        status_filter = self.request.query_params.get('status')
        
        if inspection_id:
            queryset = queryset.filter(inspection_id=inspection_id)
        if status_filter:
            queryset = queryset.filter(status=status_filter)
        
        return queryset
    
    @action(detail=False, methods=['post'])
    def bulk_update(self, request):
        """Bulk update checklist items (for QI mobile app)"""
        items_data = request.data.get('items', [])
        
        updated = []
        for item_data in items_data:
            try:
                item = InspectionChecklistItem.objects.get(id=item_data['id'])
                item.status = item_data.get('status', item.status)
                item.notes = item_data.get('notes', item.notes)
                item.photos = item_data.get('photos', item.photos)
                item.checked_at = timezone.now()
                item.checked_by = request.user
                item.save()
                updated.append(item.id)
            except InspectionChecklistItem.DoesNotExist:
                continue
        
        return Response({
            'updated_count': len(updated),
            'updated_ids': updated
        })


# ==================== INSPECTION FLAG VIEWSET ====================
class InspectionFlagViewSet(viewsets.ModelViewSet):
    """Manage inspection flags (system-generated alerts)"""
    queryset = InspectionFlag.objects.all()
    serializer_class = InspectionFlagSerializer
    permission_classes = [AllowAny]
    
    def get_queryset(self):
        queryset = super().get_queryset()
        
        status_filter = self.request.query_params.get('status')
        inspection_id = self.request.query_params.get('inspection')
        qi_id = self.request.query_params.get('qi')
        
        if status_filter:
            queryset = queryset.filter(status=status_filter)
        if inspection_id:
            queryset = queryset.filter(inspection_id=inspection_id)
        if qi_id:
            queryset = queryset.filter(inspection__assigned_qi_id=qi_id)
        
        return queryset.select_related('inspection', 'reviewed_by')
    
    @action(detail=False, methods=['get'])
    def pending_review(self, request):
        """Get all flags pending QI review"""
        qi_id = request.query_params.get('qi_id')
        
        flags = self.queryset.filter(
            status='PENDING_QI_REVIEW',
            requires_action=True
        )
        
        if qi_id:
            flags = flags.filter(inspection__assigned_qi_id=qi_id)
        
        serializer = self.get_serializer(flags, many=True)
        return Response(serializer.data)
    
    @action(detail=True, methods=['post'])
    def generate_ai_suggestions(self, request, pk=None):
        """Generate AI suggestions for defect grouping"""
        flag = self.get_object()
        
        # Get failed items
        failed_items = InspectionChecklistItem.objects.filter(
            inspection=flag.inspection,
            status='Fail'
        )
        
        # AI logic to group and suggest defects
        suggestions = self._generate_defect_suggestions(failed_items)
        
        # Save suggestions to flag
        flag.ai_suggestions = {'suggestions': suggestions}
        flag.save()
        
        return Response({
            'suggestions': suggestions,
            'total_suggestions': len(suggestions)
        })
    
    def _generate_defect_suggestions(self, failed_items):
        """
        AI/Rule-based logic to group failed items into suggested defects
        This is simplified - in production, use ML models
        """
        suggestions = []
        
        # Group by category
        categories = {}
        for item in failed_items:
            cat = item.item_category or 'General'
            if cat not in categories:
                categories[cat] = []
            categories[cat].append(item)
        
        # Create suggestions for each category
        for category, items in categories.items():
            item_names = [item.item_name for item in items]
            
            # Determine severity based on keywords
            severity = 'MINOR'
            critical_keywords = ['safety', 'structural', 'electrical', 'fire']
            major_keywords = ['code', 'violation', 'non-compliance']
            
            combined_text = ' '.join(item_names).lower()
            if any(keyword in combined_text for keyword in critical_keywords):
                severity = 'CRITICAL'
            elif any(keyword in combined_text for keyword in major_keywords):
                severity = 'MAJOR'
            
            suggestions.append({
                'suggested_defect_type': f"{category} Non-Compliance",
                'suggested_severity': severity,
                'suggested_description': f"Multiple {category.lower()} issues found: {', '.join(item_names[:3])}{'...' if len(item_names) > 3 else ''}",
                'related_item_ids': [item.id for item in items],
                'confidence_score': 0.85 if len(items) > 2 else 0.65,
                'reasoning': f"Grouped {len(items)} related failures in {category} category"
            })
        
        return suggestions
    
    @action(detail=True, methods=['patch'])
    def dismiss(self, request, pk=None):
        """QI dismisses flag without creating defects"""
        flag = self.get_object()
        
        reason = request.data.get('reason', '')
        
        flag.status = 'DISMISSED'
        flag.reviewed_at = timezone.now()
        flag.reviewed_by = request.user
        flag.ai_suggestions['dismissal_reason'] = reason
        flag.save()
        
        return Response({'status': 'dismissed'})


# ==================== DEFECT REPORT VIEWSET ====================
class DefectReportViewSet(viewsets.ModelViewSet):
    """Manage formal defect reports"""
    queryset = DefectReport.objects.all()
    permission_classes = [IsAuthenticated]
    search_fields = ['created_by']
    filterset_fields = ['created_by']
    
    @action(detail=False, methods=['post'])
    def finalize_from_flag(self, request):
        """
        QI finalizes defect reports from inspection flag
        This is the HYBRID approach - QI reviews AI suggestions and creates formal defects
        """
        flag_id = request.data.get('flag_id')
        defects_data = request.data.get('defects', [])
        qi_signature = request.data.get('qi_signature')
        
        if not qi_signature:
            return Response(
                {'error': 'QI signature required'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            flag = InspectionFlag.objects.get(id=flag_id)
        except InspectionFlag.DoesNotExist:
            return Response(
                {'error': 'Flag not found'},
                status=status.HTTP_404_NOT_FOUND
            )
        
        created_defects = []
        
        for defect_data in defects_data:
            # Calculate due date (7 days default)
            due_date = date.today() + timedelta(days=7)
            if defect_data.get('severity') == 'CRITICAL':
                due_date = date.today() + timedelta(days=3)
            
            defect = DefectReport.objects.create(
                inspection=flag.inspection,
                project=flag.inspection.project,
                defect_type=defect_data['defect_type'],
                defect_category=defect_data.get('defect_category', ''),
                severity=defect_data['severity'],
                description=defect_data['description'],
                related_checklist_items=defect_data.get('related_checklist_items', []),
                photos=defect_data.get('photos', []),
                location_gps=defect_data.get('location_gps', ''),
                qi_notes=defect_data.get('qi_notes', ''),
                qi_signature=qi_signature,
                created_by=request.user,
                correction_due_date=due_date,
                correction_status='OPEN'
            )
            created_defects.append(defect)
            
            # 🔥 ADD THIS SECTION HERE - Audit trail for AI modifications 🔥
            
            # Check if this was an AI-generated suggestion and log any modifications
            if defect_data.get('is_ai_generated'):
                # Find the original AI suggestion that matches this defect
                original_suggestion = None
                if flag.ai_suggestions and 'suggestions' in flag.ai_suggestions:
                    for suggestion in flag.ai_suggestions['suggestions']:
                        # Match by related checklist items
                        if set(suggestion.get('related_item_ids', [])) == set(defect.related_checklist_items):
                            original_suggestion = suggestion
                            break
                
                if original_suggestion:
                    changes = []
                    
                    # Check what the QI changed from AI suggestions
                    if original_suggestion.get('suggested_defect_type') != defect.defect_type:
                        changes.append(
                            f"Type changed from '{original_suggestion.get('suggested_defect_type')}' to '{defect.defect_type}'"
                        )
                    
                    if original_suggestion.get('suggested_severity') != defect.severity:
                        changes.append(
                            f"Severity changed from {original_suggestion.get('suggested_severity')} to {defect.severity}"
                        )
                    
                    if original_suggestion.get('suggested_description') != defect.description:
                        changes.append("Description modified by QI")
                    
                    # Create history entry with changes
                    if changes:
                        DefectCorrectionHistory.objects.create(
                            defect=defect,
                            action='SUBMITTED',
                            action_by=request.user,
                            notes=f"🤖 AI suggestion reviewed and modified. Changes: {'; '.join(changes)}"
                        )
                    else:
                        # QI accepted AI suggestion without changes
                        DefectCorrectionHistory.objects.create(
                            defect=defect,
                            action='SUBMITTED',
                            action_by=request.user,
                            notes=f"🤖 AI suggestion accepted without modifications (Confidence: {original_suggestion.get('confidence_score', 0)*100:.0f}%)"
                        )
                else:
                    # Marked as AI-generated but no matching suggestion found
                    DefectCorrectionHistory.objects.create(
                        defect=defect,
                        action='SUBMITTED',
                        action_by=request.user,
                        notes="🤖 Based on AI suggestions but significantly modified by QI"
                    )
            else:
                # Manually created defect (not from AI suggestion)
                DefectCorrectionHistory.objects.create(
                    defect=defect,
                    action='SUBMITTED',
                    action_by=request.user,
                    notes='✍️ Manually created defect by QI (not from AI suggestions)'
                )
            
            # 🔥 END OF AUDIT TRAIL SECTION 🔥
        
        # Mark flag as resolved
        flag.status = 'RESOLVED'
        flag.reviewed_at = timezone.now()
        flag.reviewed_by = request.user
        flag.save()
        
        # Send notifications to vendor
        from .email_notification_service import email_service
        for defect in created_defects:
            try:
                email_service.notify_defect_created(defect)
            except Exception as e:
                print(f"Failed to send notification: {e}")
        
        serializer = DefectReportSerializer(created_defects, many=True)
        return Response({
            'created_count': len(created_defects),
            'defects': serializer.data
        }, status=status.HTTP_201_CREATED)
    
    def get_serializer_class(self):
        if self.action in ['create', 'finalize_from_flag']:
            return DefectReportCreateSerializer
        return DefectReportSerializer
    
    def get_queryset(self):
        queryset = super().get_queryset()
        
        status_filter = self.request.query_params.get('correction_status')
        severity = self.request.query_params.get('severity')
        project_id = self.request.query_params.get('project')
        vendor_id = self.request.query_params.get('vendor')
        qi_id = self.request.query_params.get('qi')
        
        if status_filter:
            queryset = queryset.filter(correction_status=status_filter)
        if severity:
            queryset = queryset.filter(severity=severity)
        if project_id:
            queryset = queryset.filter(project_id=project_id)
        if vendor_id:
            queryset = queryset.filter(project__vendor_id=vendor_id)
        if qi_id:
            queryset = queryset.filter(created_by_id=qi_id)
        
        return queryset.select_related(
            'inspection',
            'project',
            'created_by',
            'reviewed_by'
        ).prefetch_related('correction_history')
    
    @action(detail=False, methods=['post'])
    def finalize_from_flag(self, request):
        """
        QI finalizes defect reports from inspection flag
        This is the HYBRID approach - QI reviews AI suggestions and creates formal defects
        """
        flag_id = request.data.get('flag_id')
        defects_data = request.data.get('defects', [])
        qi_signature = request.data.get('qi_signature')
        
        if not qi_signature:
            return Response(
                {'error': 'QI signature required'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            flag = InspectionFlag.objects.get(id=flag_id)
        except InspectionFlag.DoesNotExist:
            return Response(
                {'error': 'Flag not found'},
                status=status.HTTP_404_NOT_FOUND
            )
        
        created_defects = []
        
        for defect_data in defects_data:
            # Calculate due date (7 days default)
            due_date = date.today() + timedelta(days=7)
            if defect_data.get('severity') == 'CRITICAL':
                due_date = date.today() + timedelta(days=3)
            
            defect = DefectReport.objects.create(
                inspection=flag.inspection,
                project=flag.inspection.project,
                defect_type=defect_data['defect_type'],
                defect_category=defect_data.get('defect_category', ''),
                severity=defect_data['severity'],
                description=defect_data['description'],
                related_checklist_items=defect_data.get('related_checklist_items', []),
                photos=defect_data.get('photos', []),
                location_gps=defect_data.get('location_gps', ''),
                qi_notes=defect_data.get('qi_notes', ''),
                qi_signature=qi_signature,
                created_by=request.user,
                correction_due_date=due_date,
                correction_status='OPEN'
            )
            created_defects.append(defect)
            
            # Create history entry
            DefectCorrectionHistory.objects.create(
                defect=defect,
                action='SUBMITTED',
                action_by=request.user,
                notes='Initial defect report created'
            )
        
        # Mark flag as resolved
        flag.status = 'RESOLVED'
        flag.reviewed_at = timezone.now()
        flag.reviewed_by = request.user
        flag.save()
        
        # Send notifications to vendor
        from .email_notification_service import email_service
        for defect in created_defects:
            try:
                email_service.notify_defect_created(defect)
            except Exception as e:
                print(f"Failed to send notification: {e}")
        
        serializer = DefectReportSerializer(created_defects, many=True)
        return Response({
            'created_count': len(created_defects),
            'defects': serializer.data
        }, status=status.HTTP_201_CREATED)
    
    @action(detail=True, methods=['post'], url_path='submit_corrections')
    def submit_corrections(self, request, pk=None):
        """
        Handle vendor correction submissions with file uploads
        """
        try:
            inspection = self.get_object()
            
            # Check if already submitted
            if inspection.correction_status in ['SUBMITTED', 'APPROVED']:
                return Response({
                    'error': 'Corrections already submitted for this inspection',
                    'status': inspection.correction_status
                }, status=status.HTTP_400_BAD_REQUEST)
            
            # Get correction notes
            correction_notes = request.data.get('correction_notes', '')
            
            if not correction_notes:
                return Response(
                    {'error': 'Correction notes are required'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Update inspection with correction data
            inspection.correction_notes = correction_notes
            inspection.correction_completed_at = timezone.now()
            inspection.correction_status = 'SUBMITTED'
            
            # Handle uploaded photos
            uploaded_files = request.FILES.getlist('corrective_photos')
            photo_paths = []
            
            for idx, photo_file in enumerate(uploaded_files):
                # Create photo record
                correction_photo = QIInspectionCorrectionPhoto.objects.create(
                    inspection=inspection,
                    photo_file=photo_file,
                    caption=f"Correction photo {idx + 1}",
                    uploaded_by_id=request.data.get('uploaded_by')
                )
                
                # Store the relative path
                photo_paths.append(correction_photo.photo_file.url)
            
            # Save photo paths to JSON field
            inspection.correction_photos = photo_paths
            inspection.save()
            
            # Send notification email
            try:
                from .email_notification_service import email_service
                email_service.notify_correction_submitted(inspection)
            except Exception as e:
                print(f"Failed to send notification: {e}")
            
            # Return success with photo URLs
            return Response({
                'status': 'success',
                'message': 'Corrections submitted successfully',
                'inspection_id': inspection.inspection_id,
                'correction_status': inspection.correction_status,
                'photo_count': len(photo_paths),
                'photo_urls': photo_paths
            }, status=status.HTTP_200_OK)
            
        except Exception as e:
            return Response({
                'status': 'error',
                'message': str(e)
            }, status=status.HTTP_400_BAD_REQUEST)
    
    @action(detail=True, methods=['post'])
    def approve_correction(self, request, pk=None):
        """
        Supervisor approves correction and updates project status to 7 (Approved)
        """
        defect = self.get_object()
        
        if defect.correction_status != 'SUBMITTED':
            return Response(
                {'error': 'No correction submitted'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        review_notes = request.data.get('review_notes', '')
        reviewed_by = request.data.get('reviewed_by')
        
        if not review_notes.strip():
            return Response(
                {'error': 'Review notes are required'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            # Update defect status
            defect.correction_status = 'APPROVED'
            defect.reviewed_by_id = reviewed_by
            defect.reviewed_at = timezone.now()
            defect.review_notes = review_notes
            defect.save()
            
            # 🔥 UPDATE PROJECT STATUS TO 7 (APPROVED)
            if defect.project:
                # Get the ProjectStatus object with ID 7
                try:
                    approved_status = ProjectStatus.objects.get(status_id=7)
                    old_status_id = defect.project.status.status_id if defect.project.status else None
                    
                    # Update project status using the ForeignKey relationship
                    defect.project.status = approved_status
                    defect.project.save()
                    
                    # Log the status change
                    ChangeLog.objects.create(
                        table_name='projects',
                        record_id=defect.project.project_id,
                        change_type='UPDATE',
                        field_name='status_id',
                        old_value=str(old_status_id) if old_status_id else 'None',
                        new_value='7',
                        changed_by_id=reviewed_by,
                        change_reason=f'Project approved after defect correction approval - Defect #{defect.defect_id}'
                    )
                except ProjectStatus.DoesNotExist:
                    # If status 7 doesn't exist, log error but don't fail the approval
                    print(f"⚠️ Warning: ProjectStatus with ID 7 not found")
                except Exception as status_error:
                    # Log the error but don't fail the approval
                    print(f"⚠️ Error updating project status: {status_error}")
            
            # Create defect correction history
            DefectCorrectionHistory.objects.create(
                defect=defect,
                action='APPROVED',
                action_by_id=reviewed_by,
                notes=review_notes
            )
            
            # Send notification email to vendor
            try:
                from .email_notification_service import email_service
                email_service.notify_correction_approved(defect)
            except Exception as e:
                print(f"Failed to send notification: {e}")
            
            # Prepare response with updated project info
            serializer = DefectReportSerializer(defect)
            
            return Response({
                'status': 'approved',
                'message': 'Correction approved successfully. Project status updated to Approved.',
                'defect': serializer.data,
                'project_status_updated': True,
                'new_project_status': 7
            })
            
        except Exception as e:
            import traceback
            traceback.print_exc()
            return Response(
                {'error': str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )









    @action(detail=True, methods=['post'])
    def reject_correction(self, request, pk=None):
        """
        Supervisor rejects correction
        Auto-escalates to Team Leader after 3 failed attempts
        """
        defect = self.get_object()
        
        if defect.correction_status != 'SUBMITTED':
            return Response(
                {'error': 'No correction submitted'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        review_notes = request.data.get('review_notes', '')
        reviewed_by = request.data.get('reviewed_by')
        
        if not review_notes.strip():
            return Response(
                {'error': 'Review notes are required'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            # Increment failure count
            defect.failure_count += 1
            defect.correction_status = 'REJECTED'
            defect.reviewed_by_id = reviewed_by
            defect.reviewed_at = timezone.now()
            defect.review_notes = review_notes
            
            # Create history
            DefectCorrectionHistory.objects.create(
                defect=defect,
                action='REJECTED',
                action_by_id=reviewed_by,
                notes=review_notes
            )
            
            # Check for auto-escalation (3 failures = escalate)
            escalated = False
            if defect.failure_count >= 3 and not defect.is_escalated:
                defect.is_escalated = True
                defect.escalated_at = timezone.now()
                defect.escalation_reason = f"Automatically escalated after {defect.failure_count} failed correction attempts"
                escalated = True
                
                # Notify management
                try:
                    from .email_notification_service import email_service
                    email_service.notify_defect_escalation(defect)
                except Exception as e:
                    print(f"Failed to send escalation notification: {e}")
            else:
                # Notify vendor of rejection
                try:
                    from .email_notification_service import email_service
                    email_service.notify_correction_rejected(defect)
                except Exception as e:
                    print(f"Failed to send notification: {e}")
            
            defect.save()
            
            serializer = DefectReportSerializer(defect)
            
            return Response({
                'status': 'rejected',
                'message': 'Correction rejected',
                'defect': serializer.data,
                'failure_count': defect.failure_count,
                'escalated': escalated,
                'escalation_message': f'⚠️ Defect automatically escalated to Team Leader after {defect.failure_count} failed attempts' if escalated else None
            })
            
        except Exception as e:
            return Response(
                {'error': str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


    @action(detail=True, methods=['post'])
    def escalate(self, request, pk=None):
        """
        Manually escalate defect to Team Leader
        """
        defect = self.get_object()
        
        escalation_reason = request.data.get('escalation_reason', '')
        escalated_by = request.data.get('escalated_by')
        
        if not escalation_reason.strip():
            return Response(
                {'error': 'Escalation reason is required'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        if defect.is_escalated:
            return Response(
                {'error': 'Defect is already escalated'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            defect.is_escalated = True
            defect.escalated_at = timezone.now()
            defect.escalation_reason = escalation_reason
            defect.save()
            
            # Create history entry
            DefectCorrectionHistory.objects.create(
                defect=defect,
                action='RESUBMITTED',  # Or create a new 'ESCALATED' action type
                action_by_id=escalated_by,
                notes=f"Manually escalated: {escalation_reason}"
            )
            
            # Send notification to management
            try:
                from .email_notification_service import email_service
                email_service.notify_defect_escalation(defect)
            except Exception as e:
                print(f"Failed to send notification: {e}")
            
            serializer = DefectReportSerializer(defect)
            
            return Response({
                'status': 'escalated',
                'message': 'Defect escalated to Team Leader successfully',
                'defect': serializer.data
            })
            
        except Exception as e:
            return Response(
                {'error': str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
        
    @action(detail=False, methods=['get'])
    def dashboard_stats(self, request):
        """Get defect dashboard statistics"""
        
        stats = {
            'total_defects': DefectReport.objects.count(),
            'open': DefectReport.objects.filter(correction_status='OPEN').count(),
            'pending': DefectReport.objects.filter(correction_status='PENDING').count(),
            'submitted': DefectReport.objects.filter(correction_status='SUBMITTED').count(),
            'approved': DefectReport.objects.filter(correction_status='APPROVED').count(),
            'rejected': DefectReport.objects.filter(correction_status='REJECTED').count(),
            'escalated': DefectReport.objects.filter(is_escalated=True).count(),
            'overdue': DefectReport.objects.filter(
                correction_due_date__lt=date.today(),
                correction_status__in=['OPEN', 'PENDING', 'REJECTED']
            ).count(),
            'by_severity': {
                'critical': DefectReport.objects.filter(severity='CRITICAL').count(),
                'major': DefectReport.objects.filter(severity='MAJOR').count(),
                'minor': DefectReport.objects.filter(severity='MINOR').count(),
            }
        }
        
        return Response(stats)
    
    @action(detail=False, methods=['get'])
    def my_defects(self, request):
        """Get defects for current user (vendor)"""
        user = request.user
        
        # Get vendor projects
        from .models import Project
        vendor_projects = Project.objects.filter(vendor__user=user)
        
        defects = DefectReport.objects.filter(
            project__in=vendor_projects
        ).exclude(correction_status='CLOSED')
        
        serializer = self.get_serializer(defects, many=True)
        return Response(serializer.data)
    

from rest_framework import viewsets
from rest_framework.decorators import action
from rest_framework.response import Response
from .models import QIInspectionPhoto
from .serializers import QIInspectionPhotoSerializer

class QIInspectionPhotoViewSet(viewsets.ModelViewSet):
    queryset = QIInspectionPhoto.objects.all()
    serializer_class = QIInspectionPhotoSerializer

    def get_queryset(self):
        queryset = super().get_queryset()
        inspection_id = self.request.query_params.get('inspection_id', None)
        if inspection_id:
            queryset = queryset.filter(inspection_id=inspection_id)
        return queryset.order_by('-uploaded_at')
    
    


class QIInspectionCorrectionPhotoViewSet(viewsets.ModelViewSet):
    queryset = QIInspectionCorrectionPhoto.objects.all()
    serializer_class = QIInspectionCorrectionPhotoSerializer
    permission_classes = [AllowAny]
    parser_classes = [MultiPartParser, FormParser]
    
    def get_queryset(self):
        queryset = super().get_queryset()
        inspection_id = self.request.query_params.get('inspection_id')
        if inspection_id:
            queryset = queryset.filter(inspection_id=inspection_id)
        return queryset.order_by('-uploaded_at')
    
    

# FIXED VERSION OF PaymentReceiptViewSet
# This fixes the "AnonymousUser" error when creating payment receipts

from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import AllowAny
from django.db import transaction, models
from django.utils import timezone
import logging

logger = logging.getLogger(__name__)


class PaymentReceiptViewSet(viewsets.ModelViewSet):
    """
    ViewSet for payment receipt management
    - Vendors can upload receipts
    - Supervisors can approve/reject receipts
    
    FIXED: Properly handles AnonymousUser when using AllowAny permission
    """
    queryset = PaymentReceipt.objects.all().select_related(
        'invoice', 
        'uploaded_by', 
        'reviewed_by'
    ).order_by('-uploaded_at')
    serializer_class = PaymentReceiptSerializer
    permission_classes = [AllowAny]  # Change to IsAuthenticated in production
    
    def get_queryset(self):
        """Filter receipts based on query parameters"""
        queryset = super().get_queryset()
        
        # Filter by invoice
        invoice_id = self.request.query_params.get('invoice', None)
        if invoice_id:
            queryset = queryset.filter(invoice_id=invoice_id)
        
        # Filter by status
        status_param = self.request.query_params.get('status', None)
        if status_param:
            queryset = queryset.filter(status=status_param)
        
        # Filter by vendor (through invoice)
        vendor_id = self.request.query_params.get('vendor', None)
        if vendor_id:
            queryset = queryset.filter(invoice__vendor_id=vendor_id)
        
        return queryset
    
    @transaction.atomic
    def create(self, request, *args, **kwargs):
        """Create payment receipt - FIXED to handle AnonymousUser"""
        try:
            logger.info(f"Creating payment receipt: {request.data}")
            
            serializer = self.get_serializer(data=request.data)
            serializer.is_valid(raise_exception=True)
            
            # FIXED: Only set uploaded_by if user is authenticated
            # This prevents the "AnonymousUser" error when using AllowAny
            if hasattr(request, 'user') and request.user.is_authenticated:
                receipt = serializer.save(uploaded_by=request.user)
                logger.info(f"Receipt uploaded by user: {request.user.username}")
            else:
                # Save without uploaded_by (will be NULL in database)
                receipt = serializer.save()
                logger.warning("Receipt created without authenticated user")
            
            logger.info(f"Payment receipt {receipt.receipt_id} created successfully")
            
            return Response(
                serializer.data,
                status=status.HTTP_201_CREATED
            )
            
        except Exception as e:
            logger.error(f"Error creating payment receipt: {e}")
            return Response(
                {
                    'error': 'Failed to create payment receipt',
                    'detail': str(e)
                },
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    @action(detail=True, methods=['post'])
    def approve(self, request, pk=None):
        """Approve payment receipt and update invoice status"""
        try:
            receipt = self.get_object()
            
            if receipt.status != 'PENDING':
                return Response(
                    {'error': 'Only pending receipts can be approved'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Update receipt status
            receipt.status = 'APPROVED'
            # FIXED: Handle authenticated user check
            if hasattr(request, 'user') and request.user.is_authenticated:
                receipt.reviewed_by = request.user
            receipt.reviewed_at = timezone.now()
            receipt.review_notes = request.data.get('review_notes', '')
            receipt.save()
            
            # Update invoice
            invoice = receipt.invoice
            
            # Calculate total approved payments
            total_approved = PaymentReceipt.objects.filter(
                invoice=invoice,
                status='APPROVED'
            ).aggregate(
                total=models.Sum('payment_amount')
            )['total'] or 0
            
            # Update invoice status
            net_amount = float(invoice.net_amount)
            if total_approved >= net_amount:
                invoice.payment_status = 'Paid'
                invoice.payment_date = receipt.payment_date
                invoice.payment_reference = receipt.receipt_number
            elif total_approved > 0:
                invoice.payment_status = 'Partially Paid'
            
            invoice.save()
            
            logger.info(f"Receipt {receipt.receipt_id} approved. Invoice {invoice.invoice_number} status: {invoice.payment_status}")
            
            serializer = self.get_serializer(receipt)
            return Response({
                'message': 'Receipt approved successfully',
                'receipt': serializer.data,
                'invoice_status': invoice.payment_status
            })
            
        except Exception as e:
            logger.error(f"Error approving receipt: {e}")
            return Response(
                {'error': str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    @action(detail=True, methods=['post'])
    def reject(self, request, pk=None):
        """Reject payment receipt"""
        try:
            receipt = self.get_object()
            
            if receipt.status != 'PENDING':
                return Response(
                    {'error': 'Only pending receipts can be rejected'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            review_notes = request.data.get('review_notes', '')
            if not review_notes:
                return Response(
                    {'error': 'Review notes are required for rejection'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Update receipt status
            receipt.status = 'REJECTED'
            # FIXED: Handle authenticated user check
            if hasattr(request, 'user') and request.user.is_authenticated:
                receipt.reviewed_by = request.user
            receipt.reviewed_at = timezone.now()
            receipt.review_notes = review_notes
            receipt.save()
            
            logger.info(f"Receipt {receipt.receipt_id} rejected")
            
            serializer = self.get_serializer(receipt)
            return Response({
                'message': 'Receipt rejected',
                'receipt': serializer.data
            })
            
        except Exception as e:
            logger.error(f"Error rejecting receipt: {e}")
            return Response(
                {'error': str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    @action(detail=False, methods=['get'])
    def pending_count(self, request):
        """Get count of pending receipts"""
        try:
            count = PaymentReceipt.objects.filter(status='PENDING').count()
            return Response({'pending_count': count})
        except Exception as e:
            logger.error(f"Error getting pending count: {e}")
            return Response(
                {'error': str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    @action(detail=False, methods=['get'])
    def statistics(self, request):
        """Get payment receipt statistics"""
        try:
            from django.db.models import Sum, Count
            
            stats = {
                'total_receipts': PaymentReceipt.objects.count(),
                'pending': PaymentReceipt.objects.filter(status='PENDING').count(),
                'approved': PaymentReceipt.objects.filter(status='APPROVED').count(),
                'rejected': PaymentReceipt.objects.filter(status='REJECTED').count(),
                'total_amount_approved': float(
                    PaymentReceipt.objects.filter(status='APPROVED').aggregate(
                        Sum('payment_amount')
                    )['payment_amount__sum'] or 0
                ),
                'total_amount_pending': float(
                    PaymentReceipt.objects.filter(status='PENDING').aggregate(
                        Sum('payment_amount')
                    )['payment_amount__sum'] or 0
                ),
            }
            
            return Response(stats)
            
        except Exception as e:
            logger.error(f"Error getting statistics: {e}")
            return Response(
                {'error': str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


# ============================================
# ALTERNATIVE: Use authentication instead
# ============================================


class PaymentReceiptViewSetAuthenticated(viewsets.ModelViewSet):
    """
    ALTERNATIVE VERSION: Requires authentication
    This is the recommended approach for production
    """
    queryset = PaymentReceipt.objects.all().select_related(
        'invoice', 
        'uploaded_by', 
        'reviewed_by'
    ).order_by('-uploaded_at')
    serializer_class = PaymentReceiptSerializer
    permission_classes = [IsAuthenticated]  # Require authentication
    
    def get_queryset(self):
        """Filter receipts based on query parameters"""
        queryset = super().get_queryset()
        
        # Filter by invoice
        invoice_id = self.request.query_params.get('invoice', None)
        if invoice_id:
            queryset = queryset.filter(invoice_id=invoice_id)
        
        # Filter by status
        status_param = self.request.query_params.get('status', None)
        if status_param:
            queryset = queryset.filter(status=status_param)
        
        # Filter by vendor (through invoice)
        vendor_id = self.request.query_params.get('vendor', None)
        if vendor_id:
            queryset = queryset.filter(invoice__vendor_id=vendor_id)
        
        return queryset
    
    @transaction.atomic
    def create(self, request, *args, **kwargs):
        """Create payment receipt - user must be authenticated"""
        try:
            logger.info(f"Creating payment receipt by user: {request.user.username}")
            
            serializer = self.get_serializer(data=request.data)
            serializer.is_valid(raise_exception=True)
            
            # Save with authenticated user (no need to check, IsAuthenticated ensures it)
            receipt = serializer.save(uploaded_by=request.user)
            
            logger.info(f"Payment receipt {receipt.receipt_id} created successfully")
            
            return Response(
                serializer.data,
                status=status.HTTP_201_CREATED
            )
            
        except Exception as e:
            logger.error(f"Error creating payment receipt: {e}")
            return Response(
                {
                    'error': 'Failed to create payment receipt',
                    'detail': str(e)
                },
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    
    @action(detail=True, methods=['post'])
    def approve(self, request, pk=None):
        """Approve payment receipt and update invoice status"""
        try:
            receipt = self.get_object()
            
            if receipt.status != 'PENDING':
                return Response(
                    {'error': 'Only pending receipts can be approved'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Update receipt status
            receipt.status = 'APPROVED'
            receipt.reviewed_by = request.user if hasattr(request, 'user') else None
            receipt.reviewed_at = timezone.now()
            receipt.review_notes = request.data.get('review_notes', '')
            receipt.save()
            
            # Update invoice
            invoice = receipt.invoice
            
            # Calculate total approved payments
            total_approved = PaymentReceipt.objects.filter(
                invoice=invoice,
                status='APPROVED'
            ).aggregate(
                total=models.Sum('payment_amount')
            )['total'] or 0
            
            # Update invoice status
            net_amount = float(invoice.net_amount)
            if total_approved >= net_amount:
                invoice.payment_status = 'Paid'
                invoice.payment_date = receipt.payment_date
                invoice.payment_reference = receipt.receipt_number
            elif total_approved > 0:
                invoice.payment_status = 'Partially Paid'
            
            invoice.save()
            
            logger.info(f"Receipt {receipt.receipt_id} approved. Invoice {invoice.invoice_number} status: {invoice.payment_status}")
            
            serializer = self.get_serializer(receipt)
            return Response({
                'message': 'Receipt approved successfully',
                'receipt': serializer.data,
                'invoice_status': invoice.payment_status
            })
            
        except Exception as e:
            logger.error(f"Error approving receipt: {e}")
            return Response(
                {'error': str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    @action(detail=True, methods=['post'])
    def reject(self, request, pk=None):
        """Reject payment receipt"""
        try:
            receipt = self.get_object()
            
            if receipt.status != 'PENDING':
                return Response(
                    {'error': 'Only pending receipts can be rejected'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            review_notes = request.data.get('review_notes', '')
            if not review_notes:
                return Response(
                    {'error': 'Review notes are required for rejection'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Update receipt status
            receipt.status = 'REJECTED'
            receipt.reviewed_by = request.user if hasattr(request, 'user') else None
            receipt.reviewed_at = timezone.now()
            receipt.review_notes = review_notes
            receipt.save()
            
            logger.info(f"Receipt {receipt.receipt_id} rejected")
            
            serializer = self.get_serializer(receipt)
            return Response({
                'message': 'Receipt rejected',
                'receipt': serializer.data
            })
            
        except Exception as e:
            logger.error(f"Error rejecting receipt: {e}")
            return Response(
                {'error': str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    @action(detail=False, methods=['get'])
    def pending_count(self, request):
        """Get count of pending receipts"""
        try:
            count = PaymentReceipt.objects.filter(status='PENDING').count()
            return Response({'pending_count': count})
        except Exception as e:
            logger.error(f"Error getting pending count: {e}")
            return Response(
                {'error': str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    @action(detail=False, methods=['get'])
    def statistics(self, request):
        """Get payment receipt statistics"""
        try:
            from django.db.models import Sum, Count
            
            stats = {
                'total_receipts': PaymentReceipt.objects.count(),
                'pending': PaymentReceipt.objects.filter(status='PENDING').count(),
                'approved': PaymentReceipt.objects.filter(status='APPROVED').count(),
                'rejected': PaymentReceipt.objects.filter(status='REJECTED').count(),
                'total_amount_approved': float(
                    PaymentReceipt.objects.filter(status='APPROVED').aggregate(
                        Sum('payment_amount')
                    )['payment_amount__sum'] or 0
                ),
                'total_amount_pending': float(
                    PaymentReceipt.objects.filter(status='PENDING').aggregate(
                        Sum('payment_amount')
                    )['payment_amount__sum'] or 0
                ),
            }
            
            return Response(stats)
            
        except Exception as e:
            logger.error(f"Error getting statistics: {e}")
            return Response(
                {'error': str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )