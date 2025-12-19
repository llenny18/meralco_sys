# file_handlers.py - Create this new file

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from django.http import HttpResponse, JsonResponse
from django.views.decorators.http import require_http_methods
from django.views.decorators.csrf import csrf_exempt
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework import status
from datetime import datetime
from .models import *
from .serializers import *
import json
from io import BytesIO

# ============================================
# WORK ORDER EXCEL IMPORT/EXPORT
# ============================================

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def import_work_orders_excel(request):
    """
    Import work orders from Excel file (C1 sheet format)
    Expected columns match your Excel documentation
    """
    if 'file' not in request.FILES:
        return Response(
            {'error': 'No file provided'}, 
            status=status.HTTP_400_BAD_REQUEST
        )
    
    excel_file = request.FILES['file']
    
    try:
        # Read Excel file
        df = pd.read_excel(excel_file, sheet_name=0)
        
        # Column mapping (adjust based on your actual Excel columns)
        column_mapping = {
            'WO NO': 'wo_no',
            'DESCRIPTION': 'description',
            'LOCATION': 'location',
            'MUNICIPALITY': 'municipality',
            'AREA OF RESPONSIBILITY': 'area_of_responsibility',
            'Date Received Jacket from P&S': 'date_received_jacket',
            'Date Received Awarding of Wo': 'date_received_awarding',
            'Date Energized': 'date_energized',
            'VENDOR REMARKS (free text)': 'vendor_remarks',
            'C1 REMARKS (with or without issue)': 'c1_remarks',
            'ASSIGNED': 'assigned_crew',
            'SUPERVISOR': 'supervisor_code',
            'TOTAL MANHOURS': 'total_manhours',
            'TOTAL ESTIMATED COST (SERVICE ITEMS)': 'total_estimated_cost',
            'VIP': 'is_vip',
            'WO INITIATOR': 'wo_initiator',
            'EAM STATUS': 'eam_status',
        }
        
        created_count = 0
        updated_count = 0
        errors = []
        
        for index, row in df.iterrows():
            try:
                wo_no = row.get('WO NO') or row.get('wo_no')
                if pd.isna(wo_no) or not wo_no:
                    continue
                
                # Parse dates
                date_fields = ['date_received_jacket', 'date_received_awarding', 'date_energized']
                wo_data = {}
                
                for excel_col, model_field in column_mapping.items():
                    if excel_col in df.columns:
                        value = row[excel_col]
                        
                        # Handle dates
                        if model_field in date_fields and pd.notna(value):
                            if isinstance(value, pd.Timestamp):
                                wo_data[model_field] = value.date()
                            elif isinstance(value, str):
                                try:
                                    wo_data[model_field] = datetime.strptime(value, '%Y-%m-%d').date()
                                except:
                                    wo_data[model_field] = None
                        
                        # Handle boolean
                        elif model_field == 'is_vip':
                            wo_data[model_field] = str(value).upper() == 'Y' if pd.notna(value) else False
                        
                        # Handle numeric
                        elif model_field in ['total_manhours', 'total_estimated_cost']:
                            wo_data[model_field] = float(value) if pd.notna(value) else None
                        
                        # Handle text
                        else:
                            wo_data[model_field] = str(value) if pd.notna(value) else None
                
                # Get or create work order
                work_order, created = WorkOrder.objects.update_or_create(
                    wo_no=wo_no,
                    defaults=wo_data
                )
                
                if created:
                    created_count += 1
                else:
                    updated_count += 1
                    
            except Exception as e:
                errors.append(f"Row {index + 2}: {str(e)}")
        
        return Response({
            'message': 'Import completed',
            'created': created_count,
            'updated': updated_count,
            'errors': errors
        }, status=status.HTTP_200_OK)
        
    except Exception as e:
        return Response(
            {'error': f'Failed to process file: {str(e)}'}, 
            status=status.HTTP_400_BAD_REQUEST
        )


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def export_work_orders_excel(request):
    """
    Export work orders to Excel file (C1 sheet format)
    """
    # Get filter parameters
    status_filter = request.GET.get('status')
    vendor_filter = request.GET.get('vendor')
    crew_filter = request.GET.get('crew')
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    
    # Build queryset
    queryset = WorkOrder.objects.all()
    
    if status_filter:
        queryset = queryset.filter(status=status_filter)
    if vendor_filter:
        queryset = queryset.filter(vendor_id=vendor_filter)
    if crew_filter:
        queryset = queryset.filter(assigned_crew=crew_filter)
    if date_from:
        queryset = queryset.filter(date_energized__gte=date_from)
    if date_to:
        queryset = queryset.filter(date_energized__lte=date_to)
    
    # Create Excel workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "C1"
    
    # Define headers (matching your Excel structure)
    headers = [
        'Date Received Jacket from P&S',
        'Date Received Awarding of Wo',
        'VIP',
        'WO INITIATOR',
        'WO NO',
        'DESCRIPTION',
        'LOCATION',
        'MUNICIPALITY',
        'AREA OF RESPONSIBILITY',
        'VENDOR REMARKS',
        'C1 REMARKS',
        'ASSIGNED',
        'SUPERVISOR',
        'TOTAL MANHOURS',
        'TOTAL ESTIMATED COST',
        'BILLED COST',
        'Date Energized',
        'Date COC Received',
        'Days Energized to COC',
        'Date For Audit',
        'Days COC to Audit',
        'Date Audited',
        'Days Audit to Billing',
        'Total Resolution Days',
        'Is Delayed',
        'Delay Days',
        'STATUS',
        'EAM STATUS',
    ]
    
    # Style for headers
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    
    # Write headers
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Write data
    row_num = 2
    for wo in queryset:
        ws.cell(row=row_num, column=1).value = wo.date_received_jacket
        ws.cell(row=row_num, column=2).value = wo.date_received_awarding
        ws.cell(row=row_num, column=3).value = 'Y' if wo.is_vip else ''
        ws.cell(row=row_num, column=4).value = wo.wo_initiator
        ws.cell(row=row_num, column=5).value = wo.wo_no
        ws.cell(row=row_num, column=6).value = wo.description
        ws.cell(row=row_num, column=7).value = wo.location
        ws.cell(row=row_num, column=8).value = wo.municipality
        ws.cell(row=row_num, column=9).value = wo.area_of_responsibility
        ws.cell(row=row_num, column=10).value = wo.vendor_remarks
        ws.cell(row=row_num, column=11).value = wo.c1_remarks
        ws.cell(row=row_num, column=12).value = wo.assigned_crew
        ws.cell(row=row_num, column=13).value = wo.supervisor.get_full_name() if wo.supervisor else ''
        ws.cell(row=row_num, column=14).value = float(wo.total_manhours) if wo.total_manhours else 0
        ws.cell(row=row_num, column=15).value = float(wo.total_estimated_cost) if wo.total_estimated_cost else 0
        ws.cell(row=row_num, column=16).value = float(wo.billed_cost) if wo.billed_cost else 0
        ws.cell(row=row_num, column=17).value = wo.date_energized
        ws.cell(row=row_num, column=18).value = wo.date_coc_received
        ws.cell(row=row_num, column=19).value = wo.days_from_energized_to_coc
        ws.cell(row=row_num, column=20).value = wo.date_for_audit
        ws.cell(row=row_num, column=21).value = wo.days_from_coc_to_audit
        ws.cell(row=row_num, column=22).value = wo.date_audited
        ws.cell(row=row_num, column=23).value = wo.days_from_audit_to_billing
        ws.cell(row=row_num, column=24).value = wo.total_resolution_days
        ws.cell(row=row_num, column=25).value = 'Yes' if wo.is_delayed else 'No'
        ws.cell(row=row_num, column=26).value = wo.delay_days
        ws.cell(row=row_num, column=27).value = wo.status
        ws.cell(row=row_num, column=28).value = wo.eam_status
        
        row_num += 1
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column[0].column_letter].width = adjusted_width
    
    # Freeze first row
    ws.freeze_panes = 'A2'
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Create response
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=work_orders_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    
    return response


# ============================================
# CREW MONITORING EXCEL IMPORT
# ============================================

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def import_crew_monitoring_excel(request):
    """
    Import crew monitoring data from Excel
    Supports monthly sheets (MAY 2025, JUNE 2025, etc.)
    """
    if 'file' not in request.FILES:
        return Response({'error': 'No file provided'}, status=status.HTTP_400_BAD_REQUEST)
    
    excel_file = request.FILES['file']
    sheet_name = request.POST.get('sheet_name', 0)  # Default to first sheet
    
    try:
        df = pd.read_excel(excel_file, sheet_name=sheet_name)
        
        created_count = 0
        errors = []
        
        # Expected structure from your documentation
        # Row with crew codes, then daily data
        
        for index, row in df.iterrows():
            try:
                crew_code = row.get('CREW_CODE')
                monitoring_date = row.get('DATE')
                
                if pd.isna(crew_code) or pd.isna(monitoring_date):
                    continue
                
                # Get crew type
                crew_type = CrewType.objects.get(crew_code=crew_code)
                
                # Parse date
                if isinstance(monitoring_date, pd.Timestamp):
                    date_obj = monitoring_date.date()
                else:
                    date_obj = datetime.strptime(str(monitoring_date), '%Y-%m-%d').date()
                
                # Get values A, B, C, D
                value_a = int(row.get('VALUE_A', 0)) if pd.notna(row.get('VALUE_A')) else 0
                value_b = int(row.get('VALUE_B', 0)) if pd.notna(row.get('VALUE_B')) else 0
                value_c = int(row.get('VALUE_C', 0)) if pd.notna(row.get('VALUE_C')) else 0
                value_d = int(row.get('VALUE_D', 0)) if pd.notna(row.get('VALUE_D')) else 0
                daily_rate = float(row.get('RATE', 0)) if pd.notna(row.get('RATE')) else None
                
                # Create or update record
                monitoring, created = DailyCrewMonitoring.objects.update_or_create(
                    crew_type=crew_type,
                    monitoring_date=date_obj,
                    defaults={
                        'value_a': value_a,
                        'value_b': value_b,
                        'value_c': value_c,
                        'value_d': value_d,
                        'daily_rate': daily_rate,
                    }
                )
                
                if created:
                    created_count += 1
                    
            except CrewType.DoesNotExist:
                errors.append(f"Row {index + 2}: Crew type '{crew_code}' not found")
            except Exception as e:
                errors.append(f"Row {index + 2}: {str(e)}")
        
        return Response({
            'message': 'Import completed',
            'created': created_count,
            'errors': errors
        }, status=status.HTTP_200_OK)
        
    except Exception as e:
        return Response(
            {'error': f'Failed to process file: {str(e)}'}, 
            status=status.HTTP_400_BAD_REQUEST
        )


# ============================================
# QI MONITORING EXCEL IMPORT
# ============================================

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def import_qi_monitoring_excel(request):
    """
    Import QI monitoring data from Excel
    """
    if 'file' not in request.FILES:
        return Response({'error': 'No file provided'}, status=status.HTTP_400_BAD_REQUEST)
    
    excel_file = request.FILES['file']
    
    try:
        # Read weekly accomplishment sheet
        df = pd.read_excel(excel_file, sheet_name='WEEKLY ACCOMPLISHMENT')
        
        created_count = 0
        errors = []
        
        for index, row in df.iterrows():
            try:
                qi_name = row.get('QI_NAME')
                week_start = row.get('WEEK_START_DATE')
                
                if pd.isna(qi_name) or pd.isna(week_start):
                    continue
                
                # Get user by name
                qi_user = User.objects.get(
                    first_name__icontains=qi_name.split()[0],
                    last_name__icontains=qi_name.split()[-1]
                )
                
                # Parse date
                if isinstance(week_start, pd.Timestamp):
                    week_start_date = week_start.date()
                else:
                    week_start_date = datetime.strptime(str(week_start), '%Y-%m-%d').date()
                
                # Calculate week end
                week_end_date = week_start_date + timedelta(days=6)
                
                # Get daily counts
                monday_count = int(row.get('MONDAY', 0)) if pd.notna(row.get('MONDAY')) else 0
                tuesday_count = int(row.get('TUESDAY', 0)) if pd.notna(row.get('TUESDAY')) else 0
                wednesday_count = int(row.get('WEDNESDAY', 0)) if pd.notna(row.get('WEDNESDAY')) else 0
                thursday_count = int(row.get('THURSDAY', 0)) if pd.notna(row.get('THURSDAY')) else 0
                friday_count = int(row.get('FRIDAY', 0)) if pd.notna(row.get('FRIDAY')) else 0
                saturday_count = int(row.get('SATURDAY', 0)) if pd.notna(row.get('SATURDAY')) else 0
                sunday_count = int(row.get('SUNDAY', 0)) if pd.notna(row.get('SUNDAY')) else 0
                target = int(row.get('TARGET', 0)) if pd.notna(row.get('TARGET')) else 0
                
                # Create or update
                accomplishment, created = QIWeeklyAccomplishment.objects.update_or_create(
                    qi_user=qi_user,
                    week_start_date=week_start_date,
                    defaults={
                        'week_end_date': week_end_date,
                        'monday_count': monday_count,
                        'tuesday_count': tuesday_count,
                        'wednesday_count': wednesday_count,
                        'thursday_count': thursday_count,
                        'friday_count': friday_count,
                        'saturday_count': saturday_count,
                        'sunday_count': sunday_count,
                        'target_inspections': target,
                    }
                )
                
                if created:
                    created_count += 1
                    
            except User.DoesNotExist:
                errors.append(f"Row {index + 2}: User '{qi_name}' not found")
            except Exception as e:
                errors.append(f"Row {index + 2}: {str(e)}")
        
        return Response({
            'message': 'Import completed',
            'created': created_count,
            'errors': errors
        }, status=status.HTTP_200_OK)
        
    except Exception as e:
        return Response(
            {'error': f'Failed to process file: {str(e)}'}, 
            status=status.HTTP_400_BAD_REQUEST
        )


# ============================================
# BULK OPERATIONS
# ============================================

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def bulk_update_work_orders(request):
    """
    Bulk update work orders
    Payload: {
        "wo_ids": [list of work order IDs],
        "updates": {field: value}
    }
    """
    wo_ids = request.data.get('wo_ids', [])
    updates = request.data.get('updates', {})
    
    if not wo_ids or not updates:
        return Response(
            {'error': 'wo_ids and updates required'}, 
            status=status.HTTP_400_BAD_REQUEST
        )
    
    try:
        updated = WorkOrder.objects.filter(wo_id__in=wo_ids).update(**updates)
        
        return Response({
            'message': f'Successfully updated {updated} work orders',
            'updated_count': updated
        }, status=status.HTTP_200_OK)
        
    except Exception as e:
        return Response(
            {'error': str(e)}, 
            status=status.HTTP_400_BAD_REQUEST
        )


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def bulk_upload_documents(request):
    """
    Bulk upload documents for multiple work orders
    """
    files = request.FILES.getlist('files')
    wo_id = request.data.get('work_order_id')
    doc_type = request.data.get('document_type', 'OTHER')
    
    if not files or not wo_id:
        return Response(
            {'error': 'files and work_order_id required'}, 
            status=status.HTTP_400_BAD_REQUEST
        )
    
    try:
        work_order = WorkOrder.objects.get(wo_id=wo_id)
        uploaded_docs = []
        
        for file in files:
            doc = WorkOrderDocument.objects.create(
                work_order=work_order,
                document_type=doc_type,
                document_name=file.name,
                document_file=file,
                uploaded_by=request.user
            )
            uploaded_docs.append(WorkOrderDocumentSerializer(doc).data)
        
        return Response({
            'message': f'Successfully uploaded {len(uploaded_docs)} documents',
            'documents': uploaded_docs
        }, status=status.HTTP_201_CREATED)
        
    except WorkOrder.DoesNotExist:
        return Response(
            {'error': 'Work order not found'}, 
            status=status.HTTP_404_NOT_FOUND
        )
    except Exception as e:
        return Response(
            {'error': str(e)}, 
            status=status.HTTP_400_BAD_REQUEST
        )